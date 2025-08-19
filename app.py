#!/usr/bin/env python3
"""
SISAGENT - Sistema de Gestión de Operaciones Bancarias
"""

import os
import sys
from datetime import datetime, timedelta
from decimal import Decimal
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import pytz

# Configuración de zona horaria (UTC-5 para Perú)
peru_tz = pytz.timezone('America/Lima')

def get_peru_time():
    """Obtiene la hora actual en zona horaria de Perú"""
    return datetime.now(peru_tz)

def format_peru_time(dt):
    """Formatea una fecha/hora para mostrar en zona horaria de Perú"""
    if dt is None:
        return ""
    # Si la fecha ya tiene zona horaria, convertirla a Perú
    if dt.tzinfo is not None:
        return dt.astimezone(peru_tz).strftime('%H:%M:%S')
    # Si no tiene zona horaria, asumir que es UTC y convertir
    else:
        return dt.replace(tzinfo=pytz.UTC).astimezone(peru_tz).strftime('%H:%M:%S')

def format_peru_date(dt):
    """Formatea una fecha para mostrar en zona horaria de Perú"""
    if dt is None:
        return ""
    # Si la fecha ya tiene zona horaria, convertirla a Perú
    if dt.tzinfo is not None:
        return dt.astimezone(peru_tz).strftime('%d/%m/%Y')
    # Si no tiene zona horaria, asumir que es UTC y convertir
    else:
        return dt.replace(tzinfo=pytz.UTC).astimezone(peru_tz).strftime('%d/%m/%Y')

def format_peru_datetime(dt):
    """Formatea una fecha/hora completa para mostrar en zona horaria de Perú"""
    if dt is None:
        return ""
    # Si la fecha ya tiene zona horaria, convertirla a Perú
    if dt.tzinfo is not None:
        return dt.astimezone(peru_tz).strftime('%d/%m/%Y %H:%M:%S')
    # Si no tiene zona horaria, asumir que es UTC y convertir
    else:
        return dt.replace(tzinfo=pytz.UTC).astimezone(peru_tz).strftime('%d/%m/%Y %H:%M:%S')

def format_peru_datetime_short(dt):
    """Formatea una fecha/hora corta para mostrar en zona horaria de Perú"""
    if dt is None:
        return ""
    # Si la fecha ya tiene zona horaria, convertirla a Perú
    if dt.tzinfo is not None:
        return dt.astimezone(peru_tz).strftime('%d/%m/%Y %H:%M')
    # Si no tiene zona horaria, asumir que es UTC y convertir
    else:
        return dt.replace(tzinfo=pytz.UTC).astimezone(peru_tz).strftime('%d/%m/%Y %H:%M')

print("🚀 SISAGENT Flask arrancando...")
print("🔄 Actualización Railway - " + get_peru_time().strftime("%Y-%m-%d %H:%M:%S"))
print("🔧 FIX: Eliminación de sucursales mejorada con mejor manejo de errores")
print("🗑️ REMOVED: Botón de convertir en administrador eliminado de acciones rápidas")
print("✏️ FIX: Permitir editar nombre de usuario en gestión de usuarios")
print("🔧 FIX: Corregir edición de operaciones - problema con columnas de usuario")
print("🕐 FIX: Corregir zona horaria - usar hora de Perú (UTC-5) en lugar de UTC")
print("🔧 FIX: Corregir error de orden - función get_peru_time definida antes de usar")
print("🔧 FIX: Usar lambda functions para zona horaria en modelos - evitar errores de función")
print("🔧 FIX: Corregir descuadre en edición de operaciones - mostrar columnas solo para admin")
print("🕐 FIX: Corregir visualización de hora - usar función format_peru_time en templates")
print("🕐 FIX: Corregir TODAS las horas en reportes PDF/XLSX/CSV y templates de usuarios/sucursales")

# Configuración de la aplicación Flask
app = Flask(__name__)

print("✅ Flask app creada")

# Configuración para Railway
if os.environ.get('DATABASE_URL'):
    # Para Railway con PostgreSQL
    database_url = os.environ.get('DATABASE_URL')
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    print(f"✅ Usando PostgreSQL en Railway: {database_url[:20]}...")
else:
    # Para desarrollo local y Railway sin DATABASE_URL
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sisagent_consolidada.db'
    print("✅ Usando SQLite para desarrollo local")

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'tu-clave-secreta-aqui')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# OPTIMIZACIÓN: Configuraciones para mejorar rendimiento
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
    'pool_size': 10,
    'max_overflow': 20
}

print("✅ Configuración de base de datos completada")

# Configurar CORS para permitir peticiones desde cualquier origen en producción
from flask_cors import CORS
CORS(app, origins=['*'])

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

print("✅ SQLAlchemy y LoginManager configurados")

print("✅ Configuración de zona horaria completada")

# Agregar funciones de formato de hora al contexto de Jinja2
app.jinja_env.globals['format_peru_time'] = format_peru_time
app.jinja_env.globals['format_peru_date'] = format_peru_date
app.jinja_env.globals['format_peru_datetime'] = format_peru_datetime
app.jinja_env.globals['format_peru_datetime_short'] = format_peru_datetime_short

# Medios de pago se obtienen dinámicamente de la base de datos

# Modelos de base de datos
class Sucursal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    direccion = db.Column(db.String(200))
    activa = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(peru_tz))
    usuarios = db.relationship('Usuario', backref='sucursal', lazy=True)
    operaciones = db.relationship('Operacion', backref='sucursal', lazy=True)

class Usuario(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    nombre_completo = db.Column(db.String(100), nullable=False)
    es_admin = db.Column(db.Boolean, default=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=True)
    activo = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(peru_tz))
    operaciones = db.relationship('Operacion', backref='usuario', lazy=True)
    
    
class Operacion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    monto = db.Column(db.Numeric(10, 2), nullable=False)
    comision = db.Column(db.Numeric(10, 2), nullable=False)
    medio = db.Column(db.String(20), nullable=False)
    hora = db.Column(db.DateTime, nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(peru_tz))

class ComisionDiaria(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False)
    total_comision = db.Column(db.Numeric(10, 2), default=0)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(peru_tz))

class ComisionMensual(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    año = db.Column(db.Integer, nullable=False)
    mes = db.Column(db.Integer, nullable=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False)
    total_comision = db.Column(db.Numeric(10, 2), default=0)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(peru_tz))

class MedioPago(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre_abreviado = db.Column(db.String(20), unique=True, nullable=False)
    nombre_completo = db.Column(db.String(100), nullable=False)
    activo = db.Column(db.Boolean, default=True)
    orden = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(peru_tz))

class MedioSucursal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False)
    medio_pago_id = db.Column(db.Integer, db.ForeignKey('medio_pago.id'), nullable=False)
    activo = db.Column(db.Boolean, default=True)
    sucursal = db.relationship('Sucursal', backref='medios_sucursal')
    medio_pago = db.relationship('MedioPago', backref='sucursales_medio')

# Modelos para el módulo de TAREO
class Tareo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(200), nullable=False)
    descripcion = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False)
    estado = db.Column(db.String(20), default='pendiente')  # pendiente, en_progreso, completado
    fecha_creacion = db.Column(db.DateTime, default=lambda: datetime.now(peru_tz))
    fecha_completado = db.Column(db.DateTime, nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    
    # Relaciones
    usuario = db.relationship('Usuario', foreign_keys=[usuario_id], backref='tareos_asignados')
    sucursal = db.relationship('Sucursal', backref='tareos')
    creador = db.relationship('Usuario', foreign_keys=[created_by], backref='tareos_creados')
    operaciones = db.relationship('OperacionTareo', backref='tareo', lazy=True, cascade='all, delete-orphan')

class OperacionTareo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tareo_id = db.Column(db.Integer, db.ForeignKey('tareo.id'), nullable=False)
    medio = db.Column(db.String(20), nullable=False)
    destino = db.Column(db.String(100), nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    monto = db.Column(db.Numeric(10, 2), nullable=False)
    completado = db.Column(db.Boolean, default=False)
    fecha_completado = db.Column(db.DateTime, nullable=True)
    orden = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(peru_tz))

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# Health check extremadamente simple para Railway
@app.route('/')
def root_health_check():
    # Si el usuario está autenticado, redirigir a la app
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    # Si no está autenticado, mostrar página de login
    return redirect(url_for('login'))

# Health check específico para Railway (sin autenticación)
@app.route('/railway-health')
def railway_health():
    return "OK", 200

# Health check simple para Railway (ruta raíz sin autenticación)
@app.route('/health')
def health_check():
    return "OK", 200

# Health check simple para Railway
@app.route('/api/health')
def api_health_check():
    return "OK", 200

# Rutas de autenticación
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = Usuario.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password) and user.activo:
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Credenciales inválidas o usuario inactivo', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# Rutas principales
@app.route('/app')
@login_required
def dashboard():
    if current_user.es_admin:
        return redirect(url_for('admin_dashboard'))
    else:
        return redirect(url_for('user_dashboard'))

@app.route('/admin')
@login_required
def admin_dashboard():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))

    # Obtener la hora actual en Perú y convertirla a UTC naive para el filtro
    ahora_peru = datetime.now(peru_tz)
    hoy_peru = ahora_peru.date()
    mes_actual = hoy_peru.month
    año_actual = hoy_peru.year

    # Calcular el inicio y fin del día en Perú, y luego convertirlos a UTC naive
    inicio_dia_peru = datetime.combine(hoy_peru, datetime.min.time()).replace(tzinfo=peru_tz)
    fin_dia_peru = datetime.combine(hoy_peru, datetime.max.time()).replace(tzinfo=peru_tz)
    
    inicio_dia_utc_naive = inicio_dia_peru.astimezone(pytz.utc).replace(tzinfo=None)
    fin_dia_utc_naive = fin_dia_peru.astimezone(pytz.utc).replace(tzinfo=None)

    # Calcular el inicio y fin del mes en Perú, y luego convertirlos a UTC naive
    inicio_mes_peru = datetime.combine(hoy_peru.replace(day=1), datetime.min.time()).replace(tzinfo=peru_tz)
    if mes_actual == 12:
        fin_mes_peru = datetime.combine(hoy_peru.replace(year=año_actual + 1, month=1, day=1) - timedelta(days=1), datetime.max.time()).replace(tzinfo=peru_tz)
    else:
        fin_mes_peru = datetime.combine(hoy_peru.replace(month=mes_actual + 1, day=1) - timedelta(days=1), datetime.max.time()).replace(tzinfo=peru_tz)
    
    inicio_mes_utc_naive = inicio_mes_peru.astimezone(pytz.utc).replace(tzinfo=None)
    fin_mes_utc_naive = fin_mes_peru.astimezone(pytz.utc).replace(tzinfo=None)

    # Optimizar consultas con una sola query por sucursal
    sucursales = Sucursal.query.filter_by(activa=True).all()
    comisiones_hoy = []
    comisiones_mes = {}
    
    # OPTIMIZACIÓN: Usar la misma lógica que los reportes para consistencia
    comisiones_diarias = db.session.query(
        Operacion.sucursal_id,
        db.func.sum(Operacion.comision).label('total')
    ).filter(
        Operacion.hora >= inicio_dia_peru,
        Operacion.hora <= fin_dia_peru
    ).group_by(Operacion.sucursal_id).all()
    
    comisiones_mensuales = db.session.query(
        Operacion.sucursal_id,
        db.func.sum(Operacion.comision).label('total')
    ).filter(
        Operacion.hora >= inicio_mes_peru,
        Operacion.hora <= fin_mes_peru
    ).group_by(Operacion.sucursal_id).all()
    
    # Crear diccionarios para acceso rápido
    comisiones_diarias_dict = {cd.sucursal_id: float(cd.total) for cd in comisiones_diarias}
    comisiones_mensuales_dict = {cm.sucursal_id: float(cm.total) for cm in comisiones_mensuales}
    
    # Debug: Solo mostrar información esencial en desarrollo
    if app.debug:
        print(f"DEBUG ADMIN: Procesando {len(sucursales)} sucursales")
    
    for suc in sucursales:
        total_hoy = comisiones_diarias_dict.get(suc.id, 0.0)
        total_mes = comisiones_mensuales_dict.get(suc.id, 0.0)
        comisiones_hoy.append((suc.id, suc.nombre, total_hoy))
        comisiones_mes[suc.id] = total_mes
    
    total_sucursales = len(sucursales)
    total_usuarios = Usuario.query.filter_by(activo=True).count()
    return render_template('admin_dashboard.html',
        total_sucursales=total_sucursales,
        total_usuarios=total_usuarios,
        comisiones_hoy=comisiones_hoy,
        comisiones_mes=comisiones_mes
    )

@app.route('/user')
@login_required
def user_dashboard():
    if current_user.es_admin:
        return redirect(url_for('admin_dashboard'))
    
    # Usar la misma lógica de timezone que el registro de operaciones
    peru_tz = pytz.timezone('America/Lima')
    ahora = datetime.now(peru_tz)
    hoy = ahora.date()
    
    # Calcular la comisión diaria usando la misma lógica que los reportes
    inicio_dia = datetime.combine(hoy, datetime.min.time()).replace(tzinfo=peru_tz)
    fin_dia = datetime.combine(hoy, datetime.max.time()).replace(tzinfo=peru_tz)
    
    total_comision_hoy = db.session.query(db.func.coalesce(db.func.sum(Operacion.comision), 0)).filter(
        Operacion.usuario_id == current_user.id,
        Operacion.hora >= inicio_dia,
        Operacion.hora <= fin_dia
    ).scalar() or 0
    
    # OPTIMIZACIÓN: Limitar operaciones mostradas a las últimas 10 para mejorar rendimiento
    operaciones_hoy = Operacion.query.filter_by(
        usuario_id == current_user.id
    ).filter(
        Operacion.hora >= inicio_dia,
        Operacion.hora <= fin_dia
    ).order_by(Operacion.hora.desc()).limit(10).all()
    
    # Debug: Solo mostrar información esencial en desarrollo
    if app.debug:
        print(f"DEBUG: Operaciones encontradas: {len(operaciones_hoy)}")
    
    return render_template('user_dashboard.html',
                         total_comision_hoy=total_comision_hoy,
                         operaciones_hoy=operaciones_hoy)

# Gestión de sucursales (solo admin)
@app.route('/admin/sucursales')
@login_required
def admin_sucursales():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    
    sucursales = Sucursal.query.all()
    return render_template('admin_sucursales.html', sucursales=sucursales)

@app.route('/admin/sucursales/crear', methods=['GET', 'POST'])
@login_required
def crear_sucursal():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    medios = MedioPago.query.order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    if request.method == 'POST':
        nombre = request.form['nombre']
        direccion = request.form['direccion']
        sucursal = Sucursal(nombre=nombre, direccion=direccion)
        db.session.add(sucursal)
        db.session.commit()
        # Asociar medios seleccionados
        medios_ids = request.form.getlist('medios')
        for medio_id in medios_ids:
            ms = MedioSucursal(sucursal_id=sucursal.id, medio_pago_id=int(medio_id), activo=True)
            db.session.add(ms)
        db.session.commit()
        flash('Sucursal agregada exitosamente', 'success')
        return redirect(url_for('admin_sucursales'))
    return render_template('crear_sucursal.html', medios=medios)

@app.route('/admin/sucursales/<int:sucursal_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_sucursal(sucursal_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    sucursal = Sucursal.query.get_or_404(sucursal_id)
    medios = MedioPago.query.order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    medios_activos = {ms.medio_pago_id for ms in sucursal.medios_sucursal if ms.activo}
    if request.method == 'POST':
        sucursal.nombre = request.form['nombre']
        sucursal.direccion = request.form['direccion']
        db.session.commit()
        # Actualizar medios
        nuevos_ids = set(map(int, request.form.getlist('medios')))
        # Desactivar los que ya no están
        for ms in sucursal.medios_sucursal:
            ms.activo = ms.medio_pago_id in nuevos_ids
        # Agregar nuevos
        existentes = {ms.medio_pago_id for ms in sucursal.medios_sucursal}
        for medio_id in nuevos_ids - existentes:
            db.session.add(MedioSucursal(sucursal_id=sucursal.id, medio_pago_id=medio_id, activo=True))
        db.session.commit()
        flash('Sucursal editada exitosamente', 'success')
        return redirect(url_for('admin_sucursales'))
    return render_template('editar_sucursal.html', sucursal=sucursal, medios=medios, medios_activos=medios_activos)

@app.route('/admin/sucursales/<int:sucursal_id>/eliminar', methods=['POST'])
@login_required
def eliminar_sucursal(sucursal_id):
    try:
        if not current_user.es_admin:
            flash('Acceso denegado', 'error')
            return redirect(url_for('dashboard'))
        
        sucursal = Sucursal.query.get_or_404(sucursal_id)
        
        # Verificar si hay usuarios asignados a esta sucursal
        usuarios_asignados = Usuario.query.filter_by(sucursal_id=sucursal_id).count()
        if usuarios_asignados > 0:
            flash(f'No se puede eliminar la sucursal "{sucursal.nombre}" porque tiene {usuarios_asignados} usuario(s) asignado(s)', 'error')
            return redirect(url_for('admin_sucursales'))
        
        # Verificar si hay operaciones en esta sucursal
        operaciones_count = Operacion.query.filter_by(sucursal_id=sucursal_id).count()
        if operaciones_count > 0:
            flash(f'No se puede eliminar la sucursal "{sucursal.nombre}" porque tiene {operaciones_count} operación(es) registrada(s)', 'error')
            return redirect(url_for('admin_sucursales'))
        
        # Verificar si hay comisiones asociadas
        comisiones_diarias = ComisionDiaria.query.filter_by(sucursal_id=sucursal_id).count()
        comisiones_mensuales = ComisionMensual.query.filter_by(sucursal_id=sucursal_id).count()
        
        if comisiones_diarias > 0 or comisiones_mensuales > 0:
            flash(f'No se puede eliminar la sucursal "{sucursal.nombre}" porque tiene comisiones registradas', 'error')
            return redirect(url_for('admin_sucursales'))
        
        # Eliminar medios de pago asociados a la sucursal
        MedioSucursal.query.filter_by(sucursal_id=sucursal_id).delete()
        
        # Eliminar la sucursal
        db.session.delete(sucursal)
        db.session.commit()
        
        flash(f'Sucursal "{sucursal.nombre}" eliminada exitosamente', 'success')
        return redirect(url_for('admin_sucursales'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar la sucursal: {str(e)}', 'error')
        return redirect(url_for('admin_sucursales'))

# Gestión de usuarios (solo admin)
@app.route('/admin/usuarios')
@login_required
def admin_usuarios():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    
    usuarios = Usuario.query.all()
    return render_template('admin_usuarios.html', usuarios=usuarios)

@app.route('/admin/usuarios/crear', methods=['GET', 'POST'])
@login_required
def crear_usuario():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        nombre_completo = request.form['nombre_completo']
        sucursal_id = request.form['sucursal_id'] if request.form['sucursal_id'] else None
        es_admin = 'es_admin' in request.form
        
        # Verificar si el usuario ya existe
        if Usuario.query.filter_by(username=username).first():
            flash('El nombre de usuario ya existe', 'error')
            return render_template('crear_usuario.html', sucursales=Sucursal.query.all())
        
        nuevo_usuario = Usuario(
            username=username,
            email=email,
            password_hash=generate_password_hash(password),
            nombre_completo=nombre_completo,
            sucursal_id=sucursal_id,
            es_admin=es_admin
        )
        
        db.session.add(nuevo_usuario)
        db.session.commit()
        
        flash('Usuario agregado exitosamente', 'success')
        return redirect(url_for('admin_usuarios'))
    
    return render_template('crear_usuario.html', sucursales=Sucursal.query.all())

@app.route('/admin/usuarios/<int:usuario_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_usuario(usuario_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    
    usuario = Usuario.query.get_or_404(usuario_id)
    
    if request.method == 'POST':
        # Obtener el nuevo username
        nuevo_username = request.form['username']
        
        # Verificar si el nuevo username ya existe (excluyendo el usuario actual)
        if nuevo_username != usuario.username:
            usuario_existente = Usuario.query.filter_by(username=nuevo_username).first()
            if usuario_existente and usuario_existente.id != usuario.id:
                flash('El nombre de usuario ya existe', 'error')
                return render_template('editar_usuario.html', usuario=usuario, sucursales=Sucursal.query.all())
        
        # Actualizar los campos del usuario
        usuario.username = nuevo_username
        usuario.email = request.form['email']
        usuario.nombre_completo = request.form['nombre_completo']
        usuario.sucursal_id = request.form['sucursal_id'] if request.form['sucursal_id'] else None
        usuario.es_admin = 'es_admin' in request.form
        usuario.activo = 'activo' in request.form
        
        # Cambiar contraseña si se proporciona
        if request.form['password']:
            usuario.password_hash = generate_password_hash(request.form['password'])
        
        db.session.commit()
        flash('Usuario modificado exitosamente', 'warning')
        return redirect(url_for('admin_usuarios'))
    
    return render_template('editar_usuario.html', usuario=usuario, sucursales=Sucursal.query.all())

@app.route('/admin/usuarios/<int:usuario_id>/eliminar', methods=['POST'])
@login_required
def eliminar_usuario(usuario_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    
    usuario = Usuario.query.get_or_404(usuario_id)
    
    # No permitir eliminar al admin principal
    if usuario.username == 'admin':
        flash('No se puede eliminar al administrador principal', 'error')
        return redirect(url_for('admin_usuarios'))
    
    # No permitir que un admin se elimine a sí mismo
    if usuario.id == current_user.id:
        flash('No puedes eliminar tu propia cuenta', 'error')
        return redirect(url_for('admin_usuarios'))
    
    try:
        # Primero eliminar todas las operaciones asociadas al usuario usando SQL directo
        db.session.execute(db.text("DELETE FROM operacion WHERE usuario_id = :user_id"), 
                          {"user_id": usuario.id})
        
        # Luego eliminar el usuario
        db.session.delete(usuario)
        db.session.commit()
        flash('Usuario eliminado exitosamente junto con todas sus operaciones', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar usuario: {str(e)}', 'error')
    
    return redirect(url_for('admin_usuarios'))

@app.route('/admin/usuarios/<int:usuario_id>/cambiar-rol', methods=['POST'])
@login_required
def admin_cambiar_rol_usuario(usuario_id):
    if not current_user.es_admin:
        return jsonify({'success': False, 'message': 'Acceso denegado. Solo los administradores pueden cambiar roles.'})
    
    usuario = Usuario.query.get_or_404(usuario_id)
    
    # No permitir cambiar el rol del usuario admin principal
    if usuario.username == 'admin':
        return jsonify({'success': False, 'message': 'No se puede cambiar el rol del usuario administrador principal.'})
    
    # No permitir cambiar su propio rol
    if usuario.id == current_user.id:
        return jsonify({'success': False, 'message': 'No puedes cambiar tu propio rol.'})
    
    data = request.get_json()
    nuevo_rol = data.get('rol')
    
    if nuevo_rol not in ['usuario']:
        return jsonify({'success': False, 'message': 'Rol no válido'})
    
    try:
        usuario.es_admin = False
        db.session.commit()
        return jsonify({'success': True, 'message': f'Usuario convertido a {nuevo_rol} exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Error al cambiar el rol del usuario'})

@app.route('/admin/perfil', methods=['GET', 'POST'])
@login_required
def editar_perfil_admin():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        current_user.email = request.form['email']
        current_user.nombre_completo = request.form['nombre_completo']
        
        # Cambiar contraseña si se proporciona
        if request.form['password']:
            current_user.password_hash = generate_password_hash(request.form['password'])
        
        db.session.commit()
        flash('Perfil modificado exitosamente', 'warning')
        return redirect(url_for('admin_dashboard'))
    
    return render_template('editar_perfil_admin.html', usuario=current_user)

# Operaciones bancarias
@app.route('/operaciones')
@login_required
def operaciones():
    fecha = request.args.get('fecha')
    medio = request.args.get('medio')
    hora_inicio = request.args.get('hora_inicio')
    hora_fin = request.args.get('hora_fin')

    if current_user.es_admin:
        query = Operacion.query
        if request.args.get('sucursal_id'):
            query = query.filter_by(sucursal_id=request.args.get('sucursal_id'))
            # Si se accede desde dashboard admin con sucursal_id, mostrar todas las operaciones sin filtrar por fecha
            fecha = None
    elif False:  # Supervisor eliminado
        
        query = Operacion.query.filter_by(sucursal_id=current_user.sucursal_id)
    else:
        query = Operacion.query.filter(Operacion.usuario_id == current_user.id)

    # Usar la misma lógica de timezone que el registro de operaciones
    peru_tz = pytz.timezone('America/Lima')
    ahora = datetime.now(peru_tz)
    hoy = ahora.date()

    if fecha:
        fecha_objeto = datetime.strptime(fecha, '%Y-%m-%d').date()
        if not current_user.es_admin and fecha_objeto != hoy:
            flash('Solo los administradores pueden consultar operaciones de otros días', 'warning')
            fecha = None
        if fecha:
            # Usar filtro por fecha específica
            query = query.filter(db.func.date(Operacion.hora) == fecha_objeto)
    
    # Solo aplicar filtro de fecha si no es admin accediendo desde dashboard con sucursal_id
    # Usar filtro por fecha del día actual
    if (not fecha and not current_user.es_admin) or (fecha and not (current_user.es_admin and request.args.get('sucursal_id'))):
        query = query.filter(db.func.date(Operacion.hora) == hoy)
    if medio:
        query = query.filter(Operacion.medio == medio)
    if hora_inicio:
        query = query.filter(Operacion.hora >= hora_inicio)
    if hora_fin:
        query = query.filter(Operacion.hora <= hora_fin)

    # OPTIMIZACIÓN: Implementar paginación para mejorar rendimiento
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Mostrar 50 operaciones por página
    
    # Usar left join para incluir operaciones sin sucursal asignada
    operaciones_paginated = query.outerjoin(Usuario).outerjoin(Sucursal).order_by(Operacion.hora.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    operaciones = operaciones_paginated.items
    
    # Debug: Solo mostrar información esencial en desarrollo
    if app.debug:
        print(f"DEBUG OPERACIONES: Operaciones encontradas: {len(operaciones)}")
    
    filtros_aplicados = bool(fecha or medio or hora_inicio or hora_fin or (current_user.es_admin and request.args.get('sucursal_id')))
    
    # Calcular comisión del día si se accede desde dashboard admin
    comision_dia = 0.0
    sucursal_nombre = None
    if current_user.es_admin and request.args.get('sucursal_id'):
        sucursal_id = int(request.args.get('sucursal_id'))
        sucursal = Sucursal.query.get(sucursal_id)
        if sucursal:
            sucursal_nombre = sucursal.nombre
            # Calcular comisión del día para esta sucursal
            comision_diaria = ComisionDiaria.query.filter_by(
                fecha=hoy,
                sucursal_id=sucursal_id
            ).first()
            if comision_diaria:
                comision_dia = float(comision_diaria.total_comision)
    

    
    # Medios activos para todos los usuarios
    medios_pago = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    return render_template('operaciones.html',
                         operaciones=operaciones,
                         fecha_actual=fecha or datetime.now(peru_tz).strftime('%Y-%m-%d'),
                         fecha_hoy=datetime.now(peru_tz).strftime('%Y-%m-%d'),
                         filtros_aplicados=filtros_aplicados,
                         sucursales=Sucursal.query.all() if current_user.es_admin else [],
                         medios_pago=medios_pago,
                         comision_dia=comision_dia,
                         sucursal_nombre=sucursal_nombre)

@app.route('/operaciones/registrar', methods=['GET', 'POST'])
@login_required
def registrar_operacion():
    if request.method == 'POST':
        monto = float(request.form['monto'])
        comision = float(request.form['comision'])
        medio = request.form['medio']
        
        # Determinar la sucursal para la operación
        if current_user.es_admin:
            # Los administradores pueden seleccionar la sucursal
            sucursal_id = request.form.get('sucursal_id')
            if not sucursal_id:
                flash('Debe seleccionar una sucursal para la operación', 'error')
                return redirect(url_for('registrar_operacion'))
            sucursal_id = int(sucursal_id)
        else:
            # Los usuarios regulares usan su sucursal asignada
            if not current_user.sucursal_id:
                flash('Debe tener una sucursal asignada para registrar operaciones', 'error')
                return redirect(url_for('registrar_operacion'))
            sucursal_id = current_user.sucursal_id
        
        # Obtener hora actual en UTC-5 (Perú)
        hora_actual = datetime.now(peru_tz)
        
        nueva_operacion = Operacion(
            monto=monto,
            comision=comision,
            medio=medio,
            hora=hora_actual,
            usuario_id=current_user.id,
            sucursal_id=sucursal_id
        )
        
        db.session.add(nueva_operacion)
        
        # Actualizar comisión diaria
        fecha_hoy = hora_actual.date()
        comision_diaria = ComisionDiaria.query.filter_by(
            fecha=fecha_hoy,
            sucursal_id=sucursal_id
        ).first()
        
        if comision_diaria:
            comision_diaria.total_comision = float(comision_diaria.total_comision) + float(comision)
        else:
            comision_diaria = ComisionDiaria(
                fecha=fecha_hoy,
                sucursal_id=sucursal_id,
                total_comision=comision
            )
            db.session.add(comision_diaria)
        
        # Actualizar comisión mensual
        año_actual = hora_actual.year
        mes_actual = hora_actual.month
        comision_mensual = ComisionMensual.query.filter_by(
            año=año_actual,
            mes=mes_actual,
            sucursal_id=sucursal_id
        ).first()
        
        if comision_mensual:
            comision_mensual.total_comision = float(comision_mensual.total_comision) + float(comision)
        else:
            comision_mensual = ComisionMensual(
                año=año_actual,
                mes=mes_actual,
                sucursal_id=sucursal_id,
                total_comision=comision
            )
            db.session.add(comision_mensual)
        
        db.session.commit()
        flash('Operación bancaria registrada exitosamente', 'success')
        return redirect(url_for('operaciones'))
    
    # Pasar sucursales solo si es administrador
    sucursales = Sucursal.query.filter_by(activa=True).all() if current_user.es_admin else None
    return render_template('registrar_operacion.html', sucursales=sucursales)

# Editar operación
@app.route('/operaciones/<int:operacion_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_operacion(operacion_id):
    operacion = Operacion.query.get_or_404(operacion_id)
    
    # Verificar permisos
    if not current_user.es_admin and operacion.usuario_id != current_user.id:
        flash('No tienes permisos para editar esta operación', 'error')
        return redirect(url_for('operaciones'))
    
    if request.method == 'POST':
        monto_anterior = float(operacion.comision)
        sucursal_anterior = operacion.sucursal_id
        monto = float(request.form['monto'])
        comision = float(request.form['comision'])
        medio = request.form['medio']
        
        # Determinar la nueva sucursal
        if current_user.es_admin:
            nueva_sucursal_id = int(request.form.get('sucursal_id'))
        else:
            nueva_sucursal_id = operacion.sucursal_id
        
        # Actualizar operación
        operacion.monto = monto
        operacion.comision = comision
        operacion.medio = medio
        operacion.sucursal_id = nueva_sucursal_id
        
        # Actualizar comisiones diarias y mensuales
        fecha_operacion = operacion.hora.date()
        año_operacion = operacion.hora.year
        mes_operacion = operacion.hora.month
        
        # Si cambió la sucursal, actualizar ambas sucursales
        if current_user.es_admin and nueva_sucursal_id != sucursal_anterior:
            # Restar de la sucursal anterior
            comision_diaria_anterior = ComisionDiaria.query.filter_by(
                fecha=fecha_operacion,
                sucursal_id=sucursal_anterior
            ).first()
            
            if comision_diaria_anterior:
                comision_diaria_anterior.total_comision = float(comision_diaria_anterior.total_comision) - float(monto_anterior)
            
            comision_mensual_anterior = ComisionMensual.query.filter_by(
                año=año_operacion,
                mes=mes_operacion,
                sucursal_id=sucursal_anterior
            ).first()
            
            if comision_mensual_anterior:
                comision_mensual_anterior.total_comision = float(comision_mensual_anterior.total_comision) - float(monto_anterior)
            
            # Sumar a la nueva sucursal
            comision_diaria_nueva = ComisionDiaria.query.filter_by(
                fecha=fecha_operacion,
                sucursal_id=nueva_sucursal_id
            ).first()
            
            if comision_diaria_nueva:
                comision_diaria_nueva.total_comision = float(comision_diaria_nueva.total_comision) + float(comision)
            else:
                comision_diaria_nueva = ComisionDiaria(
                    fecha=fecha_operacion,
                    sucursal_id=nueva_sucursal_id,
                    total_comision=comision
                )
                db.session.add(comision_diaria_nueva)
            
            comision_mensual_nueva = ComisionMensual.query.filter_by(
                año=año_operacion,
                mes=mes_operacion,
                sucursal_id=nueva_sucursal_id
            ).first()
            
            if comision_mensual_nueva:
                comision_mensual_nueva.total_comision = float(comision_mensual_nueva.total_comision) + float(comision)
            else:
                comision_mensual_nueva = ComisionMensual(
                    año=año_operacion,
                    mes=mes_operacion,
                    sucursal_id=nueva_sucursal_id,
                    total_comision=comision
                )
                db.session.add(comision_mensual_nueva)
        else:
            # Solo actualizar la sucursal actual
            comision_diaria = ComisionDiaria.query.filter_by(
                fecha=fecha_operacion,
                sucursal_id=operacion.sucursal_id
            ).first()
            
            if comision_diaria:
                comision_diaria.total_comision = float(comision_diaria.total_comision) - float(monto_anterior) + float(comision)
            
            comision_mensual = ComisionMensual.query.filter_by(
                año=año_operacion,
                mes=mes_operacion,
                sucursal_id=operacion.sucursal_id
            ).first()
            
            if comision_mensual:
                comision_mensual.total_comision = float(comision_mensual.total_comision) - float(monto_anterior) + float(comision)
        
        db.session.commit()
        flash('Operación bancaria modificada exitosamente', 'warning')
        return redirect(url_for('operaciones'))
    
    # Pasar sucursales solo si es administrador
    sucursales = Sucursal.query.filter_by(activa=True).all() if current_user.es_admin else None
    # Obtener medios de pago activos
    medios_pago = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    
    # DEBUG TEMPORAL
    print(f"🔍 DEBUG - Función editar_operacion()")
    print(f"🔍 DEBUG - Medios obtenidos: {len(medios_pago)}")
    for medio in medios_pago:
        print(f"🔍 DEBUG - Medio: {medio.nombre_abreviado} (orden: {medio.orden})")
    
    return render_template('editar_operacion.html', operacion=operacion, medios_pago=medios_pago, sucursales=sucursales, version="v4")

# Eliminar operación
@app.route('/operaciones/<int:operacion_id>/eliminar', methods=['POST'])
@login_required
def eliminar_operacion(operacion_id):
    operacion = Operacion.query.get_or_404(operacion_id)
    
    # Verificar permisos
    if not current_user.es_admin and operacion.usuario_id != current_user.id:
        flash('No tienes permisos para eliminar esta operación', 'error')
        return redirect(url_for('operaciones'))
    
    monto_comision = float(operacion.comision)
    fecha_operacion = operacion.hora.date()
    año_operacion = operacion.hora.year
    mes_operacion = operacion.hora.month
    sucursal_id = operacion.sucursal_id
    
    # Eliminar operación
    db.session.delete(operacion)
    
    # Actualizar comisiones diarias y mensuales
    comision_diaria = ComisionDiaria.query.filter_by(
        fecha=fecha_operacion,
        sucursal_id=sucursal_id
    ).first()
    
    if comision_diaria:
        comision_diaria.total_comision = float(comision_diaria.total_comision) - float(monto_comision)
    
    comision_mensual = ComisionMensual.query.filter_by(
        año=año_operacion,
        mes=mes_operacion,
        sucursal_id=sucursal_id
    ).first()
    
    if comision_mensual:
        comision_mensual.total_comision = float(comision_mensual.total_comision) - float(monto_comision)
    
    db.session.commit()
    flash('Operación bancaria eliminada exitosamente', 'error')
    return redirect(url_for('operaciones'))

# API para obtener comisiones (solo admin puede ver comisiones mensuales)
@app.route('/api/comisiones')
@login_required
def api_comisiones():
    if current_user.es_admin:
        # Admin puede ver todas las sucursales
        sucursal_id = request.args.get('sucursal_id')
        if sucursal_id:
            sucursal_id = int(sucursal_id)
        else:
            return jsonify({'error': 'sucursal_id requerido para admin'}), 400
    else:
        # Usuario normal solo ve su sucursal
        sucursal_id = current_user.sucursal_id
    
    tipo = request.args.get('tipo', 'diaria')  # 'diaria' o 'mensual'
    
    if tipo == 'diaria':
        fecha = request.args.get('fecha', datetime.now(peru_tz).strftime('%Y-%m-%d'))
        comision = ComisionDiaria.query.filter_by(
            fecha=fecha,
            sucursal_id=sucursal_id
        ).first()
        
        return jsonify({
            'fecha': fecha,
            'total_comision': float(comision.total_comision) if comision else 0
        })
    
    elif tipo == 'mensual' and current_user.es_admin:
        año = request.args.get('año', datetime.now(peru_tz).year)
        mes = request.args.get('mes', datetime.now(peru_tz).month)
        
        comision = ComisionMensual.query.filter_by(
            año=año,
            mes=mes,
            sucursal_id=sucursal_id
        ).first()
        
        return jsonify({
            'año': año,
            'mes': mes,
            'total_comision': float(comision.total_comision) if comision else 0
        })
    
    else:
        return jsonify({'error': 'Acceso denegado'}), 403

# Rutas de reportes (solo para administradores)
@app.route('/reportes')
@login_required
def reportes():
    if not current_user.es_admin:
        flash('Acceso denegado. Solo los administradores pueden generar reportes.', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener medios de pago y sucursales
    medios_pago = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    
    if current_user.es_admin:
        sucursales = Sucursal.query.all()
    else:
        sucursales = [current_user.sucursal] if current_user.sucursal else []
    
    return render_template('reportes.html', sucursales=sucursales, medios_pago=medios_pago)



@app.route('/api/reportes/operaciones')
@login_required
def api_reportes_operaciones():
    # Obtener zona horaria de Perú
    peru_tz = pytz.timezone('America/Lima')
    
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    sucursal_id = request.args.get('sucursal_id')
    medio = request.args.get('medio')
    
    query = Operacion.query
    
    # Debug: Solo mostrar información esencial en desarrollo
    if app.debug:
        print(f"DEBUG REPORTE: Parámetros - fecha_inicio: '{fecha_inicio}', fecha_fin: '{fecha_fin}', sucursal_id: '{sucursal_id}', medio: '{medio}'")
    
    # Aplicar filtros de fecha usando timezone específico para evitar problemas de conversión
    if fecha_inicio:
        # Convertir fecha_inicio a datetime con timezone de Perú
        inicio_fecha = datetime.combine(datetime.strptime(fecha_inicio, '%Y-%m-%d').date(), datetime.min.time()).replace(tzinfo=peru_tz)
        query = query.filter(Operacion.hora >= inicio_fecha)
    
    if fecha_fin:
        # Convertir fecha_fin a datetime con timezone de Perú (hasta el final del día)
        fin_fecha = datetime.combine(datetime.strptime(fecha_fin, '%Y-%m-%d').date(), datetime.max.time()).replace(tzinfo=peru_tz)
        query = query.filter(Operacion.hora <= fin_fecha)
    
    if sucursal_id and sucursal_id.strip():
        # Convertir sucursal_id de string a integer
        try:
            sucursal_id_int = int(sucursal_id)
            query = query.filter(Operacion.sucursal_id == sucursal_id_int)
        except ValueError:
            # Si no se puede convertir a integer, ignorar el filtro
            if app.debug:
                print(f"DEBUG REPORTE: Error al convertir sucursal_id '{sucursal_id}' a integer")
    
    if medio:
        query = query.filter(Operacion.medio == medio)
    
    # OPTIMIZACIÓN: Limitar resultados y usar paginación para reportes grandes
    operaciones = query.order_by(Operacion.hora.desc()).limit(1000).all()
    
    # Debug: Mostrar información de filtros aplicados
    print(f"DEBUG REPORTE: Operaciones encontradas: {len(operaciones)} (limitado a 1000)")
    
    # Debug: Solo mostrar información esencial en desarrollo
    if app.debug:
        print(f"DEBUG REPORTE: Procesando {len(operaciones)} operaciones")
    
    # OPTIMIZACIÓN: Cargar medios de pago una sola vez para evitar N+1 queries
    medios_cache = {mp.nombre_abreviado: mp.nombre_completo for mp in MedioPago.query.all()}
    
    # Preparar datos para el reporte
    datos = []
    for op in operaciones:
        # Usar cache de medios de pago
        medio_nombre = medios_cache.get(op.medio, op.medio)
        
        datos.append({
            'id': op.id,
            'fecha': format_peru_date(op.hora),
            'hora': format_peru_time(op.hora),
            'monto': float(op.monto),
            'comision': float(op.comision),
            'medio': medio_nombre,
            'usuario': op.usuario.nombre_completo,
            'sucursal': op.sucursal.nombre if op.sucursal else 'Sin sucursal'
        })
    
    # Calcular totales
    total_operaciones = len(datos)
    total_monto = sum(op['monto'] for op in datos)
    total_comision = sum(op['comision'] for op in datos)
    
    # Debug: Solo mostrar información esencial en desarrollo
    if app.debug:
        print(f"DEBUG REPORTE: Total operaciones: {total_operaciones}, Total monto: {total_monto}, Total comisión: {total_comision}")
    
    return jsonify({
        'operaciones': datos,
        'total_operaciones': total_operaciones,
        'total_monto': total_monto,
        'total_comision': total_comision
    })

@app.route('/api/reportes/exportar/<formato>')
@login_required
def exportar_reporte(formato):
    if not current_user.es_admin:
        return 'Acceso denegado: solo administradores pueden exportar reportes.', 403
    
    try:
        # Obtener parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        sucursal_id = request.args.get('sucursal_id')
        medio = request.args.get('medio')
        
        # Query base
        query = Operacion.query
        
        # Obtener zona horaria de Perú
        peru_tz = pytz.timezone('America/Lima')
        
        # Aplicar filtros de fecha usando timezone específico para evitar problemas de conversión
        if fecha_inicio:
            # Convertir fecha_inicio a datetime con timezone de Perú
            inicio_fecha = datetime.combine(datetime.strptime(fecha_inicio, '%Y-%m-%d').date(), datetime.min.time()).replace(tzinfo=peru_tz)
            query = query.filter(Operacion.hora >= inicio_fecha)
        
        if fecha_fin:
            # Convertir fecha_fin a datetime con timezone de Perú (hasta el final del día)
            fin_fecha = datetime.combine(datetime.strptime(fecha_fin, '%Y-%m-%d').date(), datetime.max.time()).replace(tzinfo=peru_tz)
            query = query.filter(Operacion.hora <= fin_fecha)
        
        if sucursal_id and sucursal_id.strip():
            # Convertir sucursal_id de string a integer
            try:
                sucursal_id_int = int(sucursal_id)
                query = query.filter(Operacion.sucursal_id == sucursal_id_int)
            except ValueError:
                # Si no se puede convertir a integer, ignorar el filtro
                print(f"DEBUG EXPORT: Error al convertir sucursal_id '{sucursal_id}' a integer")
        if medio:
            query = query.filter(Operacion.medio == medio)
        
        # OPTIMIZACIÓN: Limitar exportación a máximo 5000 registros para evitar timeouts
        operaciones = query.order_by(Operacion.hora.desc()).limit(5000).all()
        
        # Función para obtener el nombre completo del medio
        def get_medio_nombre(medio_abreviado):
            medio_obj = MedioPago.query.filter_by(nombre_abreviado=medio_abreviado).first()
            return medio_obj.nombre_completo if medio_obj else medio_abreviado
        
        if formato == 'csv':
            import csv
            from io import StringIO
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(['N°', 'Fecha', 'Hora', 'Monto', 'Comisión', 'Medio', 'Usuario', 'Sucursal'])
            for idx, op in enumerate(operaciones, 1):
                writer.writerow([
                    idx,
                    format_peru_date(op.hora),
                    format_peru_time(op.hora),
                    float(op.monto),
                    float(op.comision),
                    get_medio_nombre(op.medio),
                    op.usuario.nombre_completo,
                    op.sucursal.nombre if op.sucursal else 'Sin sucursal'
                ])
            from flask import Response
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': 'attachment; filename=reporte_operaciones.csv'}
            )
        elif formato == 'xlsx':
            import openpyxl
            from openpyxl import Workbook
            from io import BytesIO
            wb = Workbook()
            ws = wb.active
            ws.title = "Operaciones"
            headers = ['N°', 'Fecha', 'Hora', 'Monto', 'Comisión', 'Medio', 'Usuario', 'Sucursal']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            for idx, op in enumerate(operaciones, 1):
                row = idx + 1
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=format_peru_date(op.hora))
                ws.cell(row=row, column=3, value=format_peru_time(op.hora))
                ws.cell(row=row, column=4, value=float(op.monto))
                ws.cell(row=row, column=5, value=float(op.comision))
                ws.cell(row=row, column=6, value=get_medio_nombre(op.medio))
                ws.cell(row=row, column=7, value=op.usuario.nombre_completo)
                ws.cell(row=row, column=8, value=op.sucursal.nombre if op.sucursal else 'Sin sucursal')
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            from flask import Response
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': 'attachment; filename=reporte_operaciones.xlsx'}
            )
        elif formato == 'pdf':
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from io import BytesIO
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            # Texto de prueba para verificar actualización
            elements.append(Paragraph('REPORTE ACTUALIZADO - PRUEBA DE NUMERACIÓN CONSECUTIVA', styles['Title']))
            elements.append(Spacer(1, 12))
            data = [['N°', 'Fecha', 'Hora', 'Monto', 'Comisión', 'Medio', 'Usuario', 'Sucursal']]
            for idx, op in enumerate(operaciones, 1):
                data.append([
                    str(idx),
                    format_peru_date(op.hora),
                    format_peru_time(op.hora),
                    f'S/. {float(op.monto):.2f}',
                    f'S/. {float(op.comision):.2f}',
                    get_medio_nombre(op.medio),
                    op.usuario.nombre_completo,
                    op.sucursal.nombre if op.sucursal else 'Sin sucursal'
                ])
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            doc.build(elements)
            output.seek(0)
            from flask import Response
            return Response(
                output.getvalue(),
                mimetype='application/pdf',
                headers={'Content-Disposition': 'attachment; filename=reporte_operaciones.pdf'}
            )
        else:
            return 'Formato no soportado', 400
    except Exception as e:
        return f'Error al generar reporte: {str(e)}', 500


# API para actualizar operaciones (edición inline)
@app.route('/api/operaciones/<int:operacion_id>', methods=['PUT'])
@login_required
def api_actualizar_operacion(operacion_id):
    try:
        # Obtener la operación
        operacion = Operacion.query.get_or_404(operacion_id)
        
        # Verificar permisos: solo el usuario que creó la operación o admin puede editarla
        if not current_user.es_admin and operacion.usuario_id != current_user.id:
            return jsonify({'success': False, 'message': 'No tienes permisos para editar esta operación'}), 403
        
        # Obtener datos del JSON
        data = request.get_json()
        monto = data.get('monto')
        comision = data.get('comision')
        medio = data.get('medio')
        
        # Validar datos
        if not monto or not comision or not medio:
            return jsonify({'success': False, 'message': 'Todos los campos son requeridos'}), 400
        
        try:
            monto = float(monto)
            comision = float(comision)
        except ValueError:
            return jsonify({'success': False, 'message': 'Monto y comisión deben ser números válidos'}), 400
        
        if monto <= 0:
            return jsonify({'success': False, 'message': 'El monto debe ser mayor a 0'}), 400
        
        if comision < 0:
            return jsonify({'success': False, 'message': 'La comisión no puede ser negativa'}), 400
        
        # Validar medio de pago contra la base de datos
        medio_valido = MedioPago.query.filter_by(nombre_abreviado=medio, activo=True).first()
        if not medio_valido:
            return jsonify({'success': False, 'message': 'Medio de pago no válido'}), 400
        
        # Guardar valores originales para actualizar comisiones
        monto_original = float(operacion.monto)
        comision_original = float(operacion.comision)
        fecha_operacion = operacion.hora.date()
        año_operacion = operacion.hora.year
        mes_operacion = operacion.hora.month
        sucursal_id = operacion.sucursal_id
        
        # Actualizar operación
        operacion.monto = monto
        operacion.comision = comision
        operacion.medio = medio
        
        # Actualizar comisiones diarias
        comision_diaria = ComisionDiaria.query.filter_by(
            fecha=fecha_operacion,
            sucursal_id=sucursal_id
        ).first()
        
        if comision_diaria:
            # Restar comisión original y sumar nueva
            comision_diaria.total_comision = float(comision_diaria.total_comision) - comision_original + comision
        
        # Actualizar comisiones mensuales
        comision_mensual = ComisionMensual.query.filter_by(
            año=año_operacion,
            mes=mes_operacion,
            sucursal_id=sucursal_id
        ).first()
        
        if comision_mensual:
            # Restar comisión original y sumar nueva
            comision_mensual.total_comision = float(comision_mensual.total_comision) - comision_original + comision
        
        db.session.commit()
        
        # No flash aquí para evitar duplicados en AJAX
        response_data = {
            'success': True,
            'message': 'Operación actualizada exitosamente',
            'monto': monto,
            'comision': comision,
            'medio': medio
        }
        
        print(f"✅ Respuesta: {response_data}")
        return jsonify(response_data)
        
    except Exception as e:
        print(f"❌ Error: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al actualizar la operación: {str(e)}'}), 500

# API para registrar operación (POST)
@app.route('/api/operaciones', methods=['POST'])
@login_required
def api_registrar_operacion():
    data = request.get_json()
    monto = data.get('monto')
    comision = data.get('comision')
    medio = data.get('medio')
    if not monto or not comision or not medio:
        return jsonify({'success': False, 'message': 'Todos los campos son requeridos'}), 400
    try:
        monto = float(monto)
        comision = float(comision)
    except ValueError:
        return jsonify({'success': False, 'message': 'Monto y comisión deben ser números válidos'}), 400
    if monto <= 0:
        return jsonify({'success': False, 'message': 'El monto debe ser mayor a 0'}), 400
    if comision < 0:
        return jsonify({'success': False, 'message': 'La comisión no puede ser negativa'}), 400
    medios_validos = ['BCP', 'KS', 'IBK', 'BN', 'AQP', 'NIUBIZ', 'CONFIANZA', 'BBVA', 'ICA', 'IZIPAY', 'CULQI', 'BIM', 'Interbank', 'Scotiabank', 'Efectivo', 'Yape', 'Plin']
    if medio not in medios_validos:
        return jsonify({'success': False, 'message': 'Medio de pago no válido'}), 400
    if current_user.es_admin:
        sucursal_id = data.get('sucursal_id')
        if not sucursal_id:
            return jsonify({'success': False, 'message': 'Debe seleccionar una sucursal'}), 400
        sucursal_id = int(sucursal_id)
    else:
        if not current_user.sucursal_id:
            return jsonify({'success': False, 'message': 'Debe tener una sucursal asignada'}), 400
        sucursal_id = current_user.sucursal_id
    hora_actual = datetime.now(peru_tz)
    nueva_operacion = Operacion(
        monto=monto,
        comision=comision,
        medio=medio,
        hora=hora_actual,
        usuario_id=current_user.id,
        sucursal_id=sucursal_id
    )
    db.session.add(nueva_operacion)
    # Actualizar comisión diaria
    fecha_hoy = hora_actual.date()
    comision_diaria = ComisionDiaria.query.filter_by(
        fecha=fecha_hoy,
        sucursal_id=sucursal_id
    ).first()
    if comision_diaria:
        comision_diaria.total_comision = float(comision_diaria.total_comision) + float(comision)
    else:
        comision_diaria = ComisionDiaria(
            fecha=fecha_hoy,
            sucursal_id=sucursal_id,
            total_comision=comision
        )
        db.session.add(comision_diaria)
    # Actualizar comisión mensual
    año_actual = hora_actual.year
    mes_actual = hora_actual.month
    comision_mensual = ComisionMensual.query.filter_by(
        año=año_actual,
        mes=mes_actual,
        sucursal_id=sucursal_id
    ).first()
    if comision_mensual:
        comision_mensual.total_comision = float(comision_mensual.total_comision) + float(comision)
    else:
        comision_mensual = ComisionMensual(
            año=año_actual,
            mes=mes_actual,
            sucursal_id=sucursal_id,
            total_comision=comision
        )
        db.session.add(comision_mensual)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Operación registrada exitosamente'})

# API para eliminar operación (DELETE)
@app.route('/api/operaciones/<int:operacion_id>', methods=['DELETE'])
@login_required
def api_eliminar_operacion(operacion_id):
    operacion = Operacion.query.get_or_404(operacion_id)
    if not current_user.es_admin and operacion.usuario_id != current_user.id:
        return jsonify({'success': False, 'message': 'No tienes permisos para eliminar esta operación'}), 403
    monto_comision = float(operacion.comision)
    fecha_operacion = operacion.hora.date()
    año_operacion = operacion.hora.year
    mes_operacion = operacion.hora.month
    sucursal_id = operacion.sucursal_id
    db.session.delete(operacion)
    # Actualizar comisiones diarias y mensuales
    comision_diaria = ComisionDiaria.query.filter_by(
        fecha=fecha_operacion,
        sucursal_id=sucursal_id
    ).first()
    if comision_diaria:
        comision_diaria.total_comision = float(comision_diaria.total_comision) - float(monto_comision)
    comision_mensual = ComisionMensual.query.filter_by(
        año=año_operacion,
        mes=mes_operacion,
        sucursal_id=sucursal_id
    ).first()
    if comision_mensual:
        comision_mensual.total_comision = float(comision_mensual.total_comision) - float(monto_comision)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Operación eliminada exitosamente'})

@app.route('/admin/medios', methods=['GET', 'POST'])
@login_required
def admin_medios():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        nombre_abreviado = request.form['nombre_abreviado'].strip()
        nombre_completo = request.form['nombre_completo'].strip()
        if nombre_abreviado and nombre_completo:
            if not MedioPago.query.filter_by(nombre_abreviado=nombre_abreviado).first():
                medio = MedioPago(nombre_abreviado=nombre_abreviado, nombre_completo=nombre_completo)
                db.session.add(medio)
                db.session.commit()
                flash('Medio de pago agregado', 'success')
            else:
                flash('Ese nombre abreviado ya existe', 'warning')
        return redirect(url_for('admin_medios'))
    medios = MedioPago.query.order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    return render_template('admin_medios.html', medios=medios)

@app.route('/admin/medios/<int:medio_id>/eliminar', methods=['POST'])
@login_required
def eliminar_medio(medio_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    medio = MedioPago.query.get_or_404(medio_id)
    db.session.delete(medio)
    db.session.commit()
    flash('Medio de pago eliminado', 'success')
    return redirect(url_for('admin_medios'))

@app.route('/admin/medios/<int:medio_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_medio(medio_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    medio = MedioPago.query.get_or_404(medio_id)
    if request.method == 'POST':
        nombre_abreviado = request.form['nombre_abreviado'].strip()
        nombre_completo = request.form['nombre_completo'].strip()
        if nombre_abreviado and nombre_completo:
            medio.nombre_abreviado = nombre_abreviado
            medio.nombre_completo = nombre_completo
            db.session.commit()
            flash('Medio de pago editado', 'success')
        return redirect(url_for('admin_medios'))
    return render_template('editar_medio.html', medio=medio)

@app.route('/admin/medios/<int:medio_id>/activar', methods=['POST'])
@login_required
def activar_medio(medio_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    medio = MedioPago.query.get_or_404(medio_id)
    medio.activo = not medio.activo
    db.session.commit()
    flash('Estado de medio actualizado', 'success')
    return redirect(url_for('admin_medios'))

@app.route('/admin/medios/<int:medio_id>/subir', methods=['POST'])
@login_required
def subir_medio(medio_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    medio = MedioPago.query.get_or_404(medio_id)
    anterior = MedioPago.query.filter(MedioPago.orden < medio.orden).order_by(MedioPago.orden.desc()).first()
    if anterior:
        medio.orden, anterior.orden = anterior.orden, medio.orden
        db.session.commit()
    return redirect(url_for('admin_medios'))

@app.route('/admin/medios/<int:medio_id>/bajar', methods=['POST'])
@login_required
def bajar_medio(medio_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    medio = MedioPago.query.get_or_404(medio_id)
    siguiente = MedioPago.query.filter(MedioPago.orden > medio.orden).order_by(MedioPago.orden.asc()).first()
    if siguiente:
        medio.orden, siguiente.orden = siguiente.orden, medio.orden
        db.session.commit()
    return redirect(url_for('admin_medios'))

@app.route('/admin/sucursales/<int:sucursal_id>/toggle_medio', methods=['POST'])
@login_required
def toggle_medio_sucursal(sucursal_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    try:
        medio_id = request.form.get('medio_id')
        if not medio_id:
            return jsonify({'error': 'ID del medio requerido'}), 400
        
        # Buscar o crear la relación
        ms = MedioSucursal.query.filter_by(
            sucursal_id=sucursal_id, 
            medio_pago_id=medio_id
        ).first()
        
        if ms:
            ms.activo = not ms.activo
        else:
            ms = MedioSucursal(
                sucursal_id=sucursal_id,
                medio_pago_id=medio_id,
                activo=True
            )
            db.session.add(ms)
        
        db.session.commit()
        return jsonify({'success': True, 'activo': ms.activo})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Rutas para el módulo de TAREO - ADMINISTRADOR
@app.route('/admin/tareos')
@login_required
def admin_tareos():
    if not current_user.es_admin:
        flash('Acceso denegado: solo administradores pueden gestionar tareos.', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener todos los tareos con información de usuario y sucursal
    tareos = Tareo.query.join(Usuario, Tareo.usuario_id == Usuario.id)\
                        .join(Sucursal, Tareo.sucursal_id == Sucursal.id)\
                        .add_columns(
                            Tareo.id,
                            Tareo.nombre,
                            Tareo.descripcion,
                            Tareo.estado,
                            Tareo.fecha_creacion,
                            Tareo.fecha_completado,
                            Usuario.nombre_completo.label('usuario_nombre'),
                            Sucursal.nombre.label('sucursal_nombre')
                        ).order_by(Tareo.fecha_creacion.desc()).all()
    
    # Obtener usuarios y sucursales para el formulario
    usuarios = Usuario.query.filter_by(activo=True).order_by(Usuario.nombre_completo).all()
    sucursales = Sucursal.query.filter_by(activa=True).order_by(Sucursal.nombre).all()
    medios = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden).all()
    
    return render_template('admin_tareos.html', 
                         tareos=tareos, 
                         usuarios=usuarios, 
                         sucursales=sucursales,
                         medios=medios)

@app.route('/admin/tareos/crear', methods=['GET', 'POST'])
@login_required
def crear_tareo():
    if not current_user.es_admin:
        flash('Acceso denegado: solo administradores pueden crear tareos.', 'error')
        return redirect(url_for('admin_tareos'))
    
    if request.method == 'POST':
        try:
            # Datos del tareo
            nombre = request.form.get('nombre')
            descripcion = request.form.get('descripcion', '')
            usuario_id = request.form.get('usuario_id')
            sucursal_id = request.form.get('sucursal_id')
            
            if not all([nombre, usuario_id, sucursal_id]):
                flash('Todos los campos obligatorios deben estar completos.', 'error')
                return redirect(url_for('crear_tareo'))
            
            # Crear el tareo
            tareo = Tareo(
                nombre=nombre,
                descripcion=descripcion,
                usuario_id=int(usuario_id),
                sucursal_id=int(sucursal_id),
                created_by=current_user.id
            )
            db.session.add(tareo)
            db.session.flush()  # Para obtener el ID del tareo
            
            # Procesar operaciones del tareo
            operaciones_data = request.form.getlist('operaciones[]')
            for i, operacion_str in enumerate(operaciones_data):
                if operacion_str.strip():
                    # Formato esperado: "medio|destino|nombre|monto"
                    partes = operacion_str.split('|')
                    if len(partes) == 4:
                        medio, destino, nombre_op, monto = partes
                        try:
                            operacion = OperacionTareo(
                                tareo_id=tareo.id,
                                medio=medio.strip(),
                                destino=destino.strip(),
                                nombre=nombre_op.strip(),
                                monto=float(monto.strip()),
                                orden=i + 1
                            )
                            db.session.add(operacion)
                        except ValueError:
                            flash(f'Error en el monto de la operación {i+1}: {monto}', 'error')
                            continue
            
            db.session.commit()
            flash('Tareo creado exitosamente.', 'success')
            return redirect(url_for('admin_tareos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el tareo: {str(e)}', 'error')
            return redirect(url_for('crear_tareo'))
    
    # GET: mostrar formulario
    usuarios = Usuario.query.filter_by(activo=True).order_by(Usuario.nombre_completo).all()
    sucursales = Sucursal.query.filter_by(activa=True).order_by(Sucursal.nombre).all()
    medios = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden).all()
    
    return render_template('crear_tareo.html', 
                         usuarios=usuarios, 
                         sucursales=sucursales,
                         medios=medios)

@app.route('/admin/tareos/<int:tareo_id>')
@login_required
def ver_tareo(tareo_id):
    if not current_user.es_admin:
        flash('Acceso denegado: solo administradores pueden ver tareos.', 'error')
        return redirect(url_for('admin_tareos'))
    
    tareo = Tareo.query.get_or_404(tareo_id)
    operaciones = OperacionTareo.query.filter_by(tareo_id=tareo_id).order_by(OperacionTareo.orden).all()
    
    return render_template('ver_tareo.html', tareo=tareo, operaciones=operaciones)

@app.route('/admin/tareos/<int:tareo_id>/eliminar', methods=['POST'])
@login_required
def eliminar_tareo(tareo_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    try:
        tareo = Tareo.query.get_or_404(tareo_id)
        db.session.delete(tareo)
        db.session.commit()
        flash('Tareo eliminado exitosamente.', 'success')
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# API para agregar operación a un tareo existente
@app.route('/api/tareos/<int:tareo_id>/operaciones', methods=['POST'])
@login_required
def agregar_operacion_tareo(tareo_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    try:
        tareo = Tareo.query.get_or_404(tareo_id)
        
        medio = request.json.get('medio')
        destino = request.json.get('destino')
        nombre = request.json.get('nombre')
        monto = request.json.get('monto')
        
        if not all([medio, destino, nombre, monto]):
            return jsonify({'error': 'Todos los campos son requeridos'}), 400
        
        # Obtener el siguiente orden
        ultimo_orden = db.session.query(db.func.max(OperacionTareo.orden))\
                                 .filter_by(tareo_id=tareo_id).scalar() or 0
        
        operacion = OperacionTareo(
            tareo_id=tareo_id,
            medio=medio,
            destino=destino,
            nombre=nombre,
            monto=float(monto),
            orden=ultimo_orden + 1
        )
        
        db.session.add(operacion)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'operacion': {
                'id': operacion.id,
                'medio': operacion.medio,
                'destino': operacion.destino,
                'nombre': operacion.nombre,
                'monto': float(operacion.monto),
                'orden': operacion.orden,
                'completado': operacion.completado,
                'fecha_completado': operacion.fecha_completado.isoformat() if operacion.fecha_completado else None
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Rutas para el módulo de TAREO - USUARIOS
@app.route('/tareos')
@login_required
def tareos_usuario():
    # Obtener tareos asignados al usuario actual
    tareos = Tareo.query.filter_by(usuario_id=current_user.id)\
                        .join(Sucursal, Tareo.sucursal_id == Sucursal.id)\
                        .add_columns(
                            Tareo.id,
                            Tareo.nombre,
                            Tareo.descripcion,
                            Tareo.estado,
                            Tareo.fecha_creacion,
                            Tareo.fecha_completado,
                            Sucursal.nombre.label('sucursal_nombre')
                        ).order_by(Tareo.fecha_creacion.desc()).all()
    
    return render_template('tareos_usuario.html', tareos=tareos)

@app.route('/tareos/<int:tareo_id>')
@login_required
def ver_tareo_usuario(tareo_id):
    # Verificar que el tareo esté asignado al usuario actual
    tareo = Tareo.query.filter_by(id=tareo_id, usuario_id=current_user.id).first_or_404()
    operaciones = OperacionTareo.query.filter_by(tareo_id=tareo_id).order_by(OperacionTareo.orden).all()
    
    # Verificar si el tareo está habilitado para el día actual
    fecha_actual = datetime.now(peru_tz).date()
    fecha_tareo = tareo.fecha_creacion.date()
    tareo_habilitado = fecha_actual == fecha_tareo
    
    return render_template('ver_tareo_usuario.html', 
                         tareo=tareo, 
                         operaciones=operaciones, 
                         tareo_habilitado=tareo_habilitado,
                         fecha_actual=fecha_actual,
                         fecha_tareo=fecha_tareo)

# API para marcar operación como completada - ULTRA OPTIMIZADA
@app.route('/api/tareos/operaciones/<int:operacion_id>/completar', methods=['POST'])
@login_required
def completar_operacion_tareo(operacion_id):
    try:
        # ULTRA OPTIMIZACIÓN: Obtener operación, tareo y todas las operaciones en una sola consulta
        operacion = OperacionTareo.query.join(Tareo)\
                                        .filter(
                                            OperacionTareo.id == operacion_id,
                                            Tareo.usuario_id == current_user.id
                                        ).first_or_404()
        
        tareo = operacion.tareo
        completado = request.json.get('completado', True)
        
        # Verificar si el tareo está habilitado para el día actual
        fecha_actual = get_peru_time().date()
        fecha_tareo = tareo.fecha_creacion.date()
        
        if fecha_actual != fecha_tareo:
            return jsonify({
                'success': False,
                'error': 'Este tareo no está habilitado para el día actual. Los tareos se deshabilitan automáticamente al cambiar de día.'
            }), 400
        
        # Actualizar la operación
        if completado:
            operacion.completado = True
            operacion.fecha_completado = get_peru_time()
        else:
            operacion.completado = False
            operacion.fecha_completado = None
        
        # ULTRA OPTIMIZACIÓN: Calcular estado del tareo usando consulta SQL directa
        # Esto es mucho más rápido que cargar todas las operaciones en memoria
        from sqlalchemy import func, case
        
        # Contar operaciones totales y completadas en una sola consulta
        stats = db.session.query(
            func.count(OperacionTareo.id).label('total'),
            func.sum(case((OperacionTareo.completado == True, 1), else_=0)).label('completadas')
        ).filter(OperacionTareo.tareo_id == tareo.id).first()
        
        total_operaciones = stats.total or 0
        operaciones_completadas = stats.completadas or 0
        
        # Determinar el nuevo estado del tareo
        if operaciones_completadas == total_operaciones and total_operaciones > 0:
            tareo.estado = 'completado'
            tareo.fecha_completado = get_peru_time()
        elif operaciones_completadas > 0:
            tareo.estado = 'en_progreso'
        else:
            tareo.estado = 'pendiente'
            tareo.fecha_completado = None
        
        # Commit inmediato
        db.session.commit()
        
        return jsonify({
            'success': True,
            'operacion': {
                'id': operacion.id,
                'completado': operacion.completado,
                'fecha_completado': operacion.fecha_completado.isoformat() if operacion.fecha_completado else None
            },
            'tareo': {
                'estado': tareo.estado,
                'fecha_completado': tareo.fecha_completado.isoformat() if tareo.fecha_completado else None
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# API para editar operación de tareo (solo admin)
@app.route('/api/tareos/operaciones/<int:operacion_id>/editar', methods=['POST'])
@login_required
def editar_operacion_tareo(operacion_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado. Solo administradores pueden editar operaciones.'}), 403
    
    try:
        operacion = OperacionTareo.query.get_or_404(operacion_id)
        
        # Obtener datos del formulario
        nombre = request.json.get('nombre')
        medio = request.json.get('medio')
        destino = request.json.get('destino')
        monto = request.json.get('monto')
        orden = request.json.get('orden')
        
        # Validar datos
        if not all([nombre, medio, destino, monto, orden]):
            return jsonify({'error': 'Todos los campos son requeridos.'}), 400
        
        # Actualizar operación
        operacion.nombre = nombre
        operacion.medio = medio
        operacion.destino = destino
        operacion.monto = Decimal(str(monto))
        operacion.orden = int(orden)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'mensaje': 'Operación actualizada correctamente',
            'operacion': {
                'id': operacion.id,
                'nombre': operacion.nombre,
                'medio': operacion.medio,
                'destino': operacion.destino,
                'monto': float(operacion.monto),
                'orden': operacion.orden
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# API para eliminar operación de tareo (solo admin)
@app.route('/api/tareos/operaciones/<int:operacion_id>/eliminar', methods=['DELETE'])
@login_required
def eliminar_operacion_tareo(operacion_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado. Solo administradores pueden eliminar operaciones.'}), 403
    
    try:
        operacion = OperacionTareo.query.get_or_404(operacion_id)
        tareo_id = operacion.tareo_id
        
        db.session.delete(operacion)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'mensaje': 'Operación eliminada correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# API para aleatorizar montos de operaciones de tareo (solo admin)
@app.route('/api/tareos/<int:tareo_id>/aleatorizar-montos', methods=['POST'])
@login_required
def aleatorizar_montos_tareo(tareo_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado. Solo administradores pueden aleatorizar montos.'}), 403
    
    try:
        import random
        from decimal import Decimal
        
        # Obtener el tareo y sus operaciones
        tareo = Tareo.query.get_or_404(tareo_id)
        operaciones = OperacionTareo.query.filter_by(tareo_id=tareo_id).all()
        
        if not operaciones:
            return jsonify({'error': 'No hay operaciones en este tareo para aleatorizar.'}), 400
        
        # Función para aleatorizar monto según el tipo de operación
        def aleatorizar_monto(medio):
            if medio.upper() == 'BBVA':
                return Decimal(str(random.randint(10, 40)))
            elif medio.upper() == 'KS':
                return Decimal(str(random.randint(100, 150)))
            elif medio.upper() == 'BN':
                return Decimal('10.00')
            else:
                # Para otros medios, mantener el monto actual
                return None
        
        # Actualizar montos
        operaciones_actualizadas = []
        for operacion in operaciones:
            nuevo_monto = aleatorizar_monto(operacion.medio)
            if nuevo_monto is not None:
                operacion.monto = nuevo_monto
                operaciones_actualizadas.append({
                    'id': operacion.id,
                    'medio': operacion.medio,
                    'monto_anterior': str(operacion.monto),
                    'monto_nuevo': str(nuevo_monto)
                })
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'mensaje': f'Se aleatorizaron {len(operaciones_actualizadas)} operaciones',
            'operaciones_actualizadas': operaciones_actualizadas
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# API para verificar si un tareo está habilitado para el día actual
@app.route('/api/tareos/<int:tareo_id>/verificar-habilitado', methods=['GET'])
@login_required
def verificar_tareo_habilitado(tareo_id):
    try:
        tareo = Tareo.query.filter_by(id=tareo_id, usuario_id=current_user.id).first_or_404()
        
        fecha_actual = datetime.now(peru_tz).date()
        fecha_tareo = tareo.fecha_creacion.date()
        
        habilitado = fecha_actual == fecha_tareo
        
        return jsonify({
            'success': True,
            'habilitado': habilitado,
            'fecha_actual': fecha_actual.isoformat(),
            'fecha_tareo': fecha_tareo.isoformat()
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API para aleatorización automática diaria (se ejecuta cuando se deshabilitan los checklists)
@app.route('/api/tareos/<int:tareo_id>/aleatorizacion-automatica', methods=['POST'])
@login_required
def aleatorizacion_automatica_tareo(tareo_id):
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado. Solo administradores pueden ejecutar la aleatorización automática.'}), 403
    
    try:
        import random
        from decimal import Decimal
        
        # Obtener el tareo y sus operaciones
        tareo = Tareo.query.get_or_404(tareo_id)
        operaciones = OperacionTareo.query.filter_by(tareo_id=tareo_id).all()
        
        if not operaciones:
            return jsonify({'error': 'No hay operaciones en este tareo para aleatorizar.'}), 400
        
        # Función para aleatorizar monto según el tipo de operación
        def aleatorizar_monto(medio):
            if medio.upper() == 'BBVA':
                return Decimal(str(random.randint(10, 40)))
            elif medio.upper() == 'KS':
                return Decimal(str(random.randint(100, 150)))
            elif medio.upper() == 'BN':
                return Decimal('10.00')
            else:
                # Para otros medios, mantener el monto actual
                return None
        
        # Actualizar montos y resetear estado de completado
        operaciones_actualizadas = []
        for operacion in operaciones:
            nuevo_monto = aleatorizar_monto(operacion.medio)
            if nuevo_monto is not None:
                operacion.monto = nuevo_monto
                operaciones_actualizadas.append({
                    'id': operacion.id,
                    'medio': operacion.medio,
                    'monto_nuevo': str(nuevo_monto)
                })
            
            # Resetear estado de completado para el nuevo día
            operacion.completado = False
            operacion.fecha_completado = None
        
        # Resetear estado del tareo
        tareo.estado = 'pendiente'
        tareo.fecha_completado = None
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'mensaje': f'Aleatorización automática completada. Se aleatorizaron {len(operaciones_actualizadas)} operaciones y se reseteó el estado del tareo.',
            'operaciones_actualizadas': operaciones_actualizadas
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Debug temporal para diagnosticar problema de filtro de fecha
@app.route('/debug/filtro-fecha')
@login_required
def debug_filtro_fecha():
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403
    
    # Simular el filtro exacto que está fallando
    fecha_inicio = "2025-07-26"
    fecha_fin = "2025-07-26"
    sucursal_id = "2"  # TECKNOVATION
    
    print(f"DEBUG TEMPORAL: Parámetros recibidos:")
    print(f"DEBUG TEMPORAL: - fecha_inicio: '{fecha_inicio}'")
    print(f"DEBUG TEMPORAL: - fecha_fin: '{fecha_fin}'")
    print(f"DEBUG TEMPORAL: - sucursal_id: '{sucursal_id}'")
    
    query = Operacion.query
    
    # Procesar fechas como en la función de reportes
    if fecha_inicio:
        fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        inicio_dia_peru = datetime.combine(fecha_inicio_obj, datetime.min.time()).replace(tzinfo=peru_tz)
        inicio_dia_utc_naive = inicio_dia_peru.astimezone(pytz.utc).replace(tzinfo=None)
        query = query.filter(Operacion.hora >= inicio_dia_utc_naive)
        
        print(f"DEBUG TEMPORAL: Fecha inicio procesada:")
        print(f"DEBUG TEMPORAL: - fecha_inicio_obj: {fecha_inicio_obj}")
        print(f"DEBUG TEMPORAL: - inicio_dia_peru: {inicio_dia_peru}")
        print(f"DEBUG TEMPORAL: - inicio_dia_utc_naive: {inicio_dia_utc_naive}")
    
    if fecha_fin:
        fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        fin_dia_peru = datetime.combine(fecha_fin_obj, datetime.max.time()).replace(tzinfo=peru_tz)
        fin_dia_utc_naive = fin_dia_peru.astimezone(pytz.utc).replace(tzinfo=None)
        query = query.filter(Operacion.hora <= fin_dia_utc_naive)
        
        print(f"DEBUG TEMPORAL: Fecha fin procesada:")
        print(f"DEBUG TEMPORAL: - fecha_fin_obj: {fecha_fin_obj}")
        print(f"DEBUG TEMPORAL: - fin_dia_peru: {fin_dia_peru}")
        print(f"DEBUG TEMPORAL: - fin_dia_utc_naive: {fin_dia_utc_naive}")
    
    if sucursal_id and sucursal_id.strip():
        try:
            sucursal_id_int = int(sucursal_id)
            query = query.filter(Operacion.sucursal_id == sucursal_id_int)
            print(f"DEBUG TEMPORAL: Filtro sucursal aplicado: {sucursal_id_int}")
        except ValueError:
            print(f"DEBUG TEMPORAL: Error al convertir sucursal_id '{sucursal_id}' a integer")
    
    # Ejecutar query
    operaciones = query.order_by(Operacion.hora.desc()).all()
    
    print(f"DEBUG TEMPORAL: Operaciones encontradas: {len(operaciones)}")
    
    # Agrupar por fecha
    operaciones_por_fecha = {}
    for op in operaciones:
        fecha = op.hora.date()
        if fecha not in operaciones_por_fecha:
            operaciones_por_fecha[fecha] = []
        operaciones_por_fecha[fecha].append(op)
    
    # Preparar datos para respuesta
    resultado = {
        'parametros': {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'sucursal_id': sucursal_id
        },
        'rangos_tiempo': {
            'inicio_dia_utc_naive': str(inicio_dia_utc_naive) if fecha_inicio else None,
            'fin_dia_utc_naive': str(fin_dia_utc_naive) if fecha_fin else None
        },
        'total_operaciones': len(operaciones),
        'operaciones_por_fecha': {}
    }
    
    for fecha in sorted(operaciones_por_fecha.keys(), reverse=True):
        ops = operaciones_por_fecha[fecha]
        total_fecha = sum(float(op.comision) for op in ops)
        resultado['operaciones_por_fecha'][str(fecha)] = {
            'cantidad': len(ops),
            'total_comision': total_fecha,
            'operaciones': [
                {
                    'id': op.id,
                    'hora': str(op.hora),
                    'fecha': str(op.hora.date()),
                    'comision': float(op.comision)
                }
                for op in ops[:5]  # Solo las primeras 5
            ]
        }
    
    return jsonify(resultado)

if __name__ == '__main__':
    try:
        with app.app_context():
            db.create_all()
            # Crear usuario admin por defecto si no existe
            admin = Usuario.query.filter_by(username='admin').first()
            if not admin:
                admin = Usuario(
                    username='admin',
                    email='admin@sisagent.com',
                    password_hash=generate_password_hash('61442159'),
                    nombre_completo='Administrador SISAGENT',
                    es_admin=True,
                    sucursal_id=None  # El admin no tiene sucursal asignada inicialmente
                )
                db.session.add(admin)
                db.session.commit()
    except Exception as e:
        print(f"Error durante la inicialización: {e}")
    
    import os
    port = int(os.environ.get("PORT", 5000))
    # En producción, no usar debug mode
    debug_mode = os.environ.get('FLASK_ENV') == 'development'
    app.run(host='0.0.0.0', debug=debug_mode, use_reloader=False, port=port)

print("🎉 SISAGENT Flask cargado completamente - Listo para producción!") 