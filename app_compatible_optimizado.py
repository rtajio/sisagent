#!/usr/bin/env python3
"""
SISAGENT - Sistema de Gestión de Operaciones Bancarias
VERSIÓN COMPATIBLE ULTRA OPTIMIZADA
"""

import os
import sys
import uuid
from datetime import datetime, timedelta
from decimal import Decimal
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session, Response, abort
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from flask_caching import Cache
from flask_compress import Compress
import pytz
import time
import base64
import json
import anthropic
import threading
from sqlalchemy import or_
from flask_sock import Sock

# Configuración de zona horaria (UTC-5 para Perú)
peru_tz = pytz.timezone('America/Lima')

def get_peru_time():
    """Obtiene la hora actual en zona horaria de Perú"""
    return datetime.now(peru_tz)

# Función removida - usar db.func.date() directamente en las consultas
# Las operaciones se guardan con get_peru_time() que ya incluye la zona horaria de Perú
# PostgreSQL está configurado con timezone 'America/Lima' en las conexiones

def _fecha_rango(fecha_str, es_fin=False):
    """
    Convierte una fecha 'YYYY-MM-DD' en un datetime naive en hora Perú
    comparable con los valores guardados en la DB.
    TODO se guarda como wall-clock Perú (naive).
    """
    try:
        d = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None
    if es_fin:
        dt_peru = datetime.combine(d, datetime.max.time()).replace(
            tzinfo=peru_tz, hour=23, minute=59, second=59, microsecond=999999)
    else:
        dt_peru = datetime.combine(d, datetime.min.time()).replace(tzinfo=peru_tz)

    # Retornar como naive Perú (wall-clock) para comparar directamente con BD
    return dt_peru.replace(tzinfo=None)


def _to_peru(dt):
    """
    Normaliza cualquier datetime a hora de Perú (America/Lima, UTC-5).

    Estrategia unificada: TODO se guarda como wall-clock Perú (naive) en ambos backends.
    - Con tzinfo: convertir a Perú (puede venir de librerías externas)
    - Sin tzinfo: asumir que YA es hora Perú, solo agregar zona para operaciones
    """
    if dt is None:
        return None
    if dt.tzinfo is not None:
        # Ya tiene zona horaria → convertir a Perú
        return dt.astimezone(peru_tz)
    # Naive: asumir que YA es hora Perú (guarda como wall-clock en la BD)
    return peru_tz.localize(dt)

def format_peru_time(dt):
    """Formatea una hora para mostrar en zona horaria de Perú"""
    t = _to_peru(dt)
    return t.strftime('%H:%M:%S') if t else ""

def format_peru_date(dt):
    """Formatea una fecha para mostrar en zona horaria de Perú"""
    t = _to_peru(dt)
    return t.strftime('%d/%m/%Y') if t else ""

def format_peru_datetime(dt):
    """Formatea fecha/hora completa en zona horaria de Perú"""
    t = _to_peru(dt)
    return t.strftime('%d/%m/%Y %H:%M:%S') if t else ""

def format_peru_datetime_short(dt):
    """Formatea fecha/hora corta en zona horaria de Perú"""
    t = _to_peru(dt)
    return t.strftime('%d/%m/%Y %H:%M') if t else ""


def _contexto_fecha_hora():
    """Linea de contexto para inyectar en los prompts del chatbot: fecha y hora
    REAL de Peru (UTC-5). Sin esto el modelo asume UTC al preguntarle la hora."""
    ahora = get_peru_time()
    # Formato 12h: HH:MMam/pm (ej: 09:05am, 04:29pm) — NUNCA uses UTC
    hora12 = ahora.strftime('%I:%M%p').lower().replace('am', 'am').replace('pm', 'pm')
    return (
        f"\n\nFECHA Y HORA ACTUAL en Peru (UTC-5): {ahora.strftime('%d/%m/%Y')}, {hora12}. "
        "Usa SIEMPRE esta hora/fecha de Peru (HH:MMam/pm) para responder 'que hora es', 'que dia es' o "
        "para interpretar 'hoy'. NUNCA uses UTC ni otra zona horaria."
    )

import math as _math

# Configuración del cálculo automático de comisión
# Regla: cada COMISION_RANGO_SOLES de monto → +COMISION_POR_RANGO de comisión.
# Por defecto: cada 100 soles → 1 sol de comisión (configurable solo cambiando estas constantes).
COMISION_RANGO_SOLES = 100
COMISION_POR_RANGO = 1.0

def calcular_comision_sugerida(monto):
    """Devuelve la comisión sugerida por la fórmula ⌈monto/rango⌉ × comision_por_rango.

    Ejemplos con rango=100, comision=1:
        monto=50   → 1.0  (un rango completo, aunque parcial)
        monto=100  → 1.0
        monto=101  → 2.0  (rebasa al siguiente rango)
        monto=2000 → 20.0
    """
    try:
        m = float(monto or 0)
    except (TypeError, ValueError):
        return 0.0
    if m <= 0:
        return 0.0
    rangos = _math.ceil(m / COMISION_RANGO_SOLES)
    return round(rangos * COMISION_POR_RANGO, 2)

print("[*] SISAGENT Flask COMPATIBLE ULTRA OPTIMIZADO arrancando...")
print("[!] OPTIMIZACIONES: Caché, Compresión, Consultas optimizadas, Paginación")
print("[~] Actualización Railway - " + get_peru_time().strftime("%Y-%m-%d %H:%M:%S"))

# Configuración de la aplicación Flask
app = Flask(__name__)

print("[OK] Flask app creada")

# Configuración para Railway
if os.environ.get('DATABASE_URL'):
    database_url = os.environ.get('DATABASE_URL')
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    print(f"[OK] Usando PostgreSQL en Railway: {database_url[:20]}...")
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sisagent_consolidada.db'
    print("[OK] Usando SQLite para desarrollo local")

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'tu-clave-secreta-aqui')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Configuración del chatbot IA (Anthropic Claude Haiku 4.5)
app.config['ANTHROPIC_API_KEY'] = os.getenv('ANTHROPIC_API_KEY', '')
ANTHROPIC_MODEL = os.getenv('ANTHROPIC_MODEL', 'claude-haiku-4-5')

# Configuración de transcripción de voz (Google Gemini — capa gratuita 1500 req/día)
app.config['GEMINI_API_KEY'] = os.getenv('GEMINI_API_KEY', '')
GEMINI_TRANSCRIPTION_MODEL = os.getenv('GEMINI_TRANSCRIPTION_MODEL', 'gemini-2.5-flash-lite')
GEMINI_CHAT_MODEL = os.getenv('GEMINI_CHAT_MODEL', 'gemini-2.5-flash-lite')  # cerebro del chat de texto (function-calling)
GEMINI_API_URL = 'https://generativelanguage.googleapis.com/v1beta/models'
GEMINI_MAX_AUDIO_BYTES = 18 * 1024 * 1024  # 18 MB (Gemini admite hasta 20MB inline)

# Modelo Live API (transcripción en streaming via WebSocket)
GEMINI_LIVE_MODEL = os.getenv('GEMINI_LIVE_MODEL', 'models/gemini-3.1-flash-live-preview')  # acepta languageCode (transcribe español confiable)
GEMINI_LIVE_WS_URL = 'wss://generativelanguage.googleapis.com/ws/google.ai.generativelanguage.v1beta.GenerativeService.BidiGenerateContent'

# OPTIMIZACIÓN ULTRA: Configuración de caché
app.config['CACHE_TYPE'] = 'simple'
app.config['CACHE_DEFAULT_TIMEOUT'] = 300  # 5 minutos

# OPTIMIZACIÓN ULTRA: Configuración de compresión
app.config['COMPRESS_MIMETYPES'] = [
    'text/html', 'text/css', 'text/xml', 'text/javascript',
    'application/json', 'application/javascript'
]
app.config['COMPRESS_LEVEL'] = 6
app.config['COMPRESS_MIN_SIZE'] = 500

# Inicializar extensiones
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# WebSocket (para Gemini Live API streaming)
# ping/pong de keepalive para que el proxy de Railway no corte la conexion por inactividad
app.config['SOCK_SERVER_OPTIONS'] = {'ping_interval': 25}
sock = Sock(app)

# OPTIMIZACIÓN ULTRA: Inicializar caché y compresión
cache = Cache(app)
Compress(app)

# Context processor para exponer funciones de formato en templates
@app.context_processor
def inject_format_functions():
    return {
        'format_peru_time': format_peru_time,
        'format_peru_date': format_peru_date,
        'format_peru_datetime': format_peru_datetime,
        'format_peru_datetime_short': format_peru_datetime_short,
    }

# Manejo global de errores para debugging
@app.errorhandler(Exception)
def handle_error(e):
    import traceback
    error_msg = f"Error: {type(e).__name__}: {str(e)}"
    print("=" * 80)
    print("ERROR EN LA APLICACIÓN:")
    print(error_msg)
    print("=" * 80)
    traceback.print_exc()
    print("=" * 80)

    # Intentar rollback de la sesión si hay error de BD
    try:
        db.session.rollback()
    except:
        pass

    # Si el usuario está autenticado, redirigir al dashboard con mensaje de error
    try:
        from flask_login import current_user
        if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated:
            flash(f'Error: {str(e)}', 'error')
            return redirect(url_for('dashboard'))
    except:
        pass

    # Si no está autenticado, redirigir al login
    flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('login'))

print("[OK] Configuración de base de datos completada")
print("[OK] SQLAlchemy y LoginManager configurados")
print("[OK] Caché y compresión configurados")
print("[OK] Configuración de zona horaria completada")

# Modelos de base de datos COMPATIBLES con la estructura existente
class Sucursal(db.Model):
    __tablename__ = 'sucursal'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    direccion = db.Column(db.String(200))
    activa = db.Column(db.Boolean, default=True)
    operaciones = db.relationship('Operacion', backref='sucursal', lazy='dynamic')

class Usuario(UserMixin, db.Model):
    __tablename__ = 'usuario'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), nullable=False, default='')
    password_hash = db.Column(db.String(120), nullable=False)
    nombre_completo = db.Column(db.String(100), nullable=True)
    es_admin = db.Column(db.Boolean, default=False)
    es_admin_sucursal = db.Column(db.Boolean, default=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'))
    # Token de sesión: se regenera al cambiar contraseña o al forzar cierre de sesión
    session_token  = db.Column(db.String(36), nullable=True)
    # Último acceso: se actualiza en cada request autenticado (para presencia en tiempo real)
    ultimo_acceso  = db.Column(db.DateTime, nullable=True)
    # "Entrenamiento" de voz: vocabulario propio + correcciones (sesga la transcripción).
    vocabulario_voz = db.Column(db.Text, nullable=True)
    sucursal = db.relationship('Sucursal', backref='usuarios', lazy='joined')
    operaciones = db.relationship('Operacion', backref='usuario', lazy='dynamic')

    def regenerate_session_token(self):
        """Genera un nuevo token, invalidando todas las sesiones activas."""
        self.session_token = str(uuid.uuid4())

    def esta_en_linea(self):
        """True si el usuario tuvo actividad en los últimos 10 minutos.
        ultimo_acceso se guarda en hora Perú (naive wall-clock)."""
        if self.ultimo_acceso is None:
            return False
        ahora_peru = get_peru_time().replace(tzinfo=None)
        delta = ahora_peru - self.ultimo_acceso
        return delta.total_seconds() < 600  # 10 minutos

    def es_admin_de_sucursal(self):
        return self.es_admin_sucursal and self.sucursal_id is not None

    def es_admin_o_admin_sucursal(self):
        return self.es_admin or self.es_admin_de_sucursal()

    def puede_crear_usuario_en_sucursal(self, sucursal_id):
        if self.es_admin:
            return True
        if self.es_admin_de_sucursal():
            return self.sucursal_id == sucursal_id
        return False

class Operacion(db.Model):
    __tablename__ = 'operacion'
    id = db.Column(db.Integer, primary_key=True)
    monto = db.Column(db.Numeric(10, 2), nullable=False)
    comision = db.Column(db.Numeric(10, 2), nullable=False)
    medio = db.Column(db.String(50), nullable=False)
    hora = db.Column(db.DateTime, default=lambda: get_peru_time().replace(tzinfo=None))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False)
    # Nuevos campos para comisión automática + auditoría del override manual
    comision_sugerida = db.Column(db.Numeric(10, 2), nullable=True)  # lo que sugirió la fórmula
    comision_manual = db.Column(db.Boolean, default=False)            # TRUE si el operador la cambió
    motivo_descuento = db.Column(db.String(200), nullable=True)       # texto libre opcional

class MedioPago(db.Model):
    __tablename__ = 'medio_pago'
    id = db.Column(db.Integer, primary_key=True)
    nombre_abreviado = db.Column(db.String(20), unique=True, nullable=False)
    nombre_completo = db.Column(db.String(100), nullable=False)
    activo = db.Column(db.Boolean, default=True)
    orden = db.Column(db.Integer, default=0)

class MedioSucursal(db.Model):
    """Tabla intermedia para la relación many-to-many entre Sucursal y MedioPago"""
    __tablename__ = 'medio_sucursal'
    id = db.Column(db.Integer, primary_key=True)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False, index=True)
    medio_pago_id = db.Column(db.Integer, db.ForeignKey('medio_pago.id'), nullable=False, index=True)
    activo = db.Column(db.Boolean, default=True, index=True)
    sucursal = db.relationship('Sucursal', backref='medios_sucursal')
    medio_pago = db.relationship('MedioPago', backref='sucursales_medio')
    
    # Evitar duplicados
    __table_args__ = (db.UniqueConstraint('sucursal_id', 'medio_pago_id', name='uq_sucursal_medio'),)

class ComisionDiaria(db.Model):
    __tablename__ = 'comision_diaria'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False)
    total_comision = db.Column(db.Numeric(10, 2), default=0.0)

class ComisionMensual(db.Model):
    __tablename__ = 'comision_mensual'
    id = db.Column(db.Integer, primary_key=True)
    año = db.Column(db.Integer, nullable=False)
    mes = db.Column(db.Integer, nullable=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False)
    total_comision = db.Column(db.Numeric(10, 2), default=0.0)

class Producto(db.Model):
    __tablename__ = 'producto'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(150), nullable=False)
    descripcion = db.Column(db.String(255))
    precio = db.Column(db.Numeric(10, 2), nullable=False)
    stock = db.Column(db.Integer, nullable=False, default=0)
    # La foto se guarda directamente en la base de datos (funciona igual en local
    # y en Railway, donde el filesystem es efímero y se perdería un archivo subido a disco)
    foto = db.Column(db.LargeBinary)
    foto_mimetype = db.Column(db.String(50))
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False)
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=lambda: get_peru_time().replace(tzinfo=None))
    sucursal = db.relationship('Sucursal', backref='productos')

class Venta(db.Model):
    __tablename__ = 'venta'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=lambda: get_peru_time().replace(tzinfo=None))
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    cantidad = db.Column(db.Integer, nullable=False, default=1)
    precio_unitario = db.Column(db.Numeric(10, 2), nullable=False)
    monto = db.Column(db.Numeric(10, 2), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False)
    usuario = db.relationship('Usuario', backref='ventas')
    sucursal = db.relationship('Sucursal', backref='ventas')
    producto = db.relationship('Producto', backref='ventas')

class CajaVentas(db.Model):
    __tablename__ = 'caja_ventas'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False)
    sucursal_id = db.Column(db.Integer, db.ForeignKey('sucursal.id'), nullable=False)
    total_vendido = db.Column(db.Numeric(10, 2), default=0.0)
    saldo = db.Column(db.Numeric(10, 2), default=0.0)
    sucursal = db.relationship('Sucursal', backref='caja_ventas')

class PronunciacionAprendida(db.Model):
    """Banco de memoria de pronunciaciones aprendidas por el modelo de voz"""
    __tablename__ = 'pronunciacion_aprendida'
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    termino_original = db.Column(db.String(200), nullable=False)  # "Tecnovation" (cómo lo dijo el usuario)
    termino_correcto = db.Column(db.String(200), nullable=False)  # "TECKNOVATION" (cómo debe ser)
    tipo = db.Column(db.String(50), default='sucursal')  # 'sucursal', 'producto', 'medio', etc.
    frecuencia = db.Column(db.Integer, default=1)  # cuántas veces se mencionó
    fecha_aprendida = db.Column(db.DateTime, default=lambda: get_peru_time().replace(tzinfo=None))
    usuario = db.relationship('Usuario', backref='pronunciaciones_aprendidas')

    __table_args__ = (
        db.UniqueConstraint('usuario_id', 'termino_original', 'termino_correcto', name='uq_pronunciacion'),
    )

# OPTIMIZACIÓN ULTRA: Cache de medios de pago
@cache.memoize(timeout=300)
def get_medios_pago_cache():
    """Obtener medios de pago con caché"""
    return MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()

# OPTIMIZACIÓN ULTRA: Cache de sucursales activas
@cache.memoize(timeout=300)
def get_sucursales_activas_cache():
    """Obtener sucursales activas con caché"""
    return Sucursal.query.filter_by(activa=True).all()

# OPTIMIZACIÓN ULTRA: Cache de estadísticas del dashboard
@cache.memoize(timeout=60)
def get_dashboard_stats_cache(user_id, is_admin):
    """Obtener estadísticas del dashboard con caché"""
    ahora = get_peru_time()
    hoy = ahora.date()
    
    # Calcular rango de tiempo para el día en hora de Perú (00:00:00 a 23:59:59)
    inicio_dia = datetime.combine(hoy, datetime.min.time()).replace(tzinfo=peru_tz)
    fin_dia = datetime.combine(hoy, datetime.max.time()).replace(tzinfo=peru_tz)
    # Ajustar fin_dia para que sea 23:59:59.999999
    fin_dia = fin_dia.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    if is_admin:
        # Estadísticas para admin - usar rango de tiempo en hora de Perú
        operaciones_hoy = db.session.query(db.func.count(Operacion.id)).filter(
            Operacion.hora >= inicio_dia,
            Operacion.hora <= fin_dia
        ).scalar() or 0
        
        total_operaciones = db.session.query(db.func.count(Operacion.id)).scalar() or 0
        
        comision_hoy = db.session.query(db.func.coalesce(db.func.sum(Operacion.comision), 0.0)).filter(
            Operacion.hora >= inicio_dia,
            Operacion.hora <= fin_dia
        ).scalar() or 0
        
        total_comision = db.session.query(db.func.coalesce(db.func.sum(Operacion.comision), 0.0)).scalar() or 0
        
        sucursales_activas = db.session.query(db.func.count(Sucursal.id)).filter_by(activa=True).scalar() or 0
        
        usuarios_total = db.session.query(db.func.count(Usuario.id)).scalar() or 0
        
    else:
        # Estadísticas para usuario normal - usar rango de tiempo en hora de Perú
        operaciones_hoy = db.session.query(db.func.count(Operacion.id)).filter(
            Operacion.usuario_id == user_id,
            Operacion.hora >= inicio_dia,
            Operacion.hora <= fin_dia
        ).scalar() or 0
        
        total_operaciones = db.session.query(db.func.count(Operacion.id)).filter(
            Operacion.usuario_id == user_id
        ).scalar() or 0
        
        comision_hoy = db.session.query(db.func.coalesce(db.func.sum(Operacion.comision), 0.0)).filter(
            Operacion.usuario_id == user_id,
            Operacion.hora >= inicio_dia,
            Operacion.hora <= fin_dia
        ).scalar() or 0
        
        total_comision = db.session.query(db.func.coalesce(db.func.sum(Operacion.comision), 0.0)).filter(
            Operacion.usuario_id == user_id
        ).scalar() or 0
        
        sucursales_activas = 1  # Usuario normal solo ve su sucursal
        usuarios_total = 1  # Usuario normal solo se ve a sí mismo
    
    return {
        'operaciones_hoy': operaciones_hoy,
        'total_operaciones': total_operaciones,
        'comision_hoy': float(comision_hoy),
        'total_comision': float(total_comision),
        'sucursales_activas': sucursales_activas,
        'usuarios_total': usuarios_total
    }

@login_manager.user_loader
def load_user(user_id):
    user = Usuario.query.get(int(user_id))
    if user is None:
        return None
    # Si el usuario tiene token definido, validar que coincida con el de la sesión actual
    if user.session_token is not None:
        if session.get('_user_session_token') != user.session_token:
            return None  # Token inválido → fuerza logout
    return user

@app.before_request
def actualizar_ultimo_acceso():
    """Actualiza el timestamp de último acceso del usuario autenticado.
    Guarda en hora Perú (wall-clock naive) para consistencia total."""
    if current_user and current_user.is_authenticated:
        ahora_peru = get_peru_time().replace(tzinfo=None)
        ultimo = current_user.ultimo_acceso
        # Solo escribir en DB si pasaron más de 60 s (evita escritura en cada request)
        if ultimo is None or (ahora_peru - ultimo).total_seconds() > 60:
            try:
                current_user.ultimo_acceso = ahora_peru
                db.session.commit()
            except Exception:
                db.session.rollback()

# OPTIMIZACIÓN ULTRA: Función para limpiar caché
def clear_cache():
    """Limpiar todo el caché"""
    cache.clear()

def asegurar_admin_existe():
    """Asegurar que el usuario admin exista con la contraseña 'vivalavida'"""
    try:
        admin = Usuario.query.filter_by(username='admin').first()
        password_hash = generate_password_hash('vivalavida')
        
        if not admin:
            # Crear admin si no existe
            admin = Usuario(
                username='admin',
                email='admin@sisagent.local',
                password_hash=password_hash,
                es_admin=True,
                es_admin_sucursal=False,
                nombre_completo='Administrador Global'
            )
            db.session.add(admin)
            db.session.commit()
            print("[OK] Usuario 'admin' creado con contraseña 'vivalavida'")
        else:
            # Actualizar contraseña y asegurar que es admin
            admin.password_hash = password_hash
            admin.es_admin = True
            admin.es_admin_sucursal = False
            # Asegurar que tenga email si no lo tiene
            if not admin.email:
                admin.email = 'admin@sisagent.local'
            if not admin.nombre_completo:
                admin.nombre_completo = 'Administrador Global'
            db.session.commit()
            print("[OK] Usuario 'admin' actualizado con contraseña 'vivalavida'")
    except Exception as e:
        print(f"[!!]  Error al asegurar admin: {e}")
        db.session.rollback()

# Rutas optimizadas
@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Asegurar que el admin exista antes de procesar el login
    try:
        asegurar_admin_existe()
    except:
        pass
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = Usuario.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            # Asegurar que el usuario tenga session_token
            if user.session_token is None:
                user.regenerate_session_token()
                db.session.commit()
            login_user(user)
            session['_user_session_token'] = user.session_token
            clear_cache()
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Redirige al dashboard según rol usando templates existentes"""
    try:
        stats = get_dashboard_stats_cache(current_user.id, current_user.es_admin)
        base_ctx = {
            'total_sucursales': stats.get('sucursales_activas', 0),
            'total_usuarios': stats.get('usuarios_total', 0),
            'comisiones_hoy': [],
            'comisiones_mes': {},
            **stats,
        }

        if current_user.es_admin:
            # Calcular comisiones por sucursal (día y mes) para admin global
            ahora = get_peru_time()
            hoy = ahora.date()
            año_actual = ahora.year
            mes_actual = ahora.month
            
            # Calcular rango de tiempo para hoy en hora de Perú (00:00:00 a 23:59:59)
            inicio_hoy = datetime.combine(hoy, datetime.min.time()).replace(tzinfo=peru_tz)
            fin_hoy = datetime.combine(hoy, datetime.max.time()).replace(tzinfo=peru_tz)
            fin_hoy = fin_hoy.replace(hour=23, minute=59, second=59, microsecond=999999)
            
            # Calcular rango de tiempo para el mes (desde el primer día del mes hasta hoy)
            inicio_mes = datetime.combine(datetime(año_actual, mes_actual, 1).date(), datetime.min.time()).replace(tzinfo=peru_tz)

            # Obtener todas las sucursales activas
            sucursales_activas = Sucursal.query.filter_by(activa=True).all()
            
            # Obtener comisiones del día por sucursal
            comisiones_hoy_query = db.session.query(
                Operacion.sucursal_id,
                db.func.coalesce(db.func.sum(Operacion.comision), 0.0).label('total')
            ).filter(
                Operacion.hora >= inicio_hoy,
                Operacion.hora <= fin_hoy
            ).group_by(Operacion.sucursal_id).all()
            
            comisiones_hoy_dict = {suc_id: float(total) for suc_id, total in comisiones_hoy_query}

            # Obtener comisiones del mes por sucursal
            comisiones_mes_query = db.session.query(
                Operacion.sucursal_id,
                db.func.coalesce(db.func.sum(Operacion.comision), 0.0).label('total')
            ).filter(
                Operacion.hora >= inicio_mes,
                Operacion.hora <= fin_hoy
            ).group_by(Operacion.sucursal_id).all()
            
            comisiones_mes_dict = {suc_id: float(total) for suc_id, total in comisiones_mes_query}

            # Crear lista con todas las sucursales, incluso si no tienen comisiones
            comisiones_hoy_list = []
            comisiones_mes_dict_final = {}
            
            for sucursal in sucursales_activas:
                comision_hoy = comisiones_hoy_dict.get(sucursal.id, 0.0)
                comision_mes = comisiones_mes_dict.get(sucursal.id, 0.0)
                comisiones_hoy_list.append((sucursal.id, sucursal.nombre, comision_hoy))
                comisiones_mes_dict_final[sucursal.id] = comision_mes

            base_ctx['comisiones_hoy'] = comisiones_hoy_list
            base_ctx['comisiones_mes'] = comisiones_mes_dict_final

            return render_template('admin_dashboard.html', **base_ctx)
        elif current_user.es_admin_de_sucursal():
            # Dashboard para administrador de sucursal
            ahora = get_peru_time()
            hoy = ahora.date()
            año_actual = ahora.year
            mes_actual = ahora.month
            
            # Calcular rango de tiempo para hoy en hora de Perú (00:00:00 a 23:59:59)
            inicio_hoy = datetime.combine(hoy, datetime.min.time()).replace(tzinfo=peru_tz)
            fin_hoy = datetime.combine(hoy, datetime.max.time()).replace(tzinfo=peru_tz)
            fin_hoy = fin_hoy.replace(hour=23, minute=59, second=59, microsecond=999999)
            
            # Calcular rango de tiempo para el mes (desde el primer día del mes hasta hoy)
            inicio_mes = datetime.combine(datetime(año_actual, mes_actual, 1).date(), datetime.min.time()).replace(tzinfo=peru_tz)
            
            sucursal_id = current_user.sucursal_id
            
            # Obtener comisiones del día para la sucursal
            comision_hoy = db.session.query(
                db.func.coalesce(db.func.sum(Operacion.comision), 0.0)
            ).filter(
                Operacion.sucursal_id == sucursal_id,
                Operacion.hora >= inicio_hoy,
                Operacion.hora <= fin_hoy
            ).scalar() or 0.0
            
            # Obtener comisiones del mes para la sucursal
            comision_mes = db.session.query(
                db.func.coalesce(db.func.sum(Operacion.comision), 0.0)
            ).filter(
                Operacion.sucursal_id == sucursal_id,
                Operacion.hora >= inicio_mes,
                Operacion.hora <= fin_hoy
            ).scalar() or 0.0
            
            # Obtener usuarios de la sucursal
            usuarios_sucursal = Usuario.query.filter_by(sucursal_id=sucursal_id).all()
            
            # Calcular comisiones por usuario del día
            comisiones_usuarios_hoy = []
            for usuario in usuarios_sucursal:
                comision_usuario_hoy = db.session.query(
                    db.func.coalesce(db.func.sum(Operacion.comision), 0.0)
                ).filter(
                    Operacion.usuario_id == usuario.id,
                    Operacion.hora >= inicio_hoy,
                    Operacion.hora <= fin_hoy
                ).scalar() or 0.0
                comisiones_usuarios_hoy.append({
                    'usuario_id': usuario.id,
                    'nombre': usuario.nombre_completo or usuario.username,
                    'comision': float(comision_usuario_hoy)
                })
            
            base_ctx['comision_hoy'] = float(comision_hoy)
            base_ctx['comision_mes'] = float(comision_mes)
            base_ctx['usuarios_sucursal'] = usuarios_sucursal
            base_ctx['comisiones_usuarios_hoy'] = comisiones_usuarios_hoy
            base_ctx['sucursal'] = current_user.sucursal

            return render_template('admin_sucursal_dashboard.html', **base_ctx)
        else:
            # Para usuarios normales: agregar variables que el template espera
            ahora = get_peru_time()
            hoy = ahora.date()
            # Calcular rango de tiempo para el día en hora de Perú (00:00:00 a 23:59:59)
            inicio_dia = datetime.combine(hoy, datetime.min.time()).replace(tzinfo=peru_tz)
            fin_dia = datetime.combine(hoy, datetime.max.time()).replace(tzinfo=peru_tz)
            fin_dia = fin_dia.replace(hour=23, minute=59, second=59, microsecond=999999)
            
            base_ctx['total_comision_hoy'] = stats.get('comision_hoy', 0.0)
            # Obtener operaciones del día para el usuario usando rango de tiempo
            operaciones_hoy_list = Operacion.query.filter(
                Operacion.usuario_id == current_user.id,
                Operacion.hora >= inicio_dia,
                Operacion.hora <= fin_dia
            ).order_by(Operacion.hora.desc()).limit(10).all()
            base_ctx['operaciones_hoy'] = operaciones_hoy_list

            return render_template('user_dashboard.html', **base_ctx)
    except Exception as e:
        import traceback
        traceback.print_exc()
        # Respuesta explícita para depurar en producción
        return f"Error en dashboard: {str(e)}", 500

@app.route('/operaciones')
@login_required
def operaciones():
    # OPTIMIZACIÓN ULTRA: Paginación para operaciones
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Reducido para mayor velocidad
    
    # Obtener parámetros de filtro
    fecha = request.args.get('fecha')
    medio = request.args.get('medio')
    hora_inicio = request.args.get('hora_inicio')
    hora_fin = request.args.get('hora_fin')
    
    # Query base optimizada según el rol del usuario
    if current_user.es_admin:
        # Admin global: puede ver todas las operaciones
        query = Operacion.query
        if request.args.get('sucursal_id'):
            query = query.filter_by(sucursal_id=request.args.get('sucursal_id'))
    elif current_user.es_admin_de_sucursal():
        # Admin de sucursal: puede ver operaciones de su sucursal
        query = Operacion.query.filter_by(sucursal_id=current_user.sucursal_id)
    else:
        # Usuario normal: solo puede ver sus propias operaciones
        query = Operacion.query.filter_by(usuario_id=current_user.id)
    
    # Obtener fecha actual para comparación en hora de Perú
    ahora = get_peru_time()
    hoy = ahora.date()
    
    # Calcular rango de tiempo para hoy en hora de Perú (00:00:00 a 23:59:59)
    inicio_hoy = datetime.combine(hoy, datetime.min.time()).replace(tzinfo=peru_tz)
    fin_hoy = datetime.combine(hoy, datetime.max.time()).replace(tzinfo=peru_tz)
    fin_hoy = fin_hoy.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Aplicar filtros
    if fecha:
        fecha_objeto = datetime.strptime(fecha, '%Y-%m-%d').date()
        if not current_user.es_admin_o_admin_sucursal() and fecha_objeto != hoy:
            flash('Solo los administradores pueden consultar operaciones de otros días', 'warning')
            fecha = None
        
        if fecha:
            # Usar rango de tiempo para la fecha específica en hora de Perú
            inicio_fecha = datetime.combine(fecha_objeto, datetime.min.time()).replace(tzinfo=peru_tz)
            fin_fecha = datetime.combine(fecha_objeto, datetime.max.time()).replace(tzinfo=peru_tz)
            fin_fecha = fin_fecha.replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(
                Operacion.hora >= inicio_fecha,
                Operacion.hora <= fin_fecha
            )
    
    if not fecha or not current_user.es_admin_o_admin_sucursal():
        # Usar rango de tiempo para hoy en hora de Perú
        query = query.filter(
            Operacion.hora >= inicio_hoy,
            Operacion.hora <= fin_hoy
        )
    
    if medio:
        query = query.filter(Operacion.medio == medio)
    
    if hora_inicio:
        query = query.filter(Operacion.hora >= hora_inicio)
    
    if hora_fin:
        query = query.filter(Operacion.hora <= hora_fin)
    
    # OPTIMIZACIÓN ULTRA: Paginación
    operaciones_paginated = query.order_by(Operacion.hora.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Detectar si hay filtros aplicados
    filtros_aplicados = bool(fecha or medio or hora_inicio or hora_fin or (current_user.es_admin and request.args.get('sucursal_id')))
    
    # OPTIMIZACIÓN ULTRA: Cargar sucursales solo si es admin global
    sucursales = []
    if current_user.es_admin:
        sucursales = get_sucursales_activas_cache()

    # Cargar medios de pago filtrados por sucursal
    if current_user.es_admin:
        # Admin: si hay sucursal_id en params, filtrar; si no, mostrar todos
        filtro_sucursal_id = request.args.get('sucursal_id', type=int)
        if filtro_sucursal_id:
            medios_pago = MedioPago.query.join(
                MedioSucursal,
                (MedioSucursal.medio_pago_id == MedioPago.id) &
                (MedioSucursal.sucursal_id == filtro_sucursal_id) &
                (MedioSucursal.activo == True)
            ).filter(MedioPago.activo == True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
        else:
            medios_pago = get_medios_pago_cache()
    else:
        # Usuario/admin de sucursal: filtrar medios habilitados para su sucursal
        sucursal_id_usuario = current_user.sucursal_id
        if sucursal_id_usuario:
            medios_pago = MedioPago.query.join(
                MedioSucursal,
                (MedioSucursal.medio_pago_id == MedioPago.id) &
                (MedioSucursal.sucursal_id == sucursal_id_usuario) &
                (MedioSucursal.activo == True)
            ).filter(MedioPago.activo == True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
        else:
            medios_pago = get_medios_pago_cache()
    
    # Calcular comisión del día para sucursal específica si es admin global
    comision_dia = 0.0
    sucursal_nombre = None
    if current_user.es_admin and request.args.get('sucursal_id'):
        try:
            sucursal_id = int(request.args.get('sucursal_id'))
            sucursal = Sucursal.query.get(sucursal_id)
            if sucursal:
                sucursal_nombre = sucursal.nombre
                # Calcular comisión del día para esta sucursal
                comision_dia_query = db.session.query(
                    db.func.coalesce(db.func.sum(Operacion.comision), 0.0)
                ).filter(
                    Operacion.sucursal_id == sucursal_id,
                    Operacion.hora >= inicio_hoy,
                    Operacion.hora <= fin_hoy
                ).scalar() or 0.0
                comision_dia = float(comision_dia_query)
        except (ValueError, TypeError) as e:
            print(f"Error al procesar sucursal_id: {e}")
            comision_dia = 0.0
            sucursal_nombre = None
    
    return render_template('operaciones.html',
                         operaciones=operaciones_paginated.items,
                         pagination=operaciones_paginated,
                         fecha_actual=fecha or datetime.now(peru_tz).strftime('%Y-%m-%d'),
                         fecha_hoy=datetime.now(peru_tz).strftime('%Y-%m-%d'),
                         filtros_aplicados=filtros_aplicados,
                         sucursales=sucursales,
                         medios_pago=medios_pago,
                         comision_dia=comision_dia,
                         sucursal_nombre=sucursal_nombre)

@app.route('/api/operaciones/lista', methods=['GET'])
@login_required
def api_operaciones_lista():
    """Devuelve lista JSON de operaciones del día para live updates en tabla."""
    try:
        ahora = get_peru_time()
        hoy = ahora.date()
        inicio_hoy = datetime.combine(hoy, datetime.min.time()).replace(tzinfo=peru_tz)
        fin_hoy = datetime.combine(hoy, datetime.max.time()).replace(tzinfo=peru_tz)
        fin_hoy = fin_hoy.replace(hour=23, minute=59, second=59, microsecond=999999)

        if current_user.es_admin:
            query = Operacion.query
        elif current_user.es_admin_de_sucursal():
            query = Operacion.query.filter_by(sucursal_id=current_user.sucursal_id)
        else:
            query = Operacion.query.filter_by(usuario_id=current_user.id)

        operaciones = query.filter(
            Operacion.hora >= inicio_hoy,
            Operacion.hora <= fin_hoy
        ).order_by(Operacion.hora.desc()).all()

        return jsonify({
            'success': True,
            'operaciones': [{
                'id': op.id,
                'hora': format_peru_time(op.hora),
                'monto': float(op.monto),
                'comision': float(op.comision),
                'medio': op.medio,
                'usuario': op.usuario.nombre_completo or op.usuario.username,
                'sucursal': op.usuario.sucursal.nombre if op.usuario.sucursal else 'N/A'
            } for op in operaciones]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/operaciones/<int:operacion_id>/eliminar', methods=['POST'])
@login_required
def eliminar_operacion(operacion_id):
    """Eliminar operación; admin puede todas, usuario solo las suyas o de su sucursal."""
    operacion = Operacion.query.get_or_404(operacion_id)
    if not current_user.es_admin:
        # Restringir a sucursal y, si no es admin de sucursal, a su propio registro
        if operacion.sucursal_id != current_user.sucursal_id:
            flash('No tienes permisos para eliminar esta operación', 'error')
            return redirect(url_for('operaciones'))
        if not current_user.es_admin_de_sucursal() and operacion.usuario_id != current_user.id:
            flash('No tienes permisos para eliminar esta operación', 'error')
            return redirect(url_for('operaciones'))
    db.session.delete(operacion)
    db.session.commit()
    flash('Operación eliminada', 'success')
    return redirect(url_for('operaciones'))


def _corregir_autoincrement_operacion():
    """Asegura que el auto-increment esté correcto tras registrar una operación."""
    try:
        max_id = db.session.query(db.func.max(Operacion.id)).scalar() or 0
        if max_id > 0:
            next_id = max_id + 1
            if os.environ.get('DATABASE_URL'):
                # PostgreSQL
                db.session.execute(
                    db.text(f"SELECT setval(pg_get_serial_sequence('operacion', 'id'), {max_id}, true)")
                )
            else:
                # SQLite
                try:
                    db.session.execute(db.text("DELETE FROM sqlite_sequence WHERE name='operacion'"))
                except:
                    pass
                db.session.execute(
                    db.text(f"INSERT INTO sqlite_sequence (name, seq) VALUES ('operacion', {next_id})")
                )
            db.session.commit()
    except Exception as e:
        print(f"[WARN] No se pudo corregir auto-increment: {e}")
        try:
            db.session.rollback()
        except:
            pass


@app.route('/operaciones/registrar', methods=['GET', 'POST'])
@login_required
def registrar_operacion():
    if request.method == 'POST':
        monto = float(request.form['monto'])
        medio = request.form['medio']

        # Comisión: si comision_manual=true, usar el valor del form (override del operador).
        # Si no, recalcular la sugerida server-side (ignorar lo que mande el cliente).
        es_manual = (request.form.get('comision_manual') or '').lower() in ('true', '1', 'on', 'yes')
        comision_sugerida = calcular_comision_sugerida(monto)
        if es_manual:
            try:
                comision = float(request.form.get('comision') or comision_sugerida)
            except (TypeError, ValueError):
                comision = comision_sugerida
            if comision < 0:
                comision = 0.0
            motivo = (request.form.get('motivo_descuento') or '').strip() or None
            # Si terminó siendo igual a la sugerida, no la marcamos como manual
            if abs(comision - comision_sugerida) < 0.001:
                es_manual = False
                motivo = None
        else:
            comision = comision_sugerida
            motivo = None

        # Determinar sucursal
        if current_user.es_admin:
            sucursal_id = int(request.form.get('sucursal_id'))
        else:
            sucursal_id = current_user.sucursal_id

        # Crear operación con hora Perú (wall-clock naive, sin tzinfo)
        hora_peru = get_peru_time().replace(tzinfo=None)
        operacion = Operacion(
            monto=monto,
            comision=comision,
            medio=medio,
            usuario_id=current_user.id,
            sucursal_id=sucursal_id,
            hora=hora_peru,
            comision_sugerida=comision_sugerida,
            comision_manual=es_manual,
            motivo_descuento=motivo,
        )

        db.session.add(operacion)
        
        # Actualizar comisiones
        hoy = get_peru_time().date()
        comision_diaria = ComisionDiaria.query.filter_by(
            fecha=hoy,
            sucursal_id=sucursal_id
        ).first()
        
        if comision_diaria:
            comision_diaria.total_comision = float(comision_diaria.total_comision) + comision
        else:
            comision_diaria = ComisionDiaria(
                fecha=hoy,
                sucursal_id=sucursal_id,
                total_comision=comision
            )
            db.session.add(comision_diaria)
        
        # Actualizar comisión mensual
        ahora = get_peru_time()
        año_actual = ahora.year
        mes_actual = ahora.month
        
        comision_mensual = ComisionMensual.query.filter_by(
            año=año_actual,
            mes=mes_actual,
            sucursal_id=sucursal_id
        ).first()
        
        if comision_mensual:
            comision_mensual.total_comision = float(comision_mensual.total_comision) + comision
        else:
            comision_mensual = ComisionMensual(
                año=año_actual,
                mes=mes_actual,
                sucursal_id=sucursal_id,
                total_comision=comision
            )
            db.session.add(comision_mensual)
        
        db.session.commit()

        # Corregir auto-increment si es necesario (evita IDs saltados)
        _corregir_autoincrement_operacion()

        # Limpiar caché después de cambios
        clear_cache()

        flash('Operación bancaria registrada exitosamente', 'success')
        return redirect(url_for('operaciones'))
    
    # OPTIMIZACIÓN ULTRA: Cargar sucursales solo si es admin
    sucursales = []
    if current_user.es_admin:
        sucursales = get_sucursales_activas_cache()

    # Cargar medios habilitados para la sucursal del usuario (o todos si admin)
    if current_user.es_admin:
        medios_pago = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    else:
        medios_pago = MedioPago.query.join(
            MedioSucursal,
            (MedioSucursal.medio_pago_id == MedioPago.id) &
            (MedioSucursal.sucursal_id == current_user.sucursal_id) &
            (MedioSucursal.activo == True)
        ).filter(MedioPago.activo == True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
        if not medios_pago:
            medios_pago = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()

    # Constantes para que el JS conozca la fórmula sin duplicarla
    return render_template('registrar_operacion.html',
                           sucursales=sucursales,
                           medios_pago=medios_pago,
                           comision_rango=COMISION_RANGO_SOLES,
                           comision_por_rango=COMISION_POR_RANGO)

@app.route('/operaciones/<int:operacion_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_operacion(operacion_id):
    operacion = Operacion.query.get_or_404(operacion_id)
    
    if not current_user.es_admin and operacion.usuario_id != current_user.id:
        flash('No tienes permisos para editar esta operación', 'error')
        return redirect(url_for('operaciones'))
    
    if request.method == 'POST':
        monto_anterior = float(operacion.comision)
        sucursal_anterior = operacion.sucursal_id
        monto = float(request.form['monto'])
        medio = request.form['medio']

        # Misma lógica que registrar: comisión auto vs manual
        es_manual = (request.form.get('comision_manual') or '').lower() in ('true', '1', 'on', 'yes')
        comision_sugerida = calcular_comision_sugerida(monto)
        if es_manual:
            try:
                comision = float(request.form.get('comision') or comision_sugerida)
            except (TypeError, ValueError):
                comision = comision_sugerida
            if comision < 0:
                comision = 0.0
            motivo = (request.form.get('motivo_descuento') or '').strip() or None
            if abs(comision - comision_sugerida) < 0.001:
                es_manual = False
                motivo = None
        else:
            comision = comision_sugerida
            motivo = None

        if current_user.es_admin:
            nueva_sucursal_id = int(request.form.get('sucursal_id'))
        else:
            nueva_sucursal_id = operacion.sucursal_id

        # Actualizar operación
        operacion.monto = monto
        operacion.comision = comision
        operacion.medio = medio
        operacion.sucursal_id = nueva_sucursal_id
        operacion.comision_sugerida = comision_sugerida
        operacion.comision_manual = es_manual
        operacion.motivo_descuento = motivo

        # Actualizar comisiones (lógica simplificada)
        db.session.commit()
        
        # Limpiar caché después de cambios
        clear_cache()
        
        flash('Operación actualizada exitosamente', 'success')
        return redirect(url_for('operaciones'))

    # GET: cargar medios habilitados para la sucursal de la operación
    medios_pago = MedioPago.query.join(
        MedioSucursal,
        (MedioSucursal.medio_pago_id == MedioPago.id) &
        (MedioSucursal.sucursal_id == operacion.sucursal_id) &
        (MedioSucursal.activo == True)
    ).filter(MedioPago.activo == True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    if not medios_pago:
        # Si no hay restricciones por sucursal aún, traer todos los activos
        medios_pago = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    return render_template('editar_operacion.html',
                           operacion=operacion,
                           medios_pago=medios_pago,
                           comision_rango=COMISION_RANGO_SOLES,
                           comision_por_rango=COMISION_POR_RANGO)


@app.route('/api/sucursal/<int:sucursal_id>/medios')
@login_required
def api_medios_por_sucursal(sucursal_id):
    """Devuelve los medios de pago habilitados para una sucursal específica."""
    medios = MedioPago.query.join(
        MedioSucursal,
        (MedioSucursal.medio_pago_id == MedioPago.id) &
        (MedioSucursal.sucursal_id == sucursal_id) &
        (MedioSucursal.activo == True)
    ).filter(MedioPago.activo == True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    return {'medios': [{'value': m.nombre_abreviado, 'label': m.nombre_abreviado} for m in medios]}

@app.route('/api/operaciones/<int:operacion_id>', methods=['PUT'])
@login_required
def api_actualizar_operacion(operacion_id):
    """API para actualizar operaciones (edición inline)"""
    try:
        # Obtener la operación
        operacion = Operacion.query.get_or_404(operacion_id)
        
        # Verificar permisos
        if not current_user.es_admin:
            # Si no es admin, verificar que sea el dueño o admin de sucursal
            if operacion.usuario_id != current_user.id:
                if not current_user.es_admin_de_sucursal() or operacion.sucursal_id != current_user.sucursal_id:
                    return jsonify({'success': False, 'message': 'No tienes permisos para editar esta operación'}), 403
        
        # Obtener datos del JSON
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        monto = data.get('monto')
        comision = data.get('comision')
        medio = data.get('medio')
        
        # Validar datos
        if monto is None or comision is None or not medio:
            return jsonify({'success': False, 'message': 'Todos los campos son requeridos'}), 400
        
        try:
            monto = float(monto)
            comision = float(comision)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Monto y comisión deben ser números válidos'}), 400
        
        if monto <= 0:
            return jsonify({'success': False, 'message': 'El monto debe ser mayor a 0'}), 400
        
        if comision < 0:
            return jsonify({'success': False, 'message': 'La comisión no puede ser negativa'}), 400
        
        # Validar medio de pago
        medio_valido = MedioPago.query.filter_by(nombre_abreviado=medio, activo=True).first()
        if not medio_valido:
            return jsonify({'success': False, 'message': 'Medio de pago no válido'}), 400
        
        # Actualizar operación
        operacion.monto = monto
        operacion.comision = comision
        operacion.medio = medio
        
        db.session.commit()
        
        # Limpiar caché después de cambios
        clear_cache()
        
        return jsonify({
            'success': True,
            'message': 'Operación actualizada exitosamente',
            'monto': float(monto),
            'comision': float(comision),
            'medio': medio
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        try:
            db.session.rollback()
        except:
            pass
        return jsonify({'success': False, 'message': f'Error al actualizar la operación: {str(e)}'}), 500
    
    # OPTIMIZACIÓN ULTRA: Cargar sucursales solo si es admin
    sucursales = []
    if current_user.es_admin:
        sucursales = get_sucursales_activas_cache()
    
    return render_template('editar_operacion.html', operacion=operacion, sucursales=sucursales)

# ========== MÓDULO DE INVENTARIO ==========

# Tipos de imagen aceptados y peso máximo para la foto de un producto.
# La foto se guarda en la base de datos (no en disco) porque en Railway el
# filesystem es efímero y un archivo subido se perdería en cada redeploy.
FOTO_TIPOS_PERMITIDOS = {'image/png', 'image/jpeg', 'image/jpg', 'image/gif', 'image/webp'}
FOTO_TAMANO_MAXIMO = 2 * 1024 * 1024  # 2 MB

def leer_foto_producto(file_storage):
    """Lee y valida la foto subida para un producto.

    Devuelve (bytes, mimetype). Si no se subió ninguna foto devuelve (None, None).
    Lanza ValueError con un mensaje amigable si el archivo no es válido.
    """
    if not file_storage or not file_storage.filename:
        return None, None

    mimetype = (file_storage.mimetype or '').lower()
    if mimetype not in FOTO_TIPOS_PERMITIDOS:
        raise ValueError('Formato de imagen no permitido (usa PNG, JPG, GIF o WEBP)')

    datos = file_storage.read()
    if not datos:
        return None, None
    if len(datos) > FOTO_TAMANO_MAXIMO:
        raise ValueError('La imagen pesa demasiado (máximo 2 MB)')

    return datos, mimetype


def sucursales_visibles_para(usuario):
    """Sucursales que el usuario puede ver/gestionar en inventario y ventas."""
    if usuario.es_admin:
        return Sucursal.query.filter_by(activa=True).order_by(Sucursal.nombre).all()
    return [usuario.sucursal] if usuario.sucursal else []


@app.route('/inventario')
@login_required
def inventario():
    """Listado de productos en inventario (stock por sucursal)"""
    sucursales = sucursales_visibles_para(current_user)

    if current_user.es_admin:
        sucursal_id = request.args.get('sucursal_id', type=int)
        query = Producto.query.filter_by(activo=True)
        if sucursal_id:
            query = query.filter_by(sucursal_id=sucursal_id)
    else:
        sucursal_id = current_user.sucursal_id
        query = Producto.query.filter_by(activo=True, sucursal_id=current_user.sucursal_id)

    productos = query.order_by(Producto.nombre).all()
    puede_gestionar = current_user.es_admin_o_admin_sucursal()

    return render_template(
        'inventario.html',
        productos=productos,
        sucursales=sucursales,
        sucursal_id=sucursal_id,
        puede_gestionar=puede_gestionar
    )


@app.route('/inventario/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_producto():
    """Agregar un producto al inventario (solo administradores / admins de sucursal)"""
    if not current_user.es_admin_o_admin_sucursal():
        flash('No tienes permisos para gestionar el inventario', 'error')
        return redirect(url_for('inventario'))

    sucursales = sucursales_visibles_para(current_user)

    if request.method == 'POST':
        try:
            nombre = request.form.get('nombre', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            precio = request.form.get('precio', type=float)
            stock = request.form.get('stock', type=int)

            if not nombre:
                flash('Debe ingresar el nombre del producto', 'error')
                return redirect(url_for('nuevo_producto'))
            if precio is None or precio <= 0:
                flash('El precio debe ser mayor a 0', 'error')
                return redirect(url_for('nuevo_producto'))
            if stock is None or stock < 0:
                flash('El stock no puede ser negativo', 'error')
                return redirect(url_for('nuevo_producto'))

            if current_user.es_admin:
                sucursal_id = request.form.get('sucursal_id', type=int)
                if not sucursal_id:
                    flash('Debe seleccionar una sucursal', 'error')
                    return redirect(url_for('nuevo_producto'))
            else:
                sucursal_id = current_user.sucursal_id

            try:
                foto, foto_mimetype = leer_foto_producto(request.files.get('foto'))
            except ValueError as e:
                flash(str(e), 'error')
                return redirect(url_for('nuevo_producto'))

            producto = Producto(
                nombre=nombre,
                descripcion=descripcion,
                precio=precio,
                stock=stock,
                foto=foto,
                foto_mimetype=foto_mimetype,
                sucursal_id=sucursal_id
            )
            db.session.add(producto)
            db.session.commit()
            flash(f'Producto "{nombre}" agregado al inventario', 'success')
            return redirect(url_for('inventario'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar el producto: {str(e)}', 'error')
            return redirect(url_for('nuevo_producto'))

    return render_template('producto_form.html', sucursales=sucursales, producto=None)


@app.route('/inventario/<int:producto_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_producto(producto_id):
    """Editar un producto del inventario (solo administradores / admins de sucursal)"""
    if not current_user.es_admin_o_admin_sucursal():
        flash('No tienes permisos para gestionar el inventario', 'error')
        return redirect(url_for('inventario'))

    producto = Producto.query.get_or_404(producto_id)
    if not current_user.es_admin and producto.sucursal_id != current_user.sucursal_id:
        flash('No tienes permisos para editar este producto', 'error')
        return redirect(url_for('inventario'))

    sucursales = sucursales_visibles_para(current_user)

    if request.method == 'POST':
        try:
            nombre = request.form.get('nombre', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            precio = request.form.get('precio', type=float)
            stock = request.form.get('stock', type=int)

            if not nombre:
                flash('Debe ingresar el nombre del producto', 'error')
                return redirect(url_for('editar_producto', producto_id=producto_id))
            if precio is None or precio <= 0:
                flash('El precio debe ser mayor a 0', 'error')
                return redirect(url_for('editar_producto', producto_id=producto_id))
            if stock is None or stock < 0:
                flash('El stock no puede ser negativo', 'error')
                return redirect(url_for('editar_producto', producto_id=producto_id))

            try:
                foto, foto_mimetype = leer_foto_producto(request.files.get('foto'))
            except ValueError as e:
                flash(str(e), 'error')
                return redirect(url_for('editar_producto', producto_id=producto_id))

            producto.nombre = nombre
            producto.descripcion = descripcion
            producto.precio = precio
            producto.stock = stock
            if foto is not None:
                producto.foto = foto
                producto.foto_mimetype = foto_mimetype

            db.session.commit()
            flash(f'Producto "{nombre}" actualizado', 'success')
            return redirect(url_for('inventario'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el producto: {str(e)}', 'error')
            return redirect(url_for('editar_producto', producto_id=producto_id))

    return render_template('producto_form.html', sucursales=sucursales, producto=producto)


@app.route('/inventario/<int:producto_id>/eliminar', methods=['POST'])
@login_required
def eliminar_producto(producto_id):
    """Quitar (desactivar) un producto del inventario"""
    if not current_user.es_admin_o_admin_sucursal():
        flash('No tienes permisos para gestionar el inventario', 'error')
        return redirect(url_for('inventario'))

    producto = Producto.query.get_or_404(producto_id)
    if not current_user.es_admin and producto.sucursal_id != current_user.sucursal_id:
        flash('No tienes permisos para eliminar este producto', 'error')
        return redirect(url_for('inventario'))

    producto.activo = False
    db.session.commit()
    flash(f'Producto "{producto.nombre}" eliminado del inventario', 'success')
    return redirect(url_for('inventario'))


@app.route('/inventario/foto/<int:producto_id>')
@login_required
def foto_producto(producto_id):
    """Sirve la foto de un producto guardada en la base de datos"""
    producto = Producto.query.get_or_404(producto_id)
    if not producto.foto:
        abort(404)
    respuesta = Response(producto.foto, mimetype=producto.foto_mimetype or 'image/jpeg')
    respuesta.headers['Cache-Control'] = 'private, max-age=3600'
    return respuesta


# ========== MÓDULO DE VENTAS ==========

@app.route('/ventas')
@login_required
def ventas():
    """Listado de ventas del día.

    Visibilidad (igual que en Operaciones):
      - Admin global: todas las ventas (puede filtrar por sucursal)
      - Admin de sucursal: las ventas de su sucursal
      - Usuario normal: solo SUS PROPIAS ventas
    """
    sucursales = sucursales_visibles_para(current_user)

    if current_user.es_admin:
        sucursal_id = request.args.get('sucursal_id', type=int)
        query = Venta.query
        if sucursal_id:
            query = query.filter(Venta.sucursal_id == sucursal_id)
    elif current_user.es_admin_de_sucursal():
        sucursal_id = current_user.sucursal_id
        query = Venta.query.filter(Venta.sucursal_id == current_user.sucursal_id)
    else:
        sucursal_id = current_user.sucursal_id
        query = Venta.query.filter(Venta.usuario_id == current_user.id)

    # Solo las ventas de hoy (hora de Perú)
    hoy = get_peru_time().date()
    query = query.filter(db.func.date(Venta.fecha) == hoy)

    ventas_hoy = query.order_by(Venta.fecha.desc()).all()
    total_ventas = sum(float(v.monto) for v in ventas_hoy)

    return render_template(
        'ventas.html',
        ventas=ventas_hoy,
        total_ventas=total_ventas,
        sucursales=sucursales,
        sucursal_id=sucursal_id
    )

@app.route('/ventas/registrar', methods=['GET', 'POST'])
@login_required
def registrar_venta():
    """Registrar una nueva venta eligiendo un producto del inventario.

    El monto se calcula como precio_unitario * cantidad y se descuenta
    automáticamente del stock del producto en esa sucursal.
    """
    if request.method == 'POST':
        try:
            producto_id = request.form.get('producto_id', type=int)
            cantidad = request.form.get('cantidad', type=int)

            if not producto_id:
                flash('Debe seleccionar un producto', 'error')
                return redirect(url_for('registrar_venta'))

            if not cantidad or cantidad <= 0:
                flash('La cantidad debe ser mayor a 0', 'error')
                return redirect(url_for('registrar_venta'))

            producto = Producto.query.get(producto_id)
            if not producto or not producto.activo:
                flash('El producto seleccionado no es válido', 'error')
                return redirect(url_for('registrar_venta'))

            # Determinar sucursal de la venta
            if current_user.es_admin:
                sucursal_id = request.form.get('sucursal_id', type=int)
                if not sucursal_id:
                    flash('Debe seleccionar una sucursal', 'error')
                    return redirect(url_for('registrar_venta'))
            else:
                sucursal_id = current_user.sucursal_id

            if producto.sucursal_id != sucursal_id:
                flash('Ese producto no pertenece al inventario de la sucursal seleccionada', 'error')
                return redirect(url_for('registrar_venta'))

            if cantidad > producto.stock:
                flash(f'Stock insuficiente: quedan {producto.stock} unidad(es) de "{producto.nombre}"', 'error')
                return redirect(url_for('registrar_venta'))

            precio_unitario = float(producto.precio)
            monto = precio_unitario * cantidad

            # Crear venta (snapshot del precio al momento de la venta)
            venta = Venta(
                producto_id=producto.id,
                cantidad=cantidad,
                precio_unitario=precio_unitario,
                monto=monto,
                usuario_id=current_user.id,
                sucursal_id=sucursal_id,
                fecha=get_peru_time().replace(tzinfo=None)
            )
            db.session.add(venta)

            # Descontar stock del inventario
            producto.stock = producto.stock - cantidad

            # Actualizar caja de ventas del día
            hoy = get_peru_time().date()
            caja = CajaVentas.query.filter_by(
                fecha=hoy,
                sucursal_id=sucursal_id
            ).first()

            if caja:
                caja.total_vendido = float(caja.total_vendido) + monto
                caja.saldo = float(caja.saldo) + monto
            else:
                caja = CajaVentas(
                    fecha=hoy,
                    sucursal_id=sucursal_id,
                    total_vendido=monto,
                    saldo=monto
                )
                db.session.add(caja)

            db.session.commit()
            flash(f'Venta registrada: {cantidad} x "{producto.nombre}" = S/ {monto:.2f}', 'success')
            return redirect(url_for('ventas'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar venta: {str(e)}', 'error')
            return redirect(url_for('registrar_venta'))

    # GET: mostrar formulario
    sucursales = sucursales_visibles_para(current_user)
    if current_user.es_admin:
        productos = Producto.query.filter_by(activo=True).order_by(Producto.nombre).all()
    else:
        productos = Producto.query.filter_by(activo=True, sucursal_id=current_user.sucursal_id).order_by(Producto.nombre).all()

    return render_template('registrar_venta.html', sucursales=sucursales, productos=productos)


@app.route('/ventas/<int:venta_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_venta(venta_id):
    """Editar una venta ya registrada.

    Solo puede editarla quien la registró (o el administrador global), igual
    que con las operaciones. Al cambiar de producto/cantidad se recalculan el
    stock del/los producto(s) involucrados y el total de la caja de ventas del día.
    """
    venta = Venta.query.get_or_404(venta_id)

    if not current_user.es_admin and venta.usuario_id != current_user.id:
        flash('Solo puedes editar las ventas que tú mismo registraste', 'error')
        return redirect(url_for('ventas'))

    productos = Producto.query.filter_by(sucursal_id=venta.sucursal_id, activo=True).order_by(Producto.nombre).all()
    # Asegurar que el producto actual de la venta aparezca en la lista aunque
    # ya no esté activo (para no romper el formulario de edición)
    if venta.producto and venta.producto not in productos:
        productos = [venta.producto] + productos

    if request.method == 'POST':
        try:
            producto_id = request.form.get('producto_id', type=int)
            cantidad = request.form.get('cantidad', type=int)

            if not producto_id:
                flash('Debe seleccionar un producto', 'error')
                return redirect(url_for('editar_venta', venta_id=venta_id))

            if not cantidad or cantidad <= 0:
                flash('La cantidad debe ser mayor a 0', 'error')
                return redirect(url_for('editar_venta', venta_id=venta_id))

            nuevo_producto_obj = Producto.query.get(producto_id)
            if not nuevo_producto_obj or nuevo_producto_obj.sucursal_id != venta.sucursal_id:
                flash('El producto seleccionado no es válido para esta venta', 'error')
                return redirect(url_for('editar_venta', venta_id=venta_id))

            producto_anterior = venta.producto
            cantidad_anterior = venta.cantidad
            monto_anterior = float(venta.monto)

            # Cuánto stock queda "disponible" si liberamos lo que ya estaba
            # comprometido por la venta original
            if nuevo_producto_obj.id == producto_anterior.id:
                stock_disponible = nuevo_producto_obj.stock + cantidad_anterior
            else:
                stock_disponible = nuevo_producto_obj.stock

            if cantidad > stock_disponible:
                flash(f'Stock insuficiente: solo hay {stock_disponible} unidad(es) disponibles de "{nuevo_producto_obj.nombre}"', 'error')
                return redirect(url_for('editar_venta', venta_id=venta_id))

            # Revertir el stock comprometido por la venta original y aplicar el nuevo
            producto_anterior.stock = producto_anterior.stock + cantidad_anterior
            nuevo_producto_obj.stock = nuevo_producto_obj.stock - cantidad

            nuevo_precio_unitario = float(nuevo_producto_obj.precio)
            nuevo_monto = nuevo_precio_unitario * cantidad

            # Ajustar la caja de ventas del día de la venta original
            # (revertir el monto anterior y aplicar el nuevo)
            fecha_caja = venta.fecha.date()
            caja = CajaVentas.query.filter_by(fecha=fecha_caja, sucursal_id=venta.sucursal_id).first()
            if caja:
                caja.total_vendido = float(caja.total_vendido) - monto_anterior + nuevo_monto
                caja.saldo = float(caja.saldo) - monto_anterior + nuevo_monto
            else:
                caja = CajaVentas(
                    fecha=fecha_caja,
                    sucursal_id=venta.sucursal_id,
                    total_vendido=nuevo_monto,
                    saldo=nuevo_monto
                )
                db.session.add(caja)

            venta.producto_id = nuevo_producto_obj.id
            venta.cantidad = cantidad
            venta.precio_unitario = nuevo_precio_unitario
            venta.monto = nuevo_monto

            db.session.commit()
            flash('Venta actualizada correctamente', 'success')
            return redirect(url_for('ventas'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar la venta: {str(e)}', 'error')
            return redirect(url_for('editar_venta', venta_id=venta_id))

    return render_template('editar_venta.html', venta=venta, productos=productos)


@app.route('/api/ventas/resumen')
@login_required
def api_ventas_resumen():
    """API para obtener resumen de ventas del día"""
    from datetime import date
    hoy = get_peru_time().date()

    sucursal_id = request.args.get('sucursal_id', type=int)

    if current_user.es_admin and sucursal_id:
        caja = CajaVentas.query.filter_by(
            fecha=hoy,
            sucursal_id=sucursal_id
        ).first()
    elif not current_user.es_admin:
        caja = CajaVentas.query.filter_by(
            fecha=hoy,
            sucursal_id=current_user.sucursal_id
        ).first()
    else:
        return jsonify({'error': 'Sucursal no especificada'}), 400

    if caja:
        return jsonify({
            'total_vendido': float(caja.total_vendido),
            'saldo': float(caja.saldo)
        })

    return jsonify({
        'total_vendido': 0.0,
        'saldo': 0.0
    })

# Rutas de administración optimizadas
@app.route('/admin/usuarios')
@login_required
def admin_usuarios():
    if not current_user.es_admin_o_admin_sucursal():
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    
    # OPTIMIZACIÓN ULTRA: Paginación para usuarios
    page = request.args.get('page', 1, type=int)
    query = Usuario.query
    
    # Si es admin de sucursal, filtrar solo usuarios de su sucursal
    if current_user.es_admin_de_sucursal() and not current_user.es_admin:
        query = query.filter_by(sucursal_id=current_user.sucursal_id)
    
    usuarios_paginated = query.paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('admin_usuarios.html', 
                         usuarios=usuarios_paginated.items,
                         pagination=usuarios_paginated)

@app.route('/admin/usuarios/crear', methods=['GET', 'POST'])
@login_required
def crear_usuario():
    if not current_user.es_admin_o_admin_sucursal():
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        nombre_completo = request.form.get('nombre_completo', '').strip()
        
        # Si es admin de sucursal, asignar automáticamente a su sucursal
        if current_user.es_admin_de_sucursal() and not current_user.es_admin:
            sucursal_id = current_user.sucursal_id
            es_admin = False
            es_admin_sucursal = False
        else:
            # Admin global puede asignar sucursal y roles
            sucursal_id = request.form.get('sucursal_id')
            sucursal_id = int(sucursal_id) if sucursal_id else None
            es_admin = 'es_admin' in request.form
            es_admin_sucursal = 'es_admin_sucursal' in request.form
        
        # Verificar si el usuario ya existe
        if Usuario.query.filter_by(username=username).first():
            flash('El nombre de usuario ya existe', 'error')
            sucursales = Sucursal.query.filter_by(activa=True).all() if current_user.es_admin else []
            return render_template('crear_usuario.html', sucursales=sucursales)
        
        # Generar email por defecto si no existe (usar username como base)
        email_default = f"{username}@sisagent.local"
        
        nuevo_usuario = Usuario(
            username=username,
            email=email_default,  # Email por defecto basado en username
            password_hash=generate_password_hash(password),
            nombre_completo=nombre_completo,
            sucursal_id=sucursal_id,
            es_admin=es_admin,
            es_admin_sucursal=es_admin_sucursal
        )
        
        db.session.add(nuevo_usuario)
        db.session.commit()
        flash('Usuario creado exitosamente', 'success')
        return redirect(url_for('admin_usuarios'))
    
    # Obtener sucursales disponibles
    if current_user.es_admin:
        sucursales = Sucursal.query.filter_by(activa=True).all()
    else:
        # Admin de sucursal solo puede ver su sucursal
        sucursales = [current_user.sucursal] if current_user.sucursal else []
    
    return render_template('crear_usuario.html', sucursales=sucursales)

@app.route('/admin/usuarios/<int:usuario_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_usuario(usuario_id):
    if not current_user.es_admin_o_admin_sucursal():
        flash('Acceso denegado', 'error')
        return redirect(url_for('admin_usuarios'))
    
    usuario = Usuario.query.get_or_404(usuario_id)
    
    # Si es admin de sucursal, solo puede editar usuarios de su sucursal
    if current_user.es_admin_de_sucursal() and not current_user.es_admin:
        if usuario.sucursal_id != current_user.sucursal_id:
            flash('No tienes permisos para editar este usuario', 'error')
            return redirect(url_for('admin_usuarios'))
    
    if request.method == 'POST':
        nuevo_username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        nombre_completo = request.form.get('nombre_completo', '').strip()
        
        if nuevo_username:
            usuario.username = nuevo_username
        if password:
            usuario.password_hash = generate_password_hash(password)
            # Regenerar token → cierra todas las sesiones activas de este usuario
            usuario.regenerate_session_token()
        if nombre_completo:
            usuario.nombre_completo = nombre_completo
        
        # Asegurar que el usuario tenga email (generar si no existe)
        if not usuario.email:
            usuario.email = f"{usuario.username}@sisagent.local"
        
        # Solo admin global puede cambiar sucursal y roles
        if current_user.es_admin:
            sucursal_id = request.form.get('sucursal_id')
            if sucursal_id:
                usuario.sucursal_id = int(sucursal_id)
            usuario.es_admin = 'es_admin' in request.form
            usuario.es_admin_sucursal = 'es_admin_sucursal' in request.form
        else:
            # Admin de sucursal solo puede asegurar que el usuario esté en su sucursal
            usuario.sucursal_id = current_user.sucursal_id
        
        db.session.commit()
        flash('Usuario actualizado', 'success')
        return redirect(url_for('admin_usuarios'))
    
    sucursales = Sucursal.query.filter_by(activa=True).all() if current_user.es_admin else [current_user.sucursal] if current_user.sucursal else []
    return render_template('editar_usuario.html', usuario=usuario, sucursales=sucursales)

# Rutas mínimas para medios de pago (compatibilidad)
@app.route('/admin/medios')
@login_required
def admin_medios():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    medios = MedioPago.query.order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    return render_template('admin_medios.html', medios=medios)

@app.route('/admin/medios/<int:medio_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_medio(medio_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    medio = MedioPago.query.get_or_404(medio_id)
    if request.method == 'POST':
        abreviado = request.form.get('nombre_abreviado', '').strip()
        completo = request.form.get('nombre_completo', '').strip()
        if abreviado and completo:
            medio.nombre_abreviado = abreviado
            medio.nombre_completo = completo
            db.session.commit()
            flash('Medio actualizado', 'success')
            return redirect(url_for('admin_medios'))
    return render_template('editar_medio.html', medio=medio)

@app.route('/admin/medios/<int:medio_id>/eliminar', methods=['POST'])
@login_required
def eliminar_medio(medio_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    medio = MedioPago.query.get_or_404(medio_id)
    db.session.delete(medio)
    db.session.commit()
    flash('Medio eliminado', 'success')
    return redirect(url_for('admin_medios'))

@app.route('/admin/medios/<int:medio_id>/activar', methods=['POST'])
@login_required
def activar_medio(medio_id):
    if not current_user.es_admin:
        return {'ok': False, 'error': 'Acceso denegado'}, 403
    medio = MedioPago.query.get_or_404(medio_id)
    medio.activo = not medio.activo
    db.session.commit()
    cache.clear()
    # Soporta tanto AJAX como form POST normal
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
        return {'ok': True, 'activo': medio.activo}
    return redirect(url_for('admin_medios'))

@app.route('/admin/medios/<int:medio_id>/subir', methods=['POST'])
@login_required
def subir_medio(medio_id):
    # Mantenido por compatibilidad — reordenamiento via API /api/medios/orden
    return redirect(url_for('admin_medios'))

@app.route('/admin/medios/<int:medio_id>/bajar', methods=['POST'])
@login_required
def bajar_medio(medio_id):
    return redirect(url_for('admin_medios'))

@app.route('/api/medios/orden', methods=['POST'])
@login_required
def api_medios_orden():
    """Guarda el nuevo orden de medios de pago (drag-and-drop)."""
    if not current_user.es_admin:
        return {'ok': False, 'error': 'Acceso denegado'}, 403
    data = request.get_json()
    ids = data.get('ids', [])
    for i, mid in enumerate(ids):
        m = MedioPago.query.get(mid)
        if m:
            m.orden = i
    db.session.commit()
    cache.clear()
    return {'ok': True}

@app.route('/api/medios/<int:medio_id>/editar', methods=['POST'])
@login_required
def api_editar_medio(medio_id):
    """Edita inline nombre abreviado y completo de un medio."""
    if not current_user.es_admin:
        return {'ok': False, 'error': 'Acceso denegado'}, 403
    medio = MedioPago.query.get_or_404(medio_id)
    data = request.get_json()
    abreviado = (data.get('nombre_abreviado') or '').strip().upper()
    completo  = (data.get('nombre_completo')  or '').strip()
    if not abreviado or not completo:
        return {'ok': False, 'error': 'Datos incompletos'}, 400
    medio.nombre_abreviado = abreviado
    medio.nombre_completo  = completo
    db.session.commit()
    cache.clear()
    return {'ok': True, 'abreviado': medio.nombre_abreviado, 'completo': medio.nombre_completo}

@app.route('/admin/usuarios/<int:usuario_id>/eliminar', methods=['POST'])
@login_required
def eliminar_usuario(usuario_id):
    """Eliminar usuario: solo admin global; protege admin/admin1."""
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('admin_usuarios'))
    usuario = Usuario.query.get_or_404(usuario_id)
    if usuario.username in ('admin', 'admin1'):
        flash('No se puede eliminar este usuario', 'warning')
        return redirect(url_for('admin_usuarios'))
    db.session.delete(usuario)
    db.session.commit()
    flash('Usuario eliminado', 'success')
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuarios/<int:usuario_id>/cerrar-sesion', methods=['POST'])
@login_required
def forzar_cierre_sesion(usuario_id):
    """Admin cierra la sesión activa de un usuario específico regenerando su token."""
    if not current_user.es_admin_o_admin_sucursal():
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'error': 'Acceso denegado'}), 403
        flash('Acceso denegado', 'error')
        return redirect(url_for('admin_usuarios'))

    usuario = Usuario.query.get_or_404(usuario_id)

    # No se puede cerrar la sesión del propio usuario admin actual
    if usuario.id == current_user.id:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'error': 'No puedes cerrar tu propia sesión desde aquí'}), 400
        flash('No puedes cerrar tu propia sesión desde aquí', 'warning')
        return redirect(url_for('admin_usuarios'))

    # Admin de sucursal solo puede afectar usuarios de su sucursal
    if current_user.es_admin_de_sucursal() and not current_user.es_admin:
        if usuario.sucursal_id != current_user.sucursal_id:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'ok': False, 'error': 'Sin permisos sobre este usuario'}), 403
            flash('Sin permisos sobre este usuario', 'error')
            return redirect(url_for('admin_usuarios'))

    usuario.regenerate_session_token()
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True, 'mensaje': f'Sesión de {usuario.username} cerrada'})
    flash(f'Sesión de {usuario.username} cerrada correctamente', 'success')
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/sucursales')
@login_required
def admin_sucursales():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Consulta segura de sucursales
        sucursales = Sucursal.query.filter_by(activa=True).all()
        
        # Pre-calcular conteos de forma segura
        for s in sucursales:
            try:
                s._usuarios_count = db.session.query(db.func.count(Usuario.id)).filter(
                    Usuario.sucursal_id == s.id
                ).scalar() or 0
            except Exception as e:
                print(f"Error contando usuarios para sucursal {s.id}: {e}")
                s._usuarios_count = 0
            
            try:
                s._operaciones_count = db.session.query(db.func.count(Operacion.id)).filter(
                    Operacion.sucursal_id == s.id
                ).scalar() or 0
            except Exception as e:
                print(f"Error contando operaciones para sucursal {s.id}: {e}")
                s._operaciones_count = 0
        
        return render_template('admin_sucursales.html', sucursales=sucursales)
    except Exception as e:
        import traceback
        error_msg = f"ERROR en admin_sucursales: {type(e).__name__}: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        try:
            db.session.rollback()
        except:
            pass
        flash(f'Error al cargar sucursales: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/admin/sucursales/crear', methods=['GET', 'POST'])
@login_required
def crear_sucursal():
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        direccion = request.form.get('direccion', '').strip()
        if not nombre:
            flash('El nombre es requerido', 'error')
            medios = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
            return render_template('crear_sucursal.html', medios=medios)
        if Sucursal.query.filter_by(nombre=nombre).first():
            flash('Ya existe una sucursal con ese nombre', 'warning')
            medios = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
            return render_template('crear_sucursal.html', medios=medios)
        suc = Sucursal(nombre=nombre, direccion=direccion, activa=True)
        db.session.add(suc)
        db.session.flush()  # Para obtener el ID de la sucursal
        
        # Manejar medios de pago seleccionados
        medios_seleccionados = request.form.getlist('medios')
        medios_seleccionados = [int(m) for m in medios_seleccionados if m]
        
        # Obtener todos los medios de pago activos
        todos_medios = MedioPago.query.filter_by(activo=True).all()
        
        # Crear relaciones con medios de pago seleccionados
        for medio in todos_medios:
            if medio.id in medios_seleccionados:
                medio_sucursal = MedioSucursal(
                    sucursal_id=suc.id,
                    medio_pago_id=medio.id,
                    activo=True
                )
                db.session.add(medio_sucursal)
        
        db.session.commit()
        flash('Sucursal creada', 'success')
        return redirect(url_for('admin_sucursales'))
    
    # Obtener todos los medios de pago activos para el formulario
    medios = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    return render_template('crear_sucursal.html', medios=medios)

@app.route('/admin/sucursales/<int:sucursal_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_sucursal(sucursal_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    suc = Sucursal.query.get_or_404(sucursal_id)
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        direccion = request.form.get('direccion', '').strip()
        activa = 'activa' in request.form
        
        if nombre:
            suc.nombre = nombre
        suc.direccion = direccion
        suc.activa = activa
        
        # Manejar medios de pago seleccionados
        medios_seleccionados = request.form.getlist('medios')
        medios_seleccionados = [int(m) for m in medios_seleccionados if m]
        
        # Obtener todos los medios de pago activos
        todos_medios = MedioPago.query.filter_by(activo=True).all()
        
        # Actualizar la relación con medios de pago
        for medio in todos_medios:
            medio_sucursal = MedioSucursal.query.filter_by(
                sucursal_id=suc.id,
                medio_pago_id=medio.id
            ).first()
            
            if medio.id in medios_seleccionados:
                # Debe estar activo
                if not medio_sucursal:
                    # Crear nueva relación
                    medio_sucursal = MedioSucursal(
                        sucursal_id=suc.id,
                        medio_pago_id=medio.id,
                        activo=True
                    )
                    db.session.add(medio_sucursal)
                else:
                    # Activar si estaba desactivado
                    medio_sucursal.activo = True
            else:
                # Debe estar desactivado
                if medio_sucursal:
                    medio_sucursal.activo = False
                # Si no existe, no crear (solo desactivar si existe)
        
        db.session.commit()
        flash('Sucursal actualizada', 'success')
        return redirect(url_for('admin_sucursales'))
    
    # Obtener medios de pago activos para esta sucursal
    medios_activos = db.session.query(MedioSucursal.medio_pago_id).filter_by(
        sucursal_id=suc.id,
        activo=True
    ).all()
    medios_activos_ids = [m[0] for m in medios_activos]
    
    # Obtener todos los medios de pago activos
    medios = MedioPago.query.filter_by(activo=True).order_by(MedioPago.orden, MedioPago.nombre_abreviado).all()
    
    return render_template('editar_sucursal.html', 
                         sucursal=suc, 
                         medios=medios, 
                         medios_activos=medios_activos_ids)

@app.route('/admin/sucursales/<int:sucursal_id>/eliminar', methods=['POST'])
@login_required
def eliminar_sucursal(sucursal_id):
    if not current_user.es_admin:
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    suc = Sucursal.query.get_or_404(sucursal_id)
    db.session.delete(suc)
    db.session.commit()
    flash('Sucursal eliminada', 'success')
    return redirect(url_for('admin_sucursales'))

@app.route('/admin/fix-operacion-sequence', methods=['POST'])
@login_required
def fix_operacion_sequence():
    """Resetear auto-increment de operaciones (admin only, emergencia)."""
    if not current_user.es_admin:
        return jsonify({'error': 'Acceso denegado'}), 403

    try:
        # Obtener el máximo ID actual
        max_id = db.session.query(db.func.max(Operacion.id)).scalar() or 0
        next_id = max_id + 1

        # Resetear auto-increment según el tipo de BD
        if os.environ.get('DATABASE_URL'):
            # PostgreSQL: setval(seq, N, true) hace que nextval() devuelva N+1
            db.session.execute(
                db.text(f"SELECT setval(pg_get_serial_sequence('operacion', 'id'), {max_id}, true)")
            )
        else:
            # SQLite: seq debe ser el PRÓXIMO valor a devolver
            db.session.execute(
                db.text(f"DELETE FROM sqlite_sequence WHERE name='operacion'")
            )
            db.session.execute(
                db.text(f"INSERT INTO sqlite_sequence (name, seq) VALUES ('operacion', {next_id})")
            )

        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Auto-increment reseteado. Próximo ID: {next_id}'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/reportes')
@login_required
def reportes():
    """Página de reportes para administradores."""
    if not current_user.es_admin_o_admin_sucursal():
        flash('Acceso denegado', 'error')
        return redirect(url_for('dashboard'))
    
    # Obtener sucursales y medios de pago
    if current_user.es_admin:
        sucursales = get_sucursales_activas_cache()
    else:
        # Admin de sucursal solo ve su sucursal
        sucursales = [current_user.sucursal] if current_user.sucursal else []
    
    medios_pago = get_medios_pago_cache()
    
    return render_template('reportes.html', sucursales=sucursales, medios_pago=medios_pago)

@app.route('/api/reportes/operaciones')
@login_required
def api_reportes_operaciones():
    """API para generar reportes de operaciones con filtros."""
    try:
        if not current_user.es_admin_o_admin_sucursal():
            return jsonify({'error': 'Acceso denegado'}), 403
        
        # Obtener parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        sucursal_id = request.args.get('sucursal_id')
        medio = request.args.get('medio')
        
        query = Operacion.query
        
        # Si es admin de sucursal, filtrar automáticamente por su sucursal (ignorar parámetro sucursal_id)
        if current_user.es_admin_de_sucursal() and not current_user.es_admin:
            if current_user.sucursal_id:
                query = query.filter(Operacion.sucursal_id == current_user.sucursal_id)
        elif sucursal_id and sucursal_id.strip():
            # Solo aplicar filtro de sucursal si es admin global
            try:
                sucursal_id_int = int(sucursal_id)
                query = query.filter(Operacion.sucursal_id == sucursal_id_int)
            except ValueError:
                pass  # Ignorar si no se puede convertir
        
        # Filtros de fecha: usar _fecha_rango para adaptarse al backend (SQLite vs PostgreSQL)
        if fecha_inicio:
            inicio = _fecha_rango(fecha_inicio, es_fin=False)
            if inicio:
                query = query.filter(Operacion.hora >= inicio)

        if fecha_fin:
            fin = _fecha_rango(fecha_fin, es_fin=True)
            if fin:
                query = query.filter(Operacion.hora <= fin)
        
        if medio:
            query = query.filter(Operacion.medio == medio)
        
        # Cache de medios de pago
        medios_cache = {mp.nombre_abreviado: mp.nombre_completo for mp in MedioPago.query.filter_by(activo=True).all()}
        
        # Obtener todas las operaciones sin límite (según filtros aplicados)
        # Usar joinedload para cargar relaciones de forma eficiente
        from sqlalchemy.orm import joinedload
        operaciones = query.options(
            joinedload(Operacion.usuario).load_only(Usuario.id, Usuario.username, Usuario.nombre_completo),
            joinedload(Operacion.sucursal).load_only(Sucursal.id, Sucursal.nombre)
        ).order_by(Operacion.hora.desc()).all()
        
        # Procesar datos
        datos = []
        total_monto = 0.0
        total_comision = 0.0
        
        for op in operaciones:
            monto = float(op.monto)
            comision = float(op.comision)
            total_monto += monto
            total_comision += comision
            
            # Obtener datos de usuario y sucursal de forma segura
            usuario_nombre = 'N/A'
            if hasattr(op, 'usuario') and op.usuario:
                usuario_nombre = op.usuario.nombre_completo if op.usuario.nombre_completo else op.usuario.username
            
            sucursal_nombre = 'Sin sucursal'
            if hasattr(op, 'sucursal') and op.sucursal:
                sucursal_nombre = op.sucursal.nombre
            
            datos.append({
                'id': op.id,
                'fecha': format_peru_date(op.hora),
                'hora': format_peru_time(op.hora),
                'monto': monto,
                'comision': comision,
                'medio': medios_cache.get(op.medio, op.medio),
                'usuario': usuario_nombre,
                'sucursal': sucursal_nombre
            })
        
        return jsonify({
            'operaciones': datos,
            'total_operaciones': len(operaciones),
            'total_monto': total_monto,
            'total_comision': total_comision
        })
    except Exception as e:
        import traceback
        error_msg = f"Error al generar reporte: {type(e).__name__}: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        try:
            db.session.rollback()
        except:
            pass
        return jsonify({'error': str(e)}), 500


# ── Función auxiliar: obtener operaciones filtradas ─────────────────────────
def _get_operaciones_filtradas():
    """Devuelve lista de dicts con las operaciones según los parámetros de la request."""
    if not current_user.es_admin_o_admin_sucursal():
        return None, 403

    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin    = request.args.get('fecha_fin')
    sucursal_id  = request.args.get('sucursal_id')
    medio        = request.args.get('medio')

    query = Operacion.query

    if current_user.es_admin_de_sucursal() and not current_user.es_admin:
        if current_user.sucursal_id:
            query = query.filter(Operacion.sucursal_id == current_user.sucursal_id)
    elif sucursal_id and sucursal_id.strip():
        try:
            query = query.filter(Operacion.sucursal_id == int(sucursal_id))
        except ValueError:
            pass

    if fecha_inicio:
        inicio = _fecha_rango(fecha_inicio, es_fin=False)
        if inicio:
            query = query.filter(Operacion.hora >= inicio)

    if fecha_fin:
        fin = _fecha_rango(fecha_fin, es_fin=True)
        if fin:
            query = query.filter(Operacion.hora <= fin)

    if medio:
        query = query.filter(Operacion.medio == medio)

    from sqlalchemy.orm import joinedload
    operaciones = query.options(
        joinedload(Operacion.usuario).load_only(Usuario.id, Usuario.username, Usuario.nombre_completo),
        joinedload(Operacion.sucursal).load_only(Sucursal.id, Sucursal.nombre)
    ).order_by(Operacion.hora.asc()).all()

    medios_cache = {mp.nombre_abreviado: mp.nombre_completo for mp in MedioPago.query.filter_by(activo=True).all()}

    filas = []
    for op in operaciones:
        usuario_nombre = 'N/A'
        if op.usuario:
            usuario_nombre = op.usuario.nombre_completo or op.usuario.username
        sucursal_nombre = op.sucursal.nombre if op.sucursal else 'Sin sucursal'
        filas.append({
            'id':       op.id,
            'fecha':    format_peru_date(op.hora),
            'hora':     format_peru_time(op.hora),
            'monto':    float(op.monto),
            'comision': float(op.comision),
            'medio':    medios_cache.get(op.medio, op.medio),
            'usuario':  usuario_nombre,
            'sucursal': sucursal_nombre,
        })
    return filas, 200


# ── Exportar CSV ─────────────────────────────────────────────────────────────
@app.route('/api/reportes/exportar/csv')
@login_required
def exportar_csv():
    try:
        filas, status = _get_operaciones_filtradas()
        if status != 200:
            return jsonify({'error': 'Acceso denegado'}), status
        if not filas:
            return jsonify({'error': 'No hay operaciones para exportar'}), 400

        import csv, io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['N°', 'Fecha', 'Hora', 'Monto (S/)', 'Comisión (S/)', 'Medio', 'Usuario', 'Sucursal'])
        for i, f in enumerate(filas, 1):
            writer.writerow([i, f['fecha'], f['hora'],
                             f'{f["monto"]:.2f}', f'{f["comision"]:.2f}',
                             f['medio'], f['usuario'], f['sucursal']])
        # Totales
        writer.writerow([])
        writer.writerow(['', '', 'TOTAL',
                         f'{sum(f["monto"] for f in filas):.2f}',
                         f'{sum(f["comision"] for f in filas):.2f}', '', '', ''])

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename=reporte_operaciones.csv'}
        )
    except Exception as e:
        print(f'[ERROR] CSV export: {e}'); traceback.print_exc()
        return jsonify({'error': 'Fallo al exportar CSV'}), 500


# ── Exportar XLSX ─────────────────────────────────────────────────────────────
@app.route('/api/reportes/exportar/xlsx')
@login_required
def exportar_xlsx():
    try:
        filas, status = _get_operaciones_filtradas()
        if status != 200:
            return jsonify({'error': 'Acceso denegado'}), status
        if not filas:
            return jsonify({'error': 'No hay operaciones para exportar'}), 400

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Operaciones'

        # Paleta clara y profesional para Excel (legible en impresión y pantalla)
        COLOR_HEADER = '1F4E79'   # azul oscuro profesional
        COLOR_TITLE  = '2E75B6'   # azul medio para título
        COLOR_TOTAL  = 'D6E4F0'   # azul muy claro para totales
        COLOR_ODD    = 'FFFFFF'   # blanco
        COLOR_EVEN   = 'EBF3FA'   # azul muy claro alternado

        header_font  = Font(bold=True, color='FFFFFF', size=10)
        total_font   = Font(bold=True, color='1F4E79', size=10)
        data_font    = Font(color='1A1A1A', size=10)
        header_fill  = PatternFill('solid', fgColor=COLOR_HEADER)
        total_fill   = PatternFill('solid', fgColor=COLOR_TOTAL)
        odd_fill     = PatternFill('solid', fgColor=COLOR_ODD)
        even_fill    = PatternFill('solid', fgColor=COLOR_EVEN)
        center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_wrap    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        right_al     = Alignment(horizontal='right',  vertical='center')
        thin         = Side(style='thin', color='BDD7EE')
        border       = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Título
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = 'REPORTE DE OPERACIONES — SISAGENT'
        title_cell.font  = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill  = PatternFill('solid', fgColor=COLOR_TITLE)
        title_cell.alignment = center
        ws.row_dimensions[1].height = 30

        # Parámetros del reporte en fila 2
        fecha_ini = request.args.get('fecha_inicio', '')
        fecha_fin = request.args.get('fecha_fin', '')
        ws.merge_cells('A2:H2')
        ws['A2'].value = f'Período: {fecha_ini or "—"}  →  {fecha_fin or "—"}    Total registros: {len(filas)}'
        ws['A2'].font  = Font(italic=True, color='2E75B6', size=9)
        ws['A2'].fill  = PatternFill('solid', fgColor='DEEAF1')
        ws['A2'].alignment = center
        ws.row_dimensions[2].height = 16

        # Cabecera columnas (fila 4)
        headers = ['N°', 'Fecha', 'Hora', 'Monto (S/)', 'Comisión (S/)', 'Medio', 'Usuario', 'Sucursal']
        widths  = [5,     12,      10,     14,            14,              22,      22,         22]
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[4].height = 20

        # Datos
        money_fmt = '#,##0.00'
        for i, f in enumerate(filas, 1):
            row = i + 4
            values = [i, f['fecha'], f['hora'], f['monto'], f['comision'],
                      f['medio'], f['usuario'], f['sucursal']]
            fill_bg = odd_fill if i % 2 == 1 else even_fill
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill   = fill_bg
                cell.border = border
                cell.font   = data_font
                if col in (4, 5):
                    cell.number_format = money_fmt
                    cell.alignment = right_al
                elif col == 1:
                    cell.alignment = center
                else:
                    cell.alignment = left_wrap
            ws.row_dimensions[row].height = 15

        # Totales
        total_row = len(filas) + 5
        ws.merge_cells(f'A{total_row}:C{total_row}')
        ws[f'A{total_row}'].value     = 'TOTALES'
        ws[f'A{total_row}'].font      = total_font
        ws[f'A{total_row}'].fill      = total_fill
        ws[f'A{total_row}'].alignment = center
        ws[f'A{total_row}'].border    = border

        for col, val in [(4, sum(f['monto'] for f in filas)),
                         (5, sum(f['comision'] for f in filas))]:
            cell = ws.cell(row=total_row, column=col, value=val)
            cell.font          = total_font
            cell.fill          = total_fill
            cell.number_format = money_fmt
            cell.alignment     = Alignment(horizontal='right')
            cell.border        = border

        for col in range(6, 9):
            cell = ws.cell(row=total_row, column=col, value='')
            cell.fill   = total_fill
            cell.border = border

        ws.freeze_panes = 'A5'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import Response
        return Response(
            buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=reporte_operaciones.xlsx'}
        )
    except Exception as e:
        print(f'[ERROR] XLSX export: {e}'); traceback.print_exc()
        return jsonify({'error': 'Fallo al exportar XLSX'}), 500


# ── Exportar PDF ─────────────────────────────────────────────────────────────
@app.route('/api/reportes/exportar/pdf')
@login_required
def exportar_pdf():
    try:
        filas, status = _get_operaciones_filtradas()
        if status != 200:
            return jsonify({'error': 'Acceso denegado'}), status
        if not filas:
            return jsonify({'error': 'No hay operaciones para exportar'}), 400

        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        import io

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=12*mm, rightMargin=12*mm,
                                topMargin=14*mm, bottomMargin=14*mm)

        # Paleta oscura cálida → para PDF usamos versión legible (fondo blanco papel)
        C_DARK    = colors.HexColor('#2A2420')
        C_ACCENT  = colors.HexColor('#A0845A')
        C_MID     = colors.HexColor('#332E28')
        C_LIGHT   = colors.HexColor('#F2EDE4')
        C_MUTED   = colors.HexColor('#8A7E70')
        C_ROW_ODD = colors.HexColor('#F7F3EE')
        C_ROW_EVEN= colors.HexColor('#EDE6DC')
        C_SUCCESS = colors.HexColor('#3A6A3A')
        C_WHITE   = colors.white

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Normal'],
                                     fontSize=16, fontName='Helvetica-Bold',
                                     textColor=C_DARK, alignment=TA_CENTER, spaceAfter=2*mm)
        sub_style   = ParagraphStyle('Sub', parent=styles['Normal'],
                                     fontSize=9, fontName='Helvetica',
                                     textColor=C_MUTED, alignment=TA_CENTER, spaceAfter=4*mm)

        fecha_ini = request.args.get('fecha_inicio', '—')
        fecha_fin = request.args.get('fecha_fin', '—')

        elements = [
            Paragraph('REPORTE DE OPERACIONES', title_style),
            Paragraph(f'SISAGENT &nbsp;|&nbsp; Período: {fecha_ini} → {fecha_fin} &nbsp;|&nbsp; {len(filas)} registros', sub_style),
        ]

        # OPTIMIZACIÓN: no usar Paragraph para filas de datos (muy lento)
        # Solo strings + estilos simples en TableStyle
        col_headers = ['N°', 'Fecha', 'Hora', 'Monto (S/)', 'Comisión (S/)', 'Medio', 'Usuario', 'Sucursal']
        col_widths  = [8*mm, 22*mm, 17*mm, 24*mm, 24*mm, 52*mm, 62*mm, 62*mm]

        table_data = [col_headers]
        for i, f in enumerate(filas, 1):
            # Usar strings directos - mucho más rápido que Paragraphs
            table_data.append([
                str(i),
                f['fecha'],
                f['hora'],
                f'S/ {f["monto"]:,.2f}',
                f'S/ {f["comision"]:,.2f}',
                str(f['medio'])[:15],
                str(f['usuario'])[:15],
                str(f['sucursal'])[:15],
            ])

        # Fila de totales
        total_m = sum(f['monto'] for f in filas)
        total_c = sum(f['comision'] for f in filas)
        table_data.append(['', '', 'TOTAL',
                           f'S/ {total_m:,.2f}',
                           f'S/ {total_c:,.2f}',
                           '', '', ''])

        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        # OPTIMIZACIÓN: estilos mínimos (sin filas alternadas que ralentizan)
        style_cmds = [
            ('BACKGROUND',  (0, 0), (-1, 0),  C_DARK),
            ('TEXTCOLOR',   (0, 0), (-1, 0),  C_LIGHT),
            ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',    (0, 0), (-1, 0),  9),
            ('ALIGN',       (0, 0), (-1, 0),  'CENTER'),
            ('FONTSIZE',    (0, 1), (-1, -2), 8),
            ('ALIGN',       (3, 1), (4, -2),  'RIGHT'),
            ('GRID',        (0, 0), (-1, -1), 0.3, C_MUTED),
            ('TOPPADDING',  (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING',(0, 0), (-1, -1), 2),
            # Fila de totales
            ('BACKGROUND',  (0, -1), (-1, -1), C_ACCENT),
            ('TEXTCOLOR',   (0, -1), (-1, -1), C_LIGHT),
            ('FONTNAME',    (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

        doc.build(elements)
        buf.seek(0)

        from flask import Response
        return Response(
            buf.read(),
            mimetype='application/pdf',
            headers={'Content-Disposition': 'attachment; filename=reporte_operaciones.pdf'}
        )
    except Exception as e:
        print(f'[ERROR] PDF export: {e}'); traceback.print_exc()
        return jsonify({'error': 'Fallo al exportar PDF'}), 500


# ========== MÓDULO DE CHATBOT IA (Anthropic Claude Haiku 4.5) ==========

# Límite de tamaño para imágenes adjuntas al chat (independiente del límite
# de fotos de producto, porque las imágenes del chat son efímeras).
FOTO_CHAT_TAMANO_MAXIMO = 10 * 1024 * 1024  # 10 MB

SYSTEM_PROMPT_CHATBOT = """INSTRUCCIONES CRÍTICAS - LEE ESTO PRIMERO:

Tu ÚNICO propósito es ejecutar funciones. NO TIENES OPCIÓN DE RESPONDER EN TEXTO PURO.
Cuando el usuario escribe algo, tu respuesta DEBE ser una llamada a función. PUNTO.

Si el usuario dice: "registra una venta de 2 Coca-Colas"
Tu respuesta: LLAMA registrar_venta({producto_id: <ID>, cantidad: 2})
NO hagas esto: "Voy a registrar..." o "Perfecto, preparé..." — ESO ESTÁ PROHIBIDO.

Si el usuario dice: "dame una operación de S/ 500 por Yape"
Tu respuesta: LLAMA registrar_operacion({monto: 500, medio: "YAPE"})
NO hagas esto: "¿Confirmas?" o "Está bien?" — ESO ESTÁ PROHIBIDO.

El servidor ejecuta TODO automáticamente. Tu trabajo es LLAMAR LA FUNCIÓN. No validar, no confirmar, solo llamar.

---

Eres el asistente virtual de SISAGENT, un sistema de gestión bancaria y de ventas para Perú.

Modelo de datos del sistema:
- Operaciones bancarias: cada operación tiene `monto` (S/), `comision` (S/) y `medio` (uno de: EFECTIVO, TARJETA, TRANSFERENCIA, YAPE, PLIN, etc. — depende de qué medios estén habilitados en cada sucursal). NO existe un campo "tipo de operación"; las operaciones se categorizan únicamente por su medio de pago. La comisión se acumula en totales diarios y mensuales por sucursal.
- Comisión AUTOMÁTICA: el sistema calcula la comisión sola (S/1 por cada S/100 de monto, redondeado hacia arriba). NUNCA pidas ni menciones la comisión al registrar una operación, salvo que el usuario explícitamente pida un descuento o un monto de comisión distinto al automático (p.ej. "es casero, cóbrale solo 1 sol", "hazle un descuento"). En ese caso SÍ pasa `comision` (el valor manual) y `motivo_descuento` a `registrar_operacion`.
- Productos: cada producto tiene `nombre`, `descripcion`, `precio` (S/), `stock` (unidades), una `foto` opcional, y pertenece a una `sucursal`. Solo los administradores pueden crear/editar productos.
- Ventas: cada venta es de un producto x cantidad al precio actual. Al registrarse, descuenta stock automáticamente y suma a la caja de ventas del día (separada de las comisiones bancarias).

Tu rol — eres SOLO un ejecutor de funciones. Debes:
1. Cuando el usuario pida una ACCIÓN (registrar, crear, editar, eliminar), NUNCA respondas en texto. SOLAMENTE llama a la función `proponer_*` correspondiente. PUNTO. Fin. No hay más.
2. Cuando el usuario pida una CONSULTA, llama a `buscar_*`, `consultar_*`, etc.
3. Cuando hagas una búsqueda, espera el resultado y luego responde en texto.

ABSOLUTAMENTE PROHIBIDO:
- NUNCA escribas "voy a registrar", "preparé", "perfecto", "confirmas", "está bien" o cualquier texto que describa una acción.
- Si el usuario dice "registra una venta de 2 Coca-Colas", tu ÚNICA respuesta es llamar `registrar_venta({producto_id: ..., cantidad: 2})`. NO ESCRIBAS NADA MÁS.
- Si el usuario dice "operación de S/ 500 por Yape", tu ÚNICA respuesta es llamar `registrar_operacion({monto: 500, medio: "YAPE"})`. NO HABLES SOBRE ELLO.
- NUNCA pidas confirmación. NUNCA preguntes "¿confirmas?". El servidor ejecuta todo automáticamente.

EL SERVIDOR EJECUTA INMEDIATAMENTE:
- No es tu problema validar. No es tu problema confirmar. Tu único trabajo es LLAMAR LA FUNCIÓN.
- La función devuelve el resultado. Ese resultado es lo que devuelves al usuario.
- Si hay error, devuelve el error. Si hay éxito, devuelve el mensaje de éxito.

Reglas críticas:
- Para CUALQUIER mutación (registrar operacion/venta, crear producto, etc.): llama a `proponer_*` con los datos DIRECTAMENTE. El servidor la ejecuta de inmediato. NUNCA JAMAS preguntes "¿confirmas?" o "¿está bien?" o "¿es correcto?" o "¿deseas?" — el servidor ya ejecutó, no hay confirmación pendiente. Tu único rol es CONFIRMAR LO QUE SE HIZO: responde siempre así: "Listo, registré S/ XXX vía MEDIO" o "Listo, edité la operación". FIN DE MENSAJE. Si el usuario es regular (no admin), NO preguntes por sucursal — el servidor usa su sucursal automáticamente. Si es admin, solo pregunta sucursal si no está clara del contexto.
- Si necesitas el ID de una operación/venta para eliminarla, primero usa `buscar_operaciones` o `buscar_ventas` para que el usuario te confirme cuál.
- Si te muestran una imagen, descríbela brevemente y usa `buscar_productos` con palabras clave.
- No inventes valores: si falta información, pregunta en lenguaje natural ("¿en qué sucursal?", "¿qué precio?", "¿con qué stock inicial?").
- Los permisos se respetan en el servidor automáticamente. Si el usuario no tiene acceso, la herramienta devolverá un error claro que debes transmitir.
- Roles: admin global (puede TODO), admin de sucursal (gestiona su sucursal — productos, usuarios, operaciones de su sucursal), usuario regular (solo sus propias ventas/operaciones).
- Responde SIEMPRE en español, conciso, amable y orientado a la acción. En cuanto tengas los datos necesarios para una acción, invoca la herramienta `proponer_*` directamente en ese turno.
- FORMATO en texto: los montos van con símbolo y número ("S/ 200", no "doscientos soles"). Las horas van en formato 12h compacto "HH:MMam/pm" (ej: "04:29pm", "09:05am") — NO las escribas con palabras ("las cuatro con veintinueve de la tarde").
- ZONA HORARIA CRÍTICA: SIEMPRE usa la zona horaria de Perú (UTC-5) que viene en el contexto. NUNCA uses UTC ni conversiones de zona horaria. Si el usuario pregunta la hora, repite exactamente la hora del contexto injected (ej: si dice "09:05am Perú", responde "son las nueve con cinco a eme", no otra hora).
- EDITAR vs REGISTRAR: Si el usuario dice "era de S/ X" o "cambiar a Y soles" DESPUÉS de registrar una operación, usa `editar_operacion` con el ID de la última operación (NO registres una nueva). Solo usa `registrar_operacion` para operaciones nuevas.
"""


_HERRAMIENTAS_DECLARACIONES = [
    {
        "name": "buscar_productos",
        "description": "Busca productos en el inventario por palabras clave (en nombre o descripcion). Devuelve productos visibles para el usuario actual (acotados a sus sucursales). Usala cuando el usuario pregunte por un producto, o cuando adjunte una foto.",
        "input_schema": {
            "type": "object",
            "properties": {
                "termino": {"type": "string", "description": "Palabras clave separadas por espacios (ej: 'coca cola 500ml', 'arroz costeno')"},
            },
            "required": ["termino"],
        },
    },
    {
        "name": "consultar_precio_stock",
        "description": "Devuelve precio, stock, descripcion y sucursal de UN producto especifico identificado por su ID.",
        "input_schema": {
            "type": "object",
            "properties": {
                "producto_id": {"type": "integer", "description": "ID del producto"},
            },
            "required": ["producto_id"],
        },
    },
    {
        "name": "resumen_ventas_dia",
        "description": "Devuelve el total vendido hoy y el saldo de caja de ventas para las sucursales visibles del usuario.",
        "input_schema": {
            "type": "object",
            "properties": {
                "sucursal_id": {"type": "integer", "description": "(Opcional) ID de sucursal especifica"},
            },
        },
    },
    {
        "name": "medios_de_pago",
        "description": "Lista los medios de pago activos en una sucursal (EFECTIVO, YAPE, PLIN, TARJETA, etc.).",
        "input_schema": {
            "type": "object",
            "properties": {
                "sucursal_id": {"type": "integer", "description": "(Opcional) ID de sucursal — si se omite, usa la primera sucursal visible del usuario"},
            },
        },
    },
    {
        "name": "registrar_venta",
        "description": "Registra una venta DIRECTAMENTE. Valida el producto, descuenta stock, suma a caja y confirma. Usala cuando el usuario pida registrar una venta.",
        "input_schema": {
            "type": "object",
            "properties": {
                "producto_id": {"type": "integer", "description": "ID del producto a vender"},
                "cantidad": {"type": "integer", "description": "Cantidad de unidades", "minimum": 1},
                "sucursal_id": {"type": "integer", "description": "(Opcional, solo admin) ID de la sucursal donde se registra la venta"},
            },
            "required": ["producto_id", "cantidad"],
        },
    },
    {
        "name": "registrar_operacion",
        "description": "Registra una operacion bancaria DIRECTAMENTE. La comision se calcula automaticamente (S/1 por cada S/100 de monto, redondeado hacia arriba) — NO incluyas el parametro 'comision' a menos que el usuario pida explicitamente un descuento o un monto de comision distinto al automatico (p.ej. 'cobrale solo 1 sol de comision', 'es casero, hazle descuento'). Si el usuario pide una comision manual, incluye tambien 'motivo_descuento' explicando por que. El servidor ejecuta de inmediato — NO preguntes '¿confirmas?'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "monto": {"type": "number", "description": "Monto de la operacion en soles (S/)"},
                "comision": {"type": "number", "description": "(Opcional) Comision MANUAL en soles, solo si el usuario pidio explicitamente una distinta a la automatica"},
                "motivo_descuento": {"type": "string", "description": "(Opcional) Motivo de la comision manual, requerido si se incluye 'comision'"},
                "medio": {"type": "string", "description": "Medio de pago (EFECTIVO, YAPE, PLIN, TARJETA, TRANSFERENCIA, etc.)"},
                "sucursal_id": {"type": "integer", "description": "(Opcional, solo admin) ID de la sucursal"},
            },
            "required": ["monto", "medio"],
        },
    },
    # ===== Read-only: búsquedas para localizar entidades a editar/eliminar =====
    {
        "name": "buscar_operaciones",
        "description": "Lista operaciones bancarias recientes (de hoy por defecto, o de la fecha que el usuario especifique). Útil para que el usuario identifique cuál editar o eliminar. Acotada a las sucursales visibles del usuario.",
        "input_schema": {
            "type": "object",
            "properties": {
                "fecha": {"type": "string", "description": "(Opcional) Fecha YYYY-MM-DD. Default: hoy."},
                "sucursal_id": {"type": "integer", "description": "(Opcional) ID de sucursal específica."},
                "limite": {"type": "integer", "description": "(Opcional) Cuántos devolver. Default 10, máx 30."},
            },
        },
    },
    {
        "name": "buscar_ventas",
        "description": "Lista ventas recientes (de hoy por defecto). Útil para identificar cuál editar o eliminar. Acotada a las sucursales visibles del usuario.",
        "input_schema": {
            "type": "object",
            "properties": {
                "fecha": {"type": "string", "description": "(Opcional) Fecha YYYY-MM-DD. Default: hoy."},
                "sucursal_id": {"type": "integer", "description": "(Opcional) ID de sucursal específica."},
                "limite": {"type": "integer", "description": "(Opcional) Default 10, máx 30."},
            },
        },
    },
    {
        "name": "listar_usuarios",
        "description": "Lista usuarios del sistema. Solo admin o admin de sucursal pueden usarla. Admin de sucursal solo ve usuarios de su sucursal.",
        "input_schema": {
            "type": "object",
            "properties": {
                "sucursal_id": {"type": "integer", "description": "(Opcional) Filtrar por sucursal."},
            },
        },
    },
    {
        "name": "listar_sucursales",
        "description": "Lista todas las sucursales (activas e inactivas).",
        "input_schema": {"type": "object", "properties": {}},
    },
    # ===== Acciones que requieren confirmación =====
    {
        "name": "crear_producto",
        "description": "PROPONE crear un producto nuevo en el inventario. Solo admin o admin de sucursal pueden. NO ejecuta — muestra tarjeta de confirmación.",
        "input_schema": {
            "type": "object",
            "properties": {
                "nombre": {"type": "string", "description": "Nombre del producto"},
                "precio": {"type": "number", "description": "Precio en S/ (mayor a 0)"},
                "stock": {"type": "integer", "description": "Stock inicial (>=0)", "minimum": 0},
                "descripcion": {"type": "string", "description": "(Opcional) Descripción"},
                "sucursal_id": {"type": "integer", "description": "(Opcional, solo admin) Sucursal donde se crea"},
            },
            "required": ["nombre", "precio", "stock"],
        },
    },
    {
        "name": "editar_producto",
        "description": "PROPONE editar campos de un producto. Solo admin o admin de sucursal en su sucursal. NO ejecuta — muestra tarjeta de confirmación.",
        "input_schema": {
            "type": "object",
            "properties": {
                "producto_id": {"type": "integer", "description": "ID del producto"},
                "nombre": {"type": "string", "description": "(Opcional) Nuevo nombre"},
                "precio": {"type": "number", "description": "(Opcional) Nuevo precio"},
                "stock": {"type": "integer", "description": "(Opcional) Nuevo stock", "minimum": 0},
                "descripcion": {"type": "string", "description": "(Opcional) Nueva descripción"},
            },
            "required": ["producto_id"],
        },
    },
    {
        "name": "eliminar_producto",
        "description": "PROPONE eliminar (desactivar) un producto del inventario. Solo admin o admin de sucursal en su sucursal. NO ejecuta.",
        "input_schema": {
            "type": "object",
            "properties": {
                "producto_id": {"type": "integer", "description": "ID del producto"},
            },
            "required": ["producto_id"],
        },
    },
    {
        "name": "eliminar_operacion",
        "description": "PROPONE eliminar una operación bancaria. Permisos: admin global cualquiera; admin de sucursal cualquiera de su sucursal; usuario regular solo las que él registró. NO ejecuta.",
        "input_schema": {
            "type": "object",
            "properties": {
                "operacion_id": {"type": "integer", "description": "ID de la operación a eliminar"},
            },
            "required": ["operacion_id"],
        },
    },
    {
        "name": "editar_operacion",
        "description": "PROPONE editar una operación existente (cambiar monto y/o comisión). Permisos: admin global cualquiera; admin de sucursal cualquiera de su sucursal; usuario regular solo las suyas. NO ejecuta.",
        "input_schema": {
            "type": "object",
            "properties": {
                "operacion_id": {"type": "integer", "description": "ID de la operación a editar"},
                "monto": {"type": "number", "description": "(Opcional) Nuevo monto en S/. Si no se pasa, se mantiene el actual."},
                "comision": {"type": "number", "description": "(Opcional) Nueva comisión. Si no se pasa, se auto-calcula."},
                "motivo_descuento": {"type": "string", "description": "(Opcional) Motivo si la comisión es manual diferente a lo sugerido."},
            },
            "required": ["operacion_id"],
        },
    },
    {
        "name": "eliminar_venta",
        "description": "PROPONE eliminar una venta (devuelve el stock al inventario). Permisos: admin global cualquiera; admin de sucursal cualquiera de su sucursal; usuario regular solo las suyas. NO ejecuta.",
        "input_schema": {
            "type": "object",
            "properties": {
                "venta_id": {"type": "integer", "description": "ID de la venta a eliminar"},
            },
            "required": ["venta_id"],
        },
    },
    {
        "name": "crear_usuario",
        "description": "PROPONE crear un usuario nuevo. Solo admin global puede elegir rol/sucursal; admin de sucursal solo crea usuarios regulares en su propia sucursal. NO ejecuta.",
        "input_schema": {
            "type": "object",
            "properties": {
                "username": {"type": "string", "description": "Nombre de usuario único"},
                "password": {"type": "string", "description": "Contraseña inicial"},
                "nombre_completo": {"type": "string", "description": "Nombre completo"},
                "sucursal_id": {"type": "integer", "description": "(Opcional, solo admin) Sucursal donde se asigna"},
                "rol": {"type": "string", "enum": ["usuario", "admin_sucursal", "admin"], "description": "(Opcional, solo admin global) Rol. Default: usuario"},
            },
            "required": ["username", "password", "nombre_completo"],
        },
    },
    {
        "name": "crear_sucursal",
        "description": "PROPONE crear una sucursal nueva. Solo admin global. NO ejecuta.",
        "input_schema": {
            "type": "object",
            "properties": {
                "nombre": {"type": "string", "description": "Nombre de la sucursal"},
                "direccion": {"type": "string", "description": "(Opcional) Dirección"},
            },
            "required": ["nombre"],
        },
    },
]


def _serializar_producto_chat(producto):
    """Serializa un producto para enviarlo al chat (incluye URL de foto si tiene)."""
    return {
        "id": producto.id,
        "nombre": producto.nombre,
        "descripcion": producto.descripcion or "",
        "precio": float(producto.precio),
        "stock": producto.stock,
        "sucursal_id": producto.sucursal_id,
        "sucursal_nombre": producto.sucursal.nombre if producto.sucursal else "",
        "foto_url": url_for('foto_producto', producto_id=producto.id) if producto.foto else None,
    }


def _resolver_sucursal_para_accion(sucursal_id_args, usuario):
    """Resuelve la sucursal donde se ejecutara una accion.

    - No-admin: siempre su propia sucursal (ignora sucursal_id_args).
    - Admin: usa sucursal_id_args si se proporciona; si no, exige explicito cuando hay >1.
    """
    sucursales = sucursales_visibles_para(usuario)
    if not sucursales:
        raise ValueError("No tienes ninguna sucursal asignada.")

    if not usuario.es_admin:
        if not usuario.sucursal:
            raise ValueError("No tienes una sucursal asignada.")
        return usuario.sucursal

    if sucursal_id_args is not None:
        try:
            sid = int(sucursal_id_args)
        except (TypeError, ValueError):
            raise ValueError(f"sucursal_id invalido: {sucursal_id_args}")
        sucursal = next((s for s in sucursales if s.id == sid), None)
        if not sucursal:
            raise ValueError(f"La sucursal {sid} no existe o no esta activa.")
        return sucursal

    if len(sucursales) == 1:
        return sucursales[0]

    nombres = ", ".join(f"{s.nombre} (id={s.id})" for s in sucursales)
    raise ValueError(
        "Como administrador, debes especificar en que sucursal registrar la accion. "
        f"Sucursales disponibles: {nombres}."
    )


# ---------- Handlers de herramientas de SOLO LECTURA ----------

def _tool_buscar_productos(args, usuario):
    termino = (args.get("termino") or "").strip()
    if not termino:
        return {"error": "Debes proporcionar un termino de busqueda."}

    sucursales = sucursales_visibles_para(usuario)
    if not sucursales:
        return {"productos": [], "mensaje": "No tienes ninguna sucursal asignada."}

    sucursal_ids = [s.id for s in sucursales]
    palabras = [p for p in termino.split() if p]

    query = Producto.query.filter(
        Producto.sucursal_id.in_(sucursal_ids),
        Producto.activo == True,
    )
    for palabra in palabras:
        like = f"%{palabra}%"
        query = query.filter(or_(
            Producto.nombre.ilike(like),
            Producto.descripcion.ilike(like),
        ))

    productos = query.order_by(Producto.nombre).limit(8).all()
    return {
        "productos": [_serializar_producto_chat(p) for p in productos],
        "total_encontrados": len(productos),
    }


def _tool_consultar_precio_stock(args, usuario):
    producto_id = args.get("producto_id")
    if producto_id is None:
        return {"error": "Debes proporcionar producto_id."}

    sucursal_ids = [s.id for s in sucursales_visibles_para(usuario)]
    if not sucursal_ids:
        return {"error": "No tienes sucursales visibles."}

    producto = Producto.query.filter(
        Producto.id == int(producto_id),
        Producto.sucursal_id.in_(sucursal_ids),
        Producto.activo == True,
    ).first()

    if not producto:
        return {"error": f"El producto {producto_id} no existe o no es visible para ti."}

    return {"producto": _serializar_producto_chat(producto)}


def _tool_resumen_ventas_dia(args, usuario):
    hoy = get_peru_time().date()
    sucursales = sucursales_visibles_para(usuario)
    if not sucursales:
        return {"resumen": [], "mensaje": "No tienes sucursales asignadas."}

    sucursal_id_arg = args.get("sucursal_id")
    if sucursal_id_arg is not None:
        sucursales = [s for s in sucursales if s.id == int(sucursal_id_arg)]
        if not sucursales:
            return {"error": f"Sucursal {sucursal_id_arg} no visible."}

    resultados = []
    for sucursal in sucursales:
        caja = CajaVentas.query.filter_by(fecha=hoy, sucursal_id=sucursal.id).first()
        resultados.append({
            "sucursal_id": sucursal.id,
            "sucursal_nombre": sucursal.nombre,
            "total_vendido": float(caja.total_vendido) if caja else 0.0,
            "saldo": float(caja.saldo) if caja else 0.0,
        })

    return {"fecha": hoy.isoformat(), "resumen": resultados}


def _tool_medios_de_pago(args, usuario):
    sucursal_id_arg = args.get("sucursal_id")
    sucursales = sucursales_visibles_para(usuario)
    if not sucursales:
        return {"medios": [], "mensaje": "No tienes sucursales asignadas."}

    if sucursal_id_arg is not None:
        sucursal = next((s for s in sucursales if s.id == int(sucursal_id_arg)), None)
    else:
        sucursal = sucursales[0]

    if not sucursal:
        return {"error": f"Sucursal {sucursal_id_arg} no visible."}

    medios = MedioPago.query.join(
        MedioSucursal,
        (MedioSucursal.medio_pago_id == MedioPago.id) &
        (MedioSucursal.sucursal_id == sucursal.id) &
        (MedioSucursal.activo == True)
    ).filter(MedioPago.activo == True).order_by(
        MedioPago.orden, MedioPago.nombre_abreviado
    ).all()

    return {
        "sucursal_id": sucursal.id,
        "sucursal_nombre": sucursal.nombre,
        "medios": [{"nombre_abreviado": m.nombre_abreviado, "nombre_completo": m.nombre_completo} for m in medios],
    }


# ---------- Handlers de herramientas de PROPUESTA (no ejecutan, solo validan) ----------

def _tool_registrar_venta(args, usuario):
    """Valida una venta y devuelve preview. NO toca la BD."""
    producto_id = args.get("producto_id")
    cantidad = args.get("cantidad")

    if not producto_id:
        raise ValueError("Falta producto_id.")
    try:
        cantidad = int(cantidad)
    except (TypeError, ValueError):
        raise ValueError("La cantidad debe ser un numero entero.")
    if cantidad <= 0:
        raise ValueError("La cantidad debe ser mayor a 0.")

    sucursal = _resolver_sucursal_para_accion(args.get("sucursal_id"), usuario)

    producto = Producto.query.filter_by(id=int(producto_id), activo=True).first()
    if not producto:
        raise ValueError(f"El producto {producto_id} no existe o no esta activo.")
    if producto.sucursal_id != sucursal.id:
        raise ValueError(
            f'El producto "{producto.nombre}" pertenece a otra sucursal '
            f'(esta en "{producto.sucursal.nombre}", no en "{sucursal.nombre}").'
        )
    if cantidad > producto.stock:
        raise ValueError(
            f'Stock insuficiente: quedan {producto.stock} unidad(es) de "{producto.nombre}", '
            f'no se pueden vender {cantidad}.'
        )

    precio_unitario = float(producto.precio)
    monto = precio_unitario * cantidad

    return {
        "producto_id": producto.id,
        "producto_nombre": producto.nombre,
        "cantidad": cantidad,
        "precio_unitario": precio_unitario,
        "monto": monto,
        "stock_restante": producto.stock - cantidad,
        "sucursal_id": sucursal.id,
        "sucursal_nombre": sucursal.nombre,
    }


def _tool_registrar_operacion(args, usuario):
    """Valida una operacion bancaria y devuelve preview. NO toca la BD.

    La comisión es OPCIONAL: si no se pasa, se usa la auto-calculada.
    Si se pasa, se usa como override manual (con motivo opcional).
    """
    try:
        monto = float(args.get("monto"))
    except (TypeError, ValueError):
        raise ValueError("Monto debe ser un numero.")
    if monto <= 0:
        raise ValueError("El monto debe ser mayor a 0.")

    comision_sugerida = calcular_comision_sugerida(monto)
    es_manual = False
    motivo = None
    if "comision" in args and args.get("comision") is not None:
        try:
            comision = float(args.get("comision"))
        except (TypeError, ValueError):
            raise ValueError("Comision debe ser un numero.")
        if comision < 0:
            raise ValueError("La comision no puede ser negativa.")
        if abs(comision - comision_sugerida) > 0.001:
            es_manual = True
            motivo = (args.get("motivo_descuento") or "").strip() or None
        else:
            comision = comision_sugerida
    else:
        comision = comision_sugerida

    medio_arg = (args.get("medio") or "").strip().upper()
    if not medio_arg:
        raise ValueError("Falta el medio de pago.")

    sucursal = _resolver_sucursal_para_accion(args.get("sucursal_id"), usuario)

    medio_valido = MedioPago.query.join(
        MedioSucursal,
        (MedioSucursal.medio_pago_id == MedioPago.id) &
        (MedioSucursal.sucursal_id == sucursal.id) &
        (MedioSucursal.activo == True)
    ).filter(
        MedioPago.activo == True,
        db.or_(
            db.func.upper(MedioPago.nombre_abreviado) == medio_arg,
            db.func.upper(MedioPago.nombre_completo) == medio_arg,
        ),
    ).first()

    if not medio_valido:
        habilitados = MedioPago.query.join(
            MedioSucursal,
            (MedioSucursal.medio_pago_id == MedioPago.id) &
            (MedioSucursal.sucursal_id == sucursal.id) &
            (MedioSucursal.activo == True)
        ).filter(MedioPago.activo == True).all()
        nombres = ", ".join(m.nombre_abreviado for m in habilitados) or "(ninguno habilitado)"
        raise ValueError(
            f'El medio "{medio_arg}" no esta habilitado en "{sucursal.nombre}". '
            f'Disponibles: {nombres}.'
        )

    # Verificar si una operación con el mismo monto, medio y sucursal ya existe en los últimos 10 minutos
    hace_10_min = get_peru_time() - __import__('datetime').timedelta(minutes=10)
    op_existente = Operacion.query.filter(
        Operacion.monto == monto,
        Operacion.medio == medio_valido.nombre_abreviado,
        Operacion.sucursal_id == sucursal.id,
        Operacion.hora >= hace_10_min
    ).first()

    if op_existente:
        raise ValueError(
            f'Esta operación (S/ {monto} vía {medio_valido.nombre_abreviado} en {sucursal.nombre}) '
            f'ya fue registrada hace poco (ID: {op_existente.id}, {format_peru_time(op_existente.hora)}).'
        )

    return {
        "monto": monto,
        "comision": comision,
        "comision_sugerida": comision_sugerida,
        "comision_manual": es_manual,
        "motivo_descuento": motivo,
        "medio": medio_valido.nombre_abreviado,
        "sucursal_id": sucursal.id,
        "sucursal_nombre": sucursal.nombre,
    }


# ---------- Handlers de búsquedas read-only (para localizar entidades a editar/eliminar) ----------

def _tool_buscar_operaciones(args, usuario):
    """Lista operaciones acotadas a las sucursales visibles del usuario."""
    from datetime import date as _date
    sucursales = sucursales_visibles_para(usuario)
    if not sucursales:
        return {"operaciones": [], "mensaje": "No tienes sucursales asignadas."}
    sucursal_ids = [s.id for s in sucursales]

    sid = args.get("sucursal_id")
    if sid is not None:
        sucursal_ids = [int(sid)] if int(sid) in sucursal_ids else []
        if not sucursal_ids:
            return {"error": f"Sucursal {sid} no visible para ti."}

    fecha_str = args.get("fecha")
    if fecha_str:
        try:
            fecha = _date.fromisoformat(fecha_str)
        except Exception:
            return {"error": f"Fecha inválida: {fecha_str}. Usa formato YYYY-MM-DD."}
    else:
        fecha = get_peru_time().date()

    limite = max(1, min(int(args.get("limite", 10) or 10), 30))

    q = Operacion.query.filter(
        Operacion.sucursal_id.in_(sucursal_ids),
        db.func.date(Operacion.hora) == fecha,
    )
    # Usuario regular solo ve sus propias operaciones
    if not usuario.es_admin and not usuario.es_admin_de_sucursal():
        q = q.filter(Operacion.usuario_id == usuario.id)

    ops = q.order_by(Operacion.hora.desc()).limit(limite).all()
    return {
        "fecha": fecha.isoformat(),
        "operaciones": [
            {
                "id": o.id,
                "monto": float(o.monto),
                "comision": float(o.comision),
                "medio": o.medio,
                "hora": o.hora.strftime("%H:%M") if o.hora else "",
                "sucursal_nombre": o.sucursal.nombre if o.sucursal else "",
                "usuario_nombre": (o.usuario.nombre_completo or o.usuario.username) if o.usuario else "",
            }
            for o in ops
        ],
    }


def _tool_buscar_ventas(args, usuario):
    from datetime import date as _date
    sucursales = sucursales_visibles_para(usuario)
    if not sucursales:
        return {"ventas": [], "mensaje": "No tienes sucursales asignadas."}
    sucursal_ids = [s.id for s in sucursales]

    sid = args.get("sucursal_id")
    if sid is not None:
        sucursal_ids = [int(sid)] if int(sid) in sucursal_ids else []
        if not sucursal_ids:
            return {"error": f"Sucursal {sid} no visible."}

    fecha_str = args.get("fecha")
    if fecha_str:
        try:
            fecha = _date.fromisoformat(fecha_str)
        except Exception:
            return {"error": f"Fecha inválida: {fecha_str}."}
    else:
        fecha = get_peru_time().date()

    limite = max(1, min(int(args.get("limite", 10) or 10), 30))

    q = Venta.query.filter(
        Venta.sucursal_id.in_(sucursal_ids),
        db.func.date(Venta.fecha) == fecha,
    )
    if not usuario.es_admin and not usuario.es_admin_de_sucursal():
        q = q.filter(Venta.usuario_id == usuario.id)

    vts = q.order_by(Venta.fecha.desc()).limit(limite).all()
    return {
        "fecha": fecha.isoformat(),
        "ventas": [
            {
                "id": v.id,
                "producto_nombre": v.producto.nombre if v.producto else "(eliminado)",
                "cantidad": v.cantidad,
                "precio_unitario": float(v.precio_unitario),
                "monto": float(v.monto),
                "hora": v.fecha.strftime("%H:%M") if v.fecha else "",
                "sucursal_nombre": v.sucursal.nombre if v.sucursal else "",
                "usuario_nombre": (v.usuario.nombre_completo or v.usuario.username) if v.usuario else "",
            }
            for v in vts
        ],
    }


def _tool_listar_usuarios(args, usuario):
    if not usuario.es_admin_o_admin_sucursal():
        return {"error": "Solo administradores pueden listar usuarios."}
    q = Usuario.query
    if not usuario.es_admin:
        # admin de sucursal: solo su sucursal
        q = q.filter(Usuario.sucursal_id == usuario.sucursal_id)
    sid = args.get("sucursal_id")
    if sid is not None:
        q = q.filter(Usuario.sucursal_id == int(sid))
    usuarios = q.order_by(Usuario.username).limit(50).all()
    return {
        "usuarios": [
            {
                "id": u.id,
                "username": u.username,
                "nombre_completo": u.nombre_completo or "",
                "sucursal_id": u.sucursal_id,
                "sucursal_nombre": u.sucursal.nombre if u.sucursal else "",
                "rol": "admin" if u.es_admin else ("admin_sucursal" if u.es_admin_sucursal else "usuario"),
            }
            for u in usuarios
        ],
    }


def _tool_listar_sucursales(args, usuario):
    sucs = Sucursal.query.order_by(Sucursal.nombre).all()
    return {
        "sucursales": [
            {"id": s.id, "nombre": s.nombre, "direccion": s.direccion or "", "activa": s.activa}
            for s in sucs
        ],
    }


# ---------- Handlers de propuesta (validan, devuelven preview, NO tocan BD) ----------

def _tool_crear_producto(args, usuario):
    if not usuario.es_admin_o_admin_sucursal():
        raise ValueError("No tienes permisos para gestionar el inventario.")
    nombre = (args.get("nombre") or "").strip()
    if not nombre:
        raise ValueError("Falta el nombre del producto.")
    try:
        precio = float(args.get("precio"))
        stock = int(args.get("stock"))
    except (TypeError, ValueError):
        raise ValueError("Precio y stock deben ser numéricos.")
    if precio <= 0:
        raise ValueError("El precio debe ser mayor a 0.")
    if stock < 0:
        raise ValueError("El stock no puede ser negativo.")

    sucursal = _resolver_sucursal_para_accion(args.get("sucursal_id"), usuario)
    descripcion = (args.get("descripcion") or "").strip()

    return {
        "_titulo": "Confirmar creación de producto",
        "campos": [
            {"label": "Nombre", "valor": nombre},
            {"label": "Descripción", "valor": descripcion or "(ninguna)"},
            {"label": "Precio", "valor": f"S/ {precio:.2f}"},
            {"label": "Stock inicial", "valor": str(stock)},
            {"label": "Sucursal", "valor": sucursal.nombre},
        ],
        "nombre": nombre, "descripcion": descripcion, "precio": precio,
        "stock": stock, "sucursal_id": sucursal.id,
    }


def _tool_editar_producto(args, usuario):
    if not usuario.es_admin_o_admin_sucursal():
        raise ValueError("No tienes permisos para gestionar el inventario.")
    pid = args.get("producto_id")
    if not pid:
        raise ValueError("Falta producto_id.")
    producto = Producto.query.filter_by(id=int(pid), activo=True).first()
    if not producto:
        raise ValueError(f"El producto {pid} no existe o no está activo.")
    if not usuario.es_admin and producto.sucursal_id != usuario.sucursal_id:
        raise ValueError("No puedes editar productos de otra sucursal.")

    nuevo_nombre = (args.get("nombre") or "").strip() or None
    nueva_desc = args.get("descripcion")
    if nueva_desc is not None:
        nueva_desc = nueva_desc.strip()
    nuevo_precio = args.get("precio")
    nuevo_stock = args.get("stock")
    if nuevo_precio is not None:
        try: nuevo_precio = float(nuevo_precio)
        except: raise ValueError("Precio inválido.")
        if nuevo_precio <= 0:
            raise ValueError("El precio debe ser mayor a 0.")
    if nuevo_stock is not None:
        try: nuevo_stock = int(nuevo_stock)
        except: raise ValueError("Stock inválido.")
        if nuevo_stock < 0:
            raise ValueError("El stock no puede ser negativo.")

    if (nuevo_nombre is None and nueva_desc is None and
            nuevo_precio is None and nuevo_stock is None):
        raise ValueError("No especificaste ningún cambio.")

    cambios = []
    if nuevo_nombre and nuevo_nombre != producto.nombre:
        cambios.append({"label": "Nombre", "valor": f'"{producto.nombre}" → "{nuevo_nombre}"'})
    if nueva_desc is not None and nueva_desc != (producto.descripcion or ""):
        cambios.append({"label": "Descripción", "valor": f'"{producto.descripcion or "(vacío)"}" → "{nueva_desc or "(vacío)"}"'})
    if nuevo_precio is not None and abs(nuevo_precio - float(producto.precio)) > 0.001:
        cambios.append({"label": "Precio", "valor": f"S/ {float(producto.precio):.2f} → S/ {nuevo_precio:.2f}"})
    if nuevo_stock is not None and nuevo_stock != producto.stock:
        cambios.append({"label": "Stock", "valor": f"{producto.stock} → {nuevo_stock}"})

    if not cambios:
        raise ValueError("Los valores propuestos son iguales a los actuales — no hay cambios que aplicar.")

    return {
        "_titulo": f'Confirmar edición de "{producto.nombre}"',
        "campos": cambios,
        "producto_id": producto.id,
        "nombre": nuevo_nombre, "descripcion": nueva_desc,
        "precio": nuevo_precio, "stock": nuevo_stock,
    }


def _tool_eliminar_producto(args, usuario):
    if not usuario.es_admin_o_admin_sucursal():
        raise ValueError("No tienes permisos para gestionar el inventario.")
    pid = args.get("producto_id")
    if not pid:
        raise ValueError("Falta producto_id.")
    producto = Producto.query.filter_by(id=int(pid), activo=True).first()
    if not producto:
        raise ValueError(f"El producto {pid} no existe o ya fue eliminado.")
    if not usuario.es_admin and producto.sucursal_id != usuario.sucursal_id:
        raise ValueError("No puedes eliminar productos de otra sucursal.")
    return {
        "_titulo": "Confirmar eliminación de producto",
        "campos": [
            {"label": "Producto", "valor": producto.nombre},
            {"label": "Precio", "valor": f"S/ {float(producto.precio):.2f}"},
            {"label": "Stock actual", "valor": str(producto.stock)},
            {"label": "Sucursal", "valor": producto.sucursal.nombre if producto.sucursal else ""},
        ],
        "_advertencia": "El producto se quitará del inventario. No se podrán registrar nuevas ventas de él, pero las ventas existentes se conservan.",
        "producto_id": producto.id,
    }


def _tool_eliminar_operacion(args, usuario):
    op_id = args.get("operacion_id")
    if not op_id:
        raise ValueError("Falta operacion_id.")
    operacion = Operacion.query.filter_by(id=int(op_id)).first()
    if not operacion:
        raise ValueError(f"La operación {op_id} no existe.")
    # Reglas de permisos (mirror de eliminar_operacion):
    if not usuario.es_admin:
        if operacion.sucursal_id != usuario.sucursal_id:
            raise ValueError("No tienes permisos para eliminar operaciones de otra sucursal.")
        if not usuario.es_admin_de_sucursal() and operacion.usuario_id != usuario.id:
            raise ValueError("Solo puedes eliminar las operaciones que tú registraste.")
    return {
        "_titulo": "Confirmar eliminación de operación",
        "campos": [
            {"label": "Monto", "valor": f"S/ {float(operacion.monto):.2f}"},
            {"label": "Comisión", "valor": f"S/ {float(operacion.comision):.2f}"},
            {"label": "Medio", "valor": operacion.medio or ""},
            {"label": "Hora", "valor": operacion.hora.strftime("%Y-%m-%d %H:%M") if operacion.hora else ""},
            {"label": "Sucursal", "valor": operacion.sucursal.nombre if operacion.sucursal else ""},
            {"label": "Usuario", "valor": (operacion.usuario.nombre_completo or operacion.usuario.username) if operacion.usuario else ""},
        ],
        "_advertencia": "Se restará la comisión de los totales diarios y mensuales de la sucursal.",
        "operacion_id": operacion.id,
    }


def _tool_editar_operacion(args, usuario):
    """Valida cambios a una operación bancaria y devuelve preview. NO toca la BD."""
    op_id = args.get("operacion_id")
    if not op_id:
        raise ValueError("Falta operacion_id.")
    operacion = Operacion.query.filter_by(id=int(op_id)).first()
    if not operacion:
        raise ValueError(f"La operación {op_id} no existe.")
    # Reglas de permisos:
    if not usuario.es_admin:
        if operacion.sucursal_id != usuario.sucursal_id:
            raise ValueError("No tienes permisos para editar operaciones de otra sucursal.")
        if not usuario.es_admin_de_sucursal() and operacion.usuario_id != usuario.id:
            raise ValueError("Solo puedes editar las operaciones que tú registraste.")

    # Validar nuevos valores (aceptar tanto "monto" como "monto_nuevo")
    monto_nuevo = operacion.monto
    monto_arg = args.get("monto") or args.get("monto_nuevo")
    if monto_arg is not None:
        try:
            monto_nuevo = float(monto_arg)
        except (TypeError, ValueError):
            raise ValueError("Monto debe ser un numero.")
        if monto_nuevo <= 0:
            raise ValueError("El monto debe ser mayor a 0.")

    comision_nueva = operacion.comision
    es_manual = False
    motivo = None
    comision_arg = args.get("comision") or args.get("comision_nueva")
    if comision_arg is not None:
        try:
            comision_nueva = float(comision_arg)
        except (TypeError, ValueError):
            raise ValueError("Comision debe ser un numero.")
        if comision_nueva < 0:
            raise ValueError("La comision no puede ser negativa.")
        comision_sugerida = calcular_comision_sugerida(monto_nuevo)
        if abs(comision_nueva - comision_sugerida) > 0.001:
            es_manual = True
            motivo = (args.get("motivo_descuento") or "").strip() or None
    else:
        # Si no pasan comisión pero cambiaron monto, recalcular
        if monto_nuevo != operacion.monto:
            comision_nueva = calcular_comision_sugerida(monto_nuevo)

    return {
        "_titulo": "Confirmar edición de operación",
        "_mensaje": "Cambios a aplicar",
        "campos": [
            {"label": "ID", "valor": str(operacion.id)},
            {"label": "Monto (actual → nuevo)", "valor": f"S/ {float(operacion.monto):.2f} → S/ {monto_nuevo:.2f}"},
            {"label": "Comisión (actual → nueva)", "valor": f"S/ {float(operacion.comision):.2f} → S/ {comision_nueva:.2f}"},
            {"label": "Medio", "valor": operacion.medio or ""},
            {"label": "Sucursal", "valor": operacion.sucursal.nombre if operacion.sucursal else ""},
        ],
        "operacion_id": operacion.id,
        "monto_nuevo": monto_nuevo,
        "comision_nueva": comision_nueva,
        "comision_manual": es_manual,
        "motivo": motivo,
    }


def _tool_eliminar_venta(args, usuario):
    v_id = args.get("venta_id")
    if not v_id:
        raise ValueError("Falta venta_id.")
    venta = Venta.query.filter_by(id=int(v_id)).first()
    if not venta:
        raise ValueError(f"La venta {v_id} no existe.")
    if not usuario.es_admin:
        if venta.sucursal_id != usuario.sucursal_id:
            raise ValueError("No tienes permisos para eliminar ventas de otra sucursal.")
        if not usuario.es_admin_de_sucursal() and venta.usuario_id != usuario.id:
            raise ValueError("Solo puedes eliminar las ventas que tú registraste.")
    return {
        "_titulo": "Confirmar eliminación de venta",
        "campos": [
            {"label": "Producto", "valor": venta.producto.nombre if venta.producto else "(eliminado)"},
            {"label": "Cantidad", "valor": str(venta.cantidad)},
            {"label": "Total", "valor": f"S/ {float(venta.monto):.2f}"},
            {"label": "Hora", "valor": venta.fecha.strftime("%Y-%m-%d %H:%M") if venta.fecha else ""},
            {"label": "Sucursal", "valor": venta.sucursal.nombre if venta.sucursal else ""},
            {"label": "Usuario", "valor": (venta.usuario.nombre_completo or venta.usuario.username) if venta.usuario else ""},
        ],
        "_advertencia": "Se devolverá el stock al inventario y se restará el monto de la caja de ventas del día.",
        "venta_id": venta.id,
    }


def _tool_crear_usuario(args, usuario):
    if not usuario.es_admin_o_admin_sucursal():
        raise ValueError("Solo administradores pueden crear usuarios.")
    username = (args.get("username") or "").strip()
    password = (args.get("password") or "").strip()
    nombre_completo = (args.get("nombre_completo") or "").strip()
    if not username:
        raise ValueError("Falta el nombre de usuario.")
    if not password or len(password) < 4:
        raise ValueError("La contraseña debe tener al menos 4 caracteres.")
    if not nombre_completo:
        raise ValueError("Falta el nombre completo.")
    if Usuario.query.filter_by(username=username).first():
        raise ValueError(f"El usuario '{username}' ya existe.")

    rol = (args.get("rol") or "usuario").strip().lower()
    if rol not in ("usuario", "admin_sucursal", "admin"):
        raise ValueError(f"Rol inválido: {rol}. Usa 'usuario', 'admin_sucursal' o 'admin'.")

    if usuario.es_admin:
        # Admin global: puede asignar cualquier sucursal y cualquier rol
        sid = args.get("sucursal_id")
        if sid is not None:
            sucursal = Sucursal.query.filter_by(id=int(sid), activa=True).first()
            if not sucursal:
                raise ValueError(f"Sucursal {sid} no existe o no está activa.")
        else:
            sucursal = None
    else:
        # Admin de sucursal: solo en su sucursal, rol forzado a 'usuario'
        if rol != "usuario":
            raise ValueError("Como admin de sucursal solo puedes crear usuarios con rol 'usuario'.")
        sucursal = usuario.sucursal
        if not sucursal:
            raise ValueError("No tienes una sucursal asignada para crear usuarios en ella.")

    return {
        "_titulo": "Confirmar creación de usuario",
        "campos": [
            {"label": "Username", "valor": username},
            {"label": "Nombre completo", "valor": nombre_completo},
            {"label": "Rol", "valor": rol},
            {"label": "Sucursal", "valor": sucursal.nombre if sucursal else "(sin asignar)"},
        ],
        "_advertencia": "Se generará un email automático: " + f"{username}@sisagent.local",
        "username": username, "password": password, "nombre_completo": nombre_completo,
        "sucursal_id": sucursal.id if sucursal else None, "rol": rol,
    }


def _tool_crear_sucursal(args, usuario):
    if not usuario.es_admin:
        raise ValueError("Solo el admin global puede crear sucursales.")
    nombre = (args.get("nombre") or "").strip()
    direccion = (args.get("direccion") or "").strip()
    if not nombre:
        raise ValueError("Falta el nombre de la sucursal.")
    if Sucursal.query.filter_by(nombre=nombre).first():
        raise ValueError(f"Ya existe una sucursal con nombre '{nombre}'.")
    return {
        "_titulo": "Confirmar creación de sucursal",
        "campos": [
            {"label": "Nombre", "valor": nombre},
            {"label": "Dirección", "valor": direccion or "(ninguna)"},
        ],
        "nombre": nombre, "direccion": direccion,
    }


# Registro: nombre -> {handler, requires_confirmation}
CHATBOT_TOOLS = {
    "buscar_productos":       {"handler": _tool_buscar_productos,       "requires_confirmation": False},
    "consultar_precio_stock": {"handler": _tool_consultar_precio_stock, "requires_confirmation": False},
    "resumen_ventas_dia":     {"handler": _tool_resumen_ventas_dia,     "requires_confirmation": False},
    "medios_de_pago":         {"handler": _tool_medios_de_pago,         "requires_confirmation": False},
    "buscar_operaciones":     {"handler": _tool_buscar_operaciones,     "requires_confirmation": False},
    "buscar_ventas":          {"handler": _tool_buscar_ventas,          "requires_confirmation": False},
    "listar_usuarios":        {"handler": _tool_listar_usuarios,        "requires_confirmation": False},
    "listar_sucursales":      {"handler": _tool_listar_sucursales,      "requires_confirmation": False},
    "registrar_venta":         {"handler": _tool_registrar_venta,             "requires_confirmation": True},
    "registrar_operacion":     {"handler": _tool_registrar_operacion,         "requires_confirmation": True},
    "crear_producto":   {"handler": _tool_crear_producto,   "requires_confirmation": True},
    "editar_producto":  {"handler": _tool_editar_producto,  "requires_confirmation": True},
    "eliminar_producto":{"handler": _tool_eliminar_producto,"requires_confirmation": True},
    "eliminar_operacion":{"handler": _tool_eliminar_operacion,"requires_confirmation": True},
    "editar_operacion":  {"handler": _tool_editar_operacion,  "requires_confirmation": True},
    "eliminar_venta":   {"handler": _tool_eliminar_venta,   "requires_confirmation": True},
    "crear_usuario":    {"handler": _tool_crear_usuario,    "requires_confirmation": True},
    "crear_sucursal":   {"handler": _tool_crear_sucursal,   "requires_confirmation": True},
}

# Agregar declaraciones de herramientas (para Claude)
for decl in _HERRAMIENTAS_DECLARACIONES:
    nombre = decl.get("name")
    if nombre in CHATBOT_TOOLS:
        CHATBOT_TOOLS[nombre]["_tool_declaration"] = decl


# ---------- Ejecutores validados (tras confirmacion del usuario) ----------

def _ejecutar_venta_validada(args, usuario):
    """Re-valida desde cero y registra la venta. Espeja la logica de registrar_venta."""
    producto_id = args.get("producto_id")
    try:
        cantidad = int(args.get("cantidad", 0))
    except (TypeError, ValueError):
        raise ValueError("Cantidad invalida.")

    if not producto_id or cantidad <= 0:
        raise ValueError("Datos de la venta incompletos.")

    sucursal = _resolver_sucursal_para_accion(args.get("sucursal_id"), usuario)

    producto = Producto.query.filter_by(id=int(producto_id), activo=True).first()
    if not producto:
        raise ValueError("El producto ya no existe o fue desactivado.")
    if producto.sucursal_id != sucursal.id:
        raise ValueError("El producto no pertenece a esta sucursal.")
    if cantidad > producto.stock:
        raise ValueError(
            f'Stock insuficiente al confirmar: quedan {producto.stock} unidad(es). '
            "Es posible que otro usuario haya registrado una venta antes que tu."
        )

    precio_unitario = float(producto.precio)
    monto = precio_unitario * cantidad

    venta = Venta(
        producto_id=producto.id,
        cantidad=cantidad,
        precio_unitario=precio_unitario,
        monto=monto,
        usuario_id=usuario.id,
        sucursal_id=sucursal.id,
        fecha=get_peru_time().replace(tzinfo=None),
    )
    db.session.add(venta)
    producto.stock = producto.stock - cantidad

    hoy = get_peru_time().date()
    caja = CajaVentas.query.filter_by(fecha=hoy, sucursal_id=sucursal.id).first()
    if caja:
        caja.total_vendido = float(caja.total_vendido) + monto
        caja.saldo = float(caja.saldo) + monto
    else:
        caja = CajaVentas(
            fecha=hoy,
            sucursal_id=sucursal.id,
            total_vendido=monto,
            saldo=monto,
        )
        db.session.add(caja)

    db.session.commit()
    clear_cache()

    # La sucursal solo se nombra a admins (un usuario regular tiene una sola).
    _msg_venta = f'Venta registrada: {cantidad} x "{producto.nombre}" = S/ {monto:.2f}'
    if getattr(usuario, 'es_admin', False):
        _msg_venta += f' en {sucursal.nombre}'
    _msg_venta += '.'
    return {
        "mensaje": _msg_venta,
        "monto": monto,
        "producto_nombre": producto.nombre,
        "cantidad": cantidad,
        "sucursal_nombre": sucursal.nombre,
    }


def _ejecutar_operacion_validada(args, usuario):
    """Re-valida desde cero y registra la operacion bancaria. Espeja registrar_operacion."""
    try:
        monto = float(args.get("monto"))
    except (TypeError, ValueError):
        raise ValueError("Monto invalido.")
    if monto <= 0:
        raise ValueError("El monto debe ser mayor a 0.")

    # Comisión: sugerida por defecto, override si se pasó manualmente
    comision_sugerida = calcular_comision_sugerida(monto)
    es_manual = bool(args.get("comision_manual"))
    if "comision" in args and args.get("comision") is not None:
        try:
            comision = float(args.get("comision"))
        except (TypeError, ValueError):
            raise ValueError("Comision invalida.")
        if comision < 0:
            raise ValueError("La comision no puede ser negativa.")
        if abs(comision - comision_sugerida) < 0.001:
            es_manual = False
        else:
            es_manual = True
    else:
        comision = comision_sugerida
        es_manual = False
    motivo = (args.get("motivo_descuento") or "").strip() or None if es_manual else None

    medio_arg = (args.get("medio") or "").strip().upper()
    if not medio_arg:
        raise ValueError("Falta el medio de pago.")

    sucursal = _resolver_sucursal_para_accion(args.get("sucursal_id"), usuario)

    medio_valido = MedioPago.query.join(
        MedioSucursal,
        (MedioSucursal.medio_pago_id == MedioPago.id) &
        (MedioSucursal.sucursal_id == sucursal.id) &
        (MedioSucursal.activo == True)
    ).filter(
        MedioPago.activo == True,
        db.or_(
            db.func.upper(MedioPago.nombre_abreviado) == medio_arg,
            db.func.upper(MedioPago.nombre_completo) == medio_arg,
        ),
    ).first()

    if not medio_valido:
        raise ValueError(f'El medio "{medio_arg}" ya no esta habilitado en "{sucursal.nombre}".')

    hora_peru = get_peru_time().replace(tzinfo=None)
    operacion = Operacion(
        monto=monto,
        comision=comision,
        medio=medio_valido.nombre_abreviado,
        usuario_id=usuario.id,
        sucursal_id=sucursal.id,
        hora=hora_peru,
        comision_sugerida=comision_sugerida,
        comision_manual=es_manual,
        motivo_descuento=motivo,
    )
    db.session.add(operacion)

    # Actualizar comisiones diarias y mensuales (mirror de registrar_operacion)
    hoy = get_peru_time().date()
    comision_diaria = ComisionDiaria.query.filter_by(fecha=hoy, sucursal_id=sucursal.id).first()
    if comision_diaria:
        comision_diaria.total_comision = float(comision_diaria.total_comision) + comision
    else:
        comision_diaria = ComisionDiaria(fecha=hoy, sucursal_id=sucursal.id, total_comision=comision)
        db.session.add(comision_diaria)

    ahora = get_peru_time()
    comision_mensual = ComisionMensual.query.filter_by(
        año=ahora.year, mes=ahora.month, sucursal_id=sucursal.id
    ).first()
    if comision_mensual:
        comision_mensual.total_comision = float(comision_mensual.total_comision) + comision
    else:
        comision_mensual = ComisionMensual(
            año=ahora.year, mes=ahora.month, sucursal_id=sucursal.id, total_comision=comision
        )
        db.session.add(comision_mensual)

    db.session.commit()

    # Corregir auto-increment si es necesario
    _corregir_autoincrement_operacion()

    clear_cache()

    # La sucursal solo se nombra a admins (un usuario regular tiene una sola, sobra decirla).
    # La comision solo se nombra si fue MANUAL (la automatica no hace falta mencionarla).
    _msg = f'Operacion registrada: S/ {monto:.2f}'
    if es_manual:
        _msg += f' con comision manual de S/ {comision:.2f}'
        if motivo:
            _msg += f' (motivo: {motivo})'
    _msg += f' via {medio_valido.nombre_abreviado}'
    if getattr(usuario, 'es_admin', False):
        _msg += f' en {sucursal.nombre}'
    _msg += '.'
    return {
        "mensaje": _msg,
        "monto": monto,
        "comision": comision,
        "comision_sugerida": comision_sugerida,
        "comision_manual": es_manual,
        "motivo_descuento": motivo,
        "medio": medio_valido.nombre_abreviado,
        "sucursal_nombre": sucursal.nombre,
    }


def _ejecutar_editar_operacion_validada(args, usuario):
    """Re-valida desde cero y edita la operacion bancaria."""
    op_id = args.get("operacion_id")
    if not op_id:
        raise ValueError("Falta operacion_id.")

    operacion = Operacion.query.filter_by(id=int(op_id)).first()
    if not operacion:
        raise ValueError(f"La operación {op_id} ya no existe.")

    # Reglas de permisos:
    if not usuario.es_admin:
        if operacion.sucursal_id != usuario.sucursal_id:
            raise ValueError("No tienes permisos para editar operaciones de otra sucursal.")
        if not usuario.es_admin_de_sucursal() and operacion.usuario_id != usuario.id:
            raise ValueError("Solo puedes editar las operaciones que tú registraste.")

    # Nuevos valores (o mantener los actuales)
    monto_nuevo = operacion.monto
    monto_arg = args.get("monto") or args.get("monto_nuevo")
    if monto_arg is not None:
        try:
            monto_nuevo = float(monto_arg)
        except (TypeError, ValueError):
            raise ValueError("Monto invalido.")
        if monto_nuevo <= 0:
            raise ValueError("El monto debe ser mayor a 0.")

    comision_nueva = operacion.comision
    es_manual = bool(args.get("comision_manual", False))
    motivo = args.get("motivo")

    comision_arg = args.get("comision") or args.get("comision_nueva")
    if comision_arg is not None:
        try:
            comision_nueva = float(comision_arg)
        except (TypeError, ValueError):
            raise ValueError("Comision invalida.")
        if comision_nueva < 0:
            raise ValueError("La comision no puede ser negativa.")
    else:
        # Si no pasaron comisión pero cambiaron monto, recalcular
        if monto_nuevo != operacion.monto:
            comision_nueva = calcular_comision_sugerida(monto_nuevo)
            es_manual = False
            motivo = None

    # Guardar valores anteriores ANTES de actualizar
    monto_anterior = operacion.monto
    comision_anterior = operacion.comision

    # Calcular diferencias
    monto_diferencia = monto_nuevo - monto_anterior
    comision_diferencia = comision_nueva - comision_anterior

    # Actualizar en la operación
    operacion.monto = monto_nuevo
    operacion.comision = comision_nueva
    operacion.comision_manual = es_manual
    operacion.motivo_descuento = motivo if es_manual else None

    # Ajustar comisiones diarias y mensuales
    hoy = get_peru_time().date()
    comision_diaria = ComisionDiaria.query.filter_by(fecha=hoy, sucursal_id=operacion.sucursal_id).first()
    if comision_diaria:
        comision_diaria.total_comision = float(comision_diaria.total_comision) + comision_diferencia
    else:
        # Crear registro si no existe (aunque es raro)
        comision_diaria = ComisionDiaria(
            fecha=hoy,
            sucursal_id=operacion.sucursal_id,
            total_comision=comision_diferencia
        )
        db.session.add(comision_diaria)

    ahora = get_peru_time()
    comision_mensual = ComisionMensual.query.filter_by(
        año=ahora.year, mes=ahora.month, sucursal_id=operacion.sucursal_id
    ).first()
    if comision_mensual:
        comision_mensual.total_comision = float(comision_mensual.total_comision) + comision_diferencia
    else:
        comision_mensual = ComisionMensual(
            año=ahora.year,
            mes=ahora.month,
            sucursal_id=operacion.sucursal_id,
            total_comision=comision_diferencia
        )
        db.session.add(comision_mensual)

    db.session.commit()
    clear_cache()

    # Mensaje
    _msg = f'Operación {operacion.id} editada'
    if monto_diferencia != 0:
        _msg += f': S/ {monto_anterior:.2f} → S/ {monto_nuevo:.2f}'
    if es_manual and comision_diferencia != 0:
        _msg += f'. Comisión manual: S/ {comision_nueva:.2f}'
        if motivo:
            _msg += f' (motivo: {motivo})'
    _msg += '.'

    return {
        "mensaje": _msg,
        "monto": monto_nuevo,
        "monto_anterior": monto_anterior,
        "comision": comision_nueva,
        "comision_anterior": comision_anterior,
        "operacion_id": operacion.id,
    }


def _ejecutar_crear_producto_validado(args, usuario):
    if not usuario.es_admin_o_admin_sucursal():
        raise ValueError("No tienes permisos para gestionar inventario.")
    nombre = (args.get("nombre") or "").strip()
    if not nombre:
        raise ValueError("Nombre vacío.")
    try:
        precio = float(args.get("precio"))
        stock = int(args.get("stock"))
    except (TypeError, ValueError):
        raise ValueError("Precio/stock inválidos.")
    if precio <= 0 or stock < 0:
        raise ValueError("Precio debe ser >0 y stock >=0.")
    sucursal = _resolver_sucursal_para_accion(args.get("sucursal_id"), usuario)
    descripcion = (args.get("descripcion") or "").strip()

    producto = Producto(
        nombre=nombre, descripcion=descripcion, precio=precio, stock=stock,
        sucursal_id=sucursal.id,
    )
    db.session.add(producto)
    db.session.commit()
    clear_cache()
    return {"mensaje": f'Producto "{nombre}" creado en {sucursal.nombre} con stock {stock} a S/ {precio:.2f}.', "producto_id": producto.id}


def _ejecutar_editar_producto_validado(args, usuario):
    if not usuario.es_admin_o_admin_sucursal():
        raise ValueError("Sin permisos.")
    pid = args.get("producto_id")
    if not pid:
        raise ValueError("Falta producto_id.")
    producto = Producto.query.filter_by(id=int(pid), activo=True).first()
    if not producto:
        raise ValueError("El producto ya no existe.")
    if not usuario.es_admin and producto.sucursal_id != usuario.sucursal_id:
        raise ValueError("No puedes editar productos de otra sucursal.")

    cambios = []
    nuevo_nombre = args.get("nombre")
    if nuevo_nombre and nuevo_nombre.strip() and nuevo_nombre.strip() != producto.nombre:
        producto.nombre = nuevo_nombre.strip(); cambios.append("nombre")
    nueva_desc = args.get("descripcion")
    if nueva_desc is not None and nueva_desc.strip() != (producto.descripcion or ""):
        producto.descripcion = nueva_desc.strip(); cambios.append("descripción")
    if args.get("precio") is not None:
        p = float(args["precio"])
        if p <= 0: raise ValueError("Precio debe ser >0.")
        if abs(p - float(producto.precio)) > 0.001:
            producto.precio = p; cambios.append("precio")
    if args.get("stock") is not None:
        st = int(args["stock"])
        if st < 0: raise ValueError("Stock no puede ser negativo.")
        if st != producto.stock:
            producto.stock = st; cambios.append("stock")

    if not cambios:
        raise ValueError("Nada que cambiar.")
    db.session.commit()
    clear_cache()
    return {"mensaje": f'Producto "{producto.nombre}" actualizado ({", ".join(cambios)}).'}


def _ejecutar_eliminar_producto_validado(args, usuario):
    if not usuario.es_admin_o_admin_sucursal():
        raise ValueError("Sin permisos.")
    pid = args.get("producto_id")
    if not pid:
        raise ValueError("Falta producto_id.")
    producto = Producto.query.filter_by(id=int(pid), activo=True).first()
    if not producto:
        raise ValueError("El producto ya no existe o fue eliminado por otro usuario.")
    if not usuario.es_admin and producto.sucursal_id != usuario.sucursal_id:
        raise ValueError("No puedes eliminar productos de otra sucursal.")
    nombre = producto.nombre
    producto.activo = False
    db.session.commit()
    clear_cache()
    return {"mensaje": f'Producto "{nombre}" eliminado del inventario.'}


def _ejecutar_eliminar_operacion_validada(args, usuario):
    op_id = args.get("operacion_id")
    if not op_id:
        raise ValueError("Falta operacion_id.")
    operacion = Operacion.query.filter_by(id=int(op_id)).first()
    if not operacion:
        raise ValueError("La operación ya no existe.")
    if not usuario.es_admin:
        if operacion.sucursal_id != usuario.sucursal_id:
            raise ValueError("Sin permisos para esta sucursal.")
        if not usuario.es_admin_de_sucursal() and operacion.usuario_id != usuario.id:
            raise ValueError("Solo puedes eliminar las operaciones que tú registraste.")
    # Revertir comisión de los totales diarios y mensuales
    comision = float(operacion.comision)
    hoy = operacion.hora.date() if operacion.hora else get_peru_time().date()
    cd = ComisionDiaria.query.filter_by(fecha=hoy, sucursal_id=operacion.sucursal_id).first()
    if cd:
        cd.total_comision = max(0.0, float(cd.total_comision) - comision)
    cm = ComisionMensual.query.filter_by(
        año=hoy.year, mes=hoy.month, sucursal_id=operacion.sucursal_id
    ).first()
    if cm:
        cm.total_comision = max(0.0, float(cm.total_comision) - comision)

    monto = float(operacion.monto)
    medio = operacion.medio
    db.session.delete(operacion)
    db.session.commit()
    clear_cache()
    return {"mensaje": f'Operación eliminada: S/ {monto:.2f} vía {medio}. Comisión S/ {comision:.2f} restada de los totales.'}


def _ejecutar_eliminar_venta_validada(args, usuario):
    v_id = args.get("venta_id")
    if not v_id:
        raise ValueError("Falta venta_id.")
    venta = Venta.query.filter_by(id=int(v_id)).first()
    if not venta:
        raise ValueError("La venta ya no existe.")
    if not usuario.es_admin:
        if venta.sucursal_id != usuario.sucursal_id:
            raise ValueError("Sin permisos para esta sucursal.")
        if not usuario.es_admin_de_sucursal() and venta.usuario_id != usuario.id:
            raise ValueError("Solo puedes eliminar las ventas que tú registraste.")

    # Devolver stock al producto (si existe)
    if venta.producto:
        venta.producto.stock = (venta.producto.stock or 0) + venta.cantidad

    # Restar de la caja de ventas
    fecha = venta.fecha.date() if venta.fecha else get_peru_time().date()
    caja = CajaVentas.query.filter_by(fecha=fecha, sucursal_id=venta.sucursal_id).first()
    if caja:
        caja.total_vendido = max(0.0, float(caja.total_vendido) - float(venta.monto))
        caja.saldo = max(0.0, float(caja.saldo) - float(venta.monto))

    nombre_prod = venta.producto.nombre if venta.producto else "(eliminado)"
    cantidad = venta.cantidad
    monto = float(venta.monto)
    db.session.delete(venta)
    db.session.commit()
    clear_cache()
    return {"mensaje": f'Venta eliminada: {cantidad} × "{nombre_prod}" = S/ {monto:.2f}. Stock devuelto al inventario.'}


def _ejecutar_crear_usuario_validado(args, usuario):
    if not usuario.es_admin_o_admin_sucursal():
        raise ValueError("Sin permisos.")
    username = (args.get("username") or "").strip()
    password = (args.get("password") or "").strip()
    nombre_completo = (args.get("nombre_completo") or "").strip()
    if not (username and password and nombre_completo):
        raise ValueError("Datos incompletos.")
    if Usuario.query.filter_by(username=username).first():
        raise ValueError(f"El usuario '{username}' ya existe (creado por otro mientras tanto).")

    rol = (args.get("rol") or "usuario").strip().lower()
    if usuario.es_admin:
        sid = args.get("sucursal_id")
        if sid is not None:
            sucursal = Sucursal.query.filter_by(id=int(sid), activa=True).first()
            if not sucursal:
                raise ValueError(f"Sucursal {sid} ya no existe o no está activa.")
            sucursal_id = sucursal.id
        else:
            sucursal_id = None
        es_admin = (rol == "admin")
        es_admin_sucursal = (rol == "admin_sucursal")
    else:
        if rol != "usuario":
            raise ValueError("Solo admin global puede crear con rol distinto a 'usuario'.")
        if not usuario.sucursal:
            raise ValueError("No tienes sucursal asignada.")
        sucursal_id = usuario.sucursal_id
        es_admin = False
        es_admin_sucursal = False

    nuevo = Usuario(
        username=username,
        email=f"{username}@sisagent.local",
        password_hash=generate_password_hash(password),
        nombre_completo=nombre_completo,
        sucursal_id=sucursal_id,
        es_admin=es_admin,
        es_admin_sucursal=es_admin_sucursal,
    )
    db.session.add(nuevo)
    db.session.commit()
    clear_cache()
    sucursal_nombre = nuevo.sucursal.nombre if nuevo.sucursal else "(sin sucursal)"
    return {"mensaje": f'Usuario "{username}" creado como {rol} en {sucursal_nombre}.', "usuario_id": nuevo.id}


def _ejecutar_crear_sucursal_validada(args, usuario):
    if not usuario.es_admin:
        raise ValueError("Solo el admin global puede crear sucursales.")
    nombre = (args.get("nombre") or "").strip()
    direccion = (args.get("direccion") or "").strip()
    if not nombre:
        raise ValueError("Nombre vacío.")
    if Sucursal.query.filter_by(nombre=nombre).first():
        raise ValueError(f"Ya existe una sucursal '{nombre}'.")
    suc = Sucursal(nombre=nombre, direccion=direccion, activa=True)
    db.session.add(suc)
    db.session.commit()
    clear_cache()
    return {"mensaje": f'Sucursal "{nombre}" creada exitosamente.', "sucursal_id": suc.id}


# Dispatcher de ejecutores directos: nombre de proponer_* -> funcion validada que ejecuta de inmediato.
# Usado tanto por el chat de texto como por voz para evitar el paso de confirmacion.
EJECUTORES_DIRECTOS = {
    'registrar_venta':              _ejecutar_venta_validada,
    'registrar_operacion':          _ejecutar_operacion_validada,
    'crear_producto':     _ejecutar_crear_producto_validado,
    'editar_producto':    _ejecutar_editar_producto_validado,
    'eliminar_producto':  _ejecutar_eliminar_producto_validado,
    'eliminar_operacion': _ejecutar_eliminar_operacion_validada,
    'editar_operacion':   _ejecutar_editar_operacion_validada,
    'eliminar_venta':     _ejecutar_eliminar_venta_validada,
    'crear_usuario':      _ejecutar_crear_usuario_validado,
    'crear_sucursal':     _ejecutar_crear_sucursal_validada,
}


# ---------- Cliente Anthropic (lazy) y loop de function-calling ----------

_anthropic_client = None

def _get_anthropic_client():
    global _anthropic_client
    if not app.config.get('ANTHROPIC_API_KEY'):
        raise RuntimeError(
            "El asistente IA no esta disponible: falta configurar ANTHROPIC_API_KEY en el servidor."
        )
    if _anthropic_client is None:
        _anthropic_client = anthropic.Anthropic(
            api_key=app.config['ANTHROPIC_API_KEY'],
            timeout=30.0,
        )
    return _anthropic_client


def _construir_mensajes_chat(historial, mensaje_usuario, imagen_bytes=None, imagen_mimetype=None):
    """Construye el array `contents` para Gemini (function-calling), incluyendo el turno actual.

    El historial del cliente es texto plano (sin function_call) — las imagenes pasadas no se reenvian.
    Gemini usa el rol "model" para las respuestas del asistente (no "assistant").
    """
    contents = []

    for turno in (historial or []):
        if not isinstance(turno, dict):
            continue
        rol = turno.get("role")
        contenido = turno.get("content")
        if rol not in ("user", "assistant") or not contenido or not isinstance(contenido, str):
            continue
        contents.append({
            "role": "model" if rol == "assistant" else "user",
            "parts": [{"text": contenido}],
        })

    # Turno actual
    parts = []
    if imagen_bytes and imagen_mimetype:
        b64 = base64.standard_b64encode(imagen_bytes).decode("ascii")
        parts.append({"inlineData": {"mimeType": imagen_mimetype, "data": b64}})
    parts.append({"text": mensaje_usuario or "(El usuario adjunto una imagen)"})

    contents.append({"role": "user", "parts": parts})
    return contents


def _ejecutar_turno_chat(mensajes, usuario, max_iteraciones=4):
    """Loop de tool-use con Claude API. Devuelve dict con 'tipo': 'texto'.

    Usa Anthropic Claude en lugar de Gemini para mejor obediencia de instrucciones.
    Los mensajes llegan en formato Gemini (role: user/model, parts: [...]) y se convierten a Claude.
    """
    if not app.config.get("ANTHROPIC_API_KEY"):
        raise RuntimeError("El asistente IA no esta disponible: falta configurar ANTHROPIC_API_KEY en el servidor.")

    client = _get_anthropic_client()

    productos_buscados = []
    productos_ids_vistos = set()

    def _registrar_productos(lista):
        for prod in (lista or []):
            pid = prod.get("id")
            if pid is None or pid in productos_ids_vistos:
                continue
            productos_ids_vistos.add(pid)
            productos_buscados.append(prod)

    # Convertir mensajes de formato Gemini a Claude
    # Gemini: {"role": "user"/"model", "parts": [{"text": "..."}, ...]}
    # Claude: {"role": "user"/"assistant", "content": "text"} o con tool_use blocks
    mensajes_claude = []
    for msg in mensajes:
        rol = msg.get("role", "").lower()
        parts = msg.get("parts") or []
        if not parts:
            continue

        # Extraer texto de los parts
        textos = []
        for part in parts:
            if isinstance(part, dict):
                if "text" in part:
                    textos.append(part["text"])
                # Ignorar inlineData por ahora (fotos) — Claude lo maneja diferente

        contenido = " ".join(textos).strip()
        if not contenido:
            continue

        if rol == "user":
            mensajes_claude.append({"role": "user", "content": contenido})
        elif rol == "model":
            mensajes_claude.append({"role": "assistant", "content": contenido})

    # Construir herramientas para Claude
    herramientas = []
    for nombre, spec in CHATBOT_TOOLS.items():
        decl = spec.get("_tool_declaration", {})
        if decl:
            herramientas.append({
                "name": nombre,
                "description": decl.get("description", ""),
                "input_schema": decl.get("input_schema", {})
            })

    system_prompt = SYSTEM_PROMPT_CHATBOT + _contexto_fecha_hora()

    for iteracion in range(max_iteraciones):
        response = client.messages.create(
            model="claude-opus-4-8-20250514",
            max_tokens=2048,
            system=system_prompt,
            tools=herramientas if herramientas else None,
            messages=mensajes_claude
        )

        # Procesar respuesta de Claude
        texto = ""
        tool_calls = []

        for block in response.content:
            if hasattr(block, "text"):
                texto = (block.text or "").strip()
            elif hasattr(block, "type") and block.type == "tool_use":
                tool_calls.append({
                    "name": block.name,
                    "id": block.id,
                    "input": block.input or {}
                })

        # Sin llamadas a herramienta -> respuesta final
        if not tool_calls:
            return {
                "tipo": "texto",
                "texto": texto or "(Sin respuesta)",
                "productos": productos_buscados,
            }

        # Procesar tool calls
        herramientas_ejecucion_directa = list(EJECUTORES_DIRECTOS.keys())
        tool_calls_no_mutativos = []

        for tool_call in tool_calls:
            nombre = tool_call.get("name")
            fargs = tool_call.get("input") or {}

            # Herramientas que ejecutan directamente (sin confirmacion)
            if nombre in herramientas_ejecucion_directa:
                try:
                    resultado = EJECUTORES_DIRECTOS[nombre](fargs, usuario)
                except ValueError as e:
                    return {
                        "tipo": "texto",
                        "texto": f"No se pudo completar la accion: {str(e)}",
                        "productos": productos_buscados,
                    }
                return {
                    "tipo": "texto",
                    "texto": resultado.get("mensaje", "Listo."),
                    "productos": productos_buscados,
                }

            # Herramientas de solo lectura - ejecutar y continuar loop
            spec = CHATBOT_TOOLS.get(nombre)
            if spec:
                try:
                    resultado = spec["handler"](fargs, usuario)
                    if not isinstance(resultado, dict):
                        resultado = {"resultado": resultado}
                    if nombre == "buscar_productos":
                        _registrar_productos(resultado.get("productos"))
                    elif nombre == "consultar_precio_stock":
                        prod = resultado.get("producto")
                        if prod:
                            _registrar_productos([prod])
                except (ValueError, Exception) as e:
                    resultado = {"error": str(e)}

                tool_calls_no_mutativos.append({
                    "tool_use_id": tool_call.get("id"),
                    "content": resultado
                })

        # Si hay tool calls para procesar, agregar al historial y continuar
        if tool_calls_no_mutativos:
            # Agregar respuesta del asistente con los tool_use blocks
            # Convertir response.content a un formato que Claude pueda procesar
            assistant_content = []
            for block in response.content:
                if hasattr(block, "text") and block.text:
                    assistant_content.append({"type": "text", "text": block.text})
                elif hasattr(block, "type") and block.type == "tool_use":
                    assistant_content.append({
                        "type": "tool_use",
                        "id": block.id,
                        "name": block.name,
                        "input": block.input
                    })

            mensajes_claude.append({"role": "assistant", "content": assistant_content})

            # Agregar resultados de las herramientas
            tool_results = []
            for tr in tool_calls_no_mutativos:
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tr["tool_use_id"],
                    "content": json.dumps(tr["content"])
                })

            mensajes_claude.append({
                "role": "user",
                "content": tool_results
            })

    return {
        "tipo": "texto",
        "texto": "He alcanzado el límite de iteraciones.",
        "productos": productos_buscados,
    }


# ---------- Rutas HTTP del chatbot ----------

@app.route('/api/chat/mensaje', methods=['POST'])
@login_required
def api_chat_mensaje():
    """Procesa un mensaje del chat. Acepta multipart/form-data con `mensaje`, `historial` (JSON) e `imagen` (opcional)."""
    try:
        mensaje = (request.form.get('mensaje') or '').strip()
        historial_json = request.form.get('historial') or '[]'
        try:
            historial = json.loads(historial_json)
            if not isinstance(historial, list):
                historial = []
        except json.JSONDecodeError:
            historial = []

        imagen_bytes = None
        imagen_mimetype = None
        imagen_file = request.files.get('imagen')
        if imagen_file and imagen_file.filename:
            mimetype = (imagen_file.mimetype or '').lower()
            if mimetype not in FOTO_TIPOS_PERMITIDOS:
                return jsonify({
                    'success': False,
                    'message': f'Tipo de imagen no permitido: {mimetype}. Usa png/jpg/gif/webp.',
                }), 400
            imagen_bytes = imagen_file.read()
            if len(imagen_bytes) > FOTO_CHAT_TAMANO_MAXIMO:
                return jsonify({
                    'success': False,
                    'message': f'La imagen excede el tamano maximo de {FOTO_CHAT_TAMANO_MAXIMO // (1024*1024)} MB.',
                }), 400
            imagen_mimetype = mimetype

        if not mensaje and not imagen_bytes:
            return jsonify({'success': False, 'message': 'Escribe un mensaje o adjunta una imagen.'}), 400

        mensajes = _construir_mensajes_chat(historial, mensaje, imagen_bytes, imagen_mimetype)
        resultado = _ejecutar_turno_chat(mensajes, current_user)

        return jsonify({'success': True, **resultado})

    except RuntimeError as e:
        return jsonify({'success': False, 'message': str(e)}), 503
    except anthropic.APIStatusError as e:
        return jsonify({
            'success': False,
            'message': f'Error de la IA ({e.status_code}). Intenta de nuevo en un momento.',
        }), 502
    except anthropic.APIConnectionError:
        return jsonify({
            'success': False,
            'message': 'No se pudo conectar con la IA. Verifica tu conexion a internet.',
        }), 502
    except Exception as e:
        import traceback
        traceback.print_exc()
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({
            'success': False,
            'message': f'Error procesando el mensaje: {str(e)}',
        }), 500


@app.route('/api/chat/confirmar_accion', methods=['POST'])
@login_required
def api_chat_confirmar_accion():
    """Ejecuta una accion previamente propuesta tras confirmacion explicita del usuario.

    REVALIDA todo desde cero — nunca confia en los args que viajaron por el cliente.
    """
    try:
        data = request.get_json(silent=True) or {}
        accion = data.get('accion')
        args = data.get('args') or {}

        if accion not in CHATBOT_TOOLS:
            return jsonify({'success': False, 'message': 'Accion desconocida.'}), 400
        if not CHATBOT_TOOLS[accion]['requires_confirmation']:
            return jsonify({'success': False, 'message': 'Esta accion no requiere confirmacion.'}), 400

        ejecutor = EJECUTORES_DIRECTOS.get(accion)
        if not ejecutor:
            return jsonify({'success': False, 'message': f'Accion {accion} no implementada.'}), 400

        resultado = ejecutor(args, current_user)
        return jsonify({'success': True, **resultado})

    except ValueError as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({
            'success': False,
            'message': f'Error ejecutando la accion: {str(e)}',
        }), 500


def _convertir_audio_a_wav_16k_mono(audio_bytes):
    """Convierte cualquier audio a WAV 16kHz mono via PyAV.

    Gemini acepta nativamente wav/mp3/aiff/aac/ogg/flac, pero los navegadores
    producen webm/opus que NO está en esa lista. Convertimos a WAV (formato
    universal que Gemini procesa eficientemente).
    """
    import io
    import av
    input_container = av.open(io.BytesIO(audio_bytes))
    try:
        audio_stream = next((s for s in input_container.streams if s.type == 'audio'), None)
        if audio_stream is None:
            raise ValueError('El archivo no contiene pista de audio.')

        output_buf = io.BytesIO()
        output_container = av.open(output_buf, mode='w', format='wav')
        output_stream = output_container.add_stream('pcm_s16le', rate=16000)
        output_stream.layout = 'mono'

        resampler = av.AudioResampler(format='s16', layout='mono', rate=16000)

        for frame in input_container.decode(audio_stream):
            for rs_frame in resampler.resample(frame):
                for packet in output_stream.encode(rs_frame):
                    output_container.mux(packet)
        # Flush
        for packet in output_stream.encode(None):
            output_container.mux(packet)
        output_container.close()

        return output_buf.getvalue()
    finally:
        input_container.close()


# ---------- "Entrenamiento" de voz: vocabulario y correcciones por usuario ----------

def _parsear_vocabulario_voz(texto):
    """Separa el vocabulario del usuario en (terminos, correcciones).

    Cada linea con 'mal => bien' (o 'mal -> bien') es una correccion;
    el resto son terminos de vocabulario (nombres, jerga, frases).
    """
    terminos, correcciones = [], []
    for linea in (texto or '').splitlines():
        linea = linea.strip()
        if not linea or linea.startswith('#'):
            continue
        sep = '=>' if '=>' in linea else ('->' if '->' in linea else None)
        if sep:
            izq, der = linea.split(sep, 1)
            izq, der = izq.strip(), der.strip()
            if izq and der:
                correcciones.append((izq, der))
        else:
            terminos.append(linea)
    return terminos[:80], correcciones[:80]


def _aplicar_correcciones_voz(texto, correcciones):
    """Reemplaza mistranscripciones conocidas (insensible a may/min)."""
    import re
    out = texto or ''
    for mal, bien in correcciones:
        try:
            out = re.sub(re.escape(mal), bien, out, flags=re.IGNORECASE)
        except Exception:
            pass
    return out


def _vocabulario_de(usuario):
    """Devuelve (terminos, correcciones) del usuario, tolerante a columna ausente."""
    try:
        return _parsear_vocabulario_voz(getattr(usuario, 'vocabulario_voz', '') or '')
    except Exception:
        return [], []


# Cache simple del vocabulario dinamico por usuario (se invalida por TTL corto).
_vocab_dinamico_cache = {}


def _vocabulario_dinamico(usuario):
    """Vocabulario de reconocimiento de voz derivado AUTOMATICAMENTE de la BD del sistema
    (sucursales, productos y medios de pago que el usuario ve). Es 'speech adaptation':
    sesga el modelo hacia las entidades reales sin que el usuario escriba nada.

    Cacheado 120s por usuario para no golpear la BD en cada chunk del wake-word.
    """
    try:
        uid = getattr(usuario, 'id', None)
        if uid is not None:
            entry = _vocab_dinamico_cache.get(uid)
            if entry and (time.time() - entry[0] < 120):
                return entry[1]

        terminos = []
        sucursales = sucursales_visibles_para(usuario)
        suc_ids = [s.id for s in sucursales if s]
        for s in sucursales:
            if s and s.nombre:
                terminos.append(s.nombre.strip())
        for m in MedioPago.query.filter_by(activo=True).all():
            if m.nombre_abreviado:
                terminos.append(m.nombre_abreviado.strip())
            if m.nombre_completo:
                terminos.append(m.nombre_completo.strip())
        if suc_ids:
            prods = (Producto.query
                     .filter(Producto.activo == True, Producto.sucursal_id.in_(suc_ids))
                     .order_by(Producto.nombre).limit(120).all())
            for p in prods:
                if p.nombre:
                    terminos.append(p.nombre.strip())

        vistos, out = set(), []
        for t in terminos:
            k = t.lower()
            if t and k not in vistos:
                vistos.add(k)
                out.append(t)
        out = out[:150]

        if uid is not None:
            _vocab_dinamico_cache[uid] = (time.time(), out)
        return out
    except Exception:
        return []


def registrar_pronunciacion_aprendida(usuario_id, termino_original, termino_correcto, tipo='sucursal'):
    """Registra una pronunciación aprendida en el banco de memoria del usuario."""
    try:
        # Buscar si ya existe esta pronunciación
        pron = PronunciacionAprendida.query.filter_by(
            usuario_id=usuario_id,
            termino_original=termino_original,
            termino_correcto=termino_correcto
        ).first()

        if pron:
            # Incrementar frecuencia
            pron.frecuencia = (pron.frecuencia or 0) + 1
        else:
            # Crear nueva
            pron = PronunciacionAprendida(
                usuario_id=usuario_id,
                termino_original=termino_original,
                termino_correcto=termino_correcto,
                tipo=tipo,
                frecuencia=1
            )
            db.session.add(pron)

        db.session.commit()
        return True
    except Exception as e:
        print(f"[WARN] Error registrando pronunciación: {e}")
        try:
            db.session.rollback()
        except:
            pass
        return False


def obtener_pronunciaciones_aprendidas(usuario_id):
    """Obtiene el banco de pronunciaciones aprendidas del usuario (para pasar al system prompt)."""
    try:
        prons = PronunciacionAprendida.query.filter_by(usuario_id=usuario_id).all()
        if not prons:
            return []
        # Formato: "Tecnovation → TECKNOVATION", ordenadas por frecuencia (más aprendidas primero)
        prons.sort(key=lambda p: p.frecuencia, reverse=True)
        return [f'{p.termino_original} → {p.termino_correcto}' for p in prons[:50]]  # máx 50
    except Exception:
        return []


@app.route('/api/chat/vocabulario', methods=['GET', 'POST'])
@login_required
def api_chat_vocabulario():
    """Lee/guarda el vocabulario de voz ('entrenamiento') del usuario actual."""
    try:
        if request.method == 'POST':
            data = request.get_json(silent=True) or {}
            texto = (data.get('vocabulario') or '').strip()[:4000]
            current_user.vocabulario_voz = texto
            db.session.commit()
            return jsonify({'success': True})
        return jsonify({'success': True, 'vocabulario': getattr(current_user, 'vocabulario_voz', '') or ''})
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': f'No se pudo guardar el vocabulario: {str(e)}'}), 500


@app.route('/api/chat/aprender_pronunciacion', methods=['POST'])
@login_required
def api_chat_aprender_pronunciacion():
    """Registra una pronunciación aprendida (cómo el usuario dice algo vs. cómo está en el sistema)."""
    try:
        data = request.get_json(silent=True) or {}
        termino_original = (data.get('termino_original') or '').strip()
        termino_correcto = (data.get('termino_correcto') or '').strip()
        tipo = (data.get('tipo') or 'sucursal').strip()

        if not termino_original or not termino_correcto:
            return jsonify({'success': False, 'message': 'Faltan términos'}), 400

        if registrar_pronunciacion_aprendida(current_user.id, termino_original, termino_correcto, tipo):
            return jsonify({'success': True, 'message': f'Aprendido: {termino_original} → {termino_correcto}'})
        else:
            return jsonify({'success': False, 'message': 'Error al registrar pronunciación'}), 500
    except Exception as e:
        try:
            db.session.rollback()
        except:
            pass
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@app.route('/api/chat/transcribir', methods=['POST'])
@login_required
def api_chat_transcribir():
    """Transcribe un blob de audio del cliente usando Google Gemini.

    Capa gratuita: 1500 requests/día, sin tarjeta. Soporta audios largos.
    Recibe multipart/form-data con campo `audio`. Devuelve
    {'success': True, 'texto': '...'} o {'success': False, 'message': '...'}.
    """
    try:
        if not app.config.get('GEMINI_API_KEY'):
            return jsonify({
                'success': False,
                'message': 'La transcripcion por voz no esta disponible: falta configurar GEMINI_API_KEY. Obten una key gratis en https://aistudio.google.com/app/apikey',
            }), 503

        audio = request.files.get('audio')
        if not audio or not audio.filename:
            return jsonify({'success': False, 'message': 'No se recibio archivo de audio.'}), 400

        audio_bytes = audio.read()
        if not audio_bytes:
            return jsonify({'success': False, 'message': 'El audio esta vacio.'}), 400
        if len(audio_bytes) > GEMINI_MAX_AUDIO_BYTES:
            return jsonify({
                'success': False,
                'message': f'El audio excede {GEMINI_MAX_AUDIO_BYTES // (1024*1024)}MB. Graba un fragmento mas corto.',
            }), 400

        # Convertir a WAV 16kHz mono (Gemini no acepta webm directo)
        try:
            wav_bytes = _convertir_audio_a_wav_16k_mono(audio_bytes)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'message': f'No se pudo decodificar el audio: {str(e)}',
            }), 400

        # Llamar a Gemini con audio embebido (inline_data + base64)
        import httpx as _httpx_local
        url = f'{GEMINI_API_URL}/{GEMINI_TRANSCRIPTION_MODEL}:generateContent'
        gemini_headers = {"x-goog-api-key": app.config["GEMINI_API_KEY"]}

        # "Entrenamiento" por prompt (vocabulary biasing): contexto del dominio bancario
        # + la frase activadora del usuario como pista, para que el modelo transcriba
        # bien palabras como "soles", "Yape", montos, y reconozca la frase aunque suene aproximada.
        es_wakeword = (request.form.get('modo') or '') == 'wakeword'

        if es_wakeword:
            # Wake-word: transcripcion LITERAL, SIN sesgo de dominio. El sesgo de vocabulario
            # bancario hacia que el modelo alucinara frases ("el monto de la venta es de 200 soles")
            # ante un audio corto. Aqui solo queremos lo que se dijo, tal cual, o vacio.
            prompt_transcripcion = (
                'Transcribe LITERALMENTE el audio en espanol, palabra por palabra, exactamente lo que '
                'se oye. NO interpretes, NO completes frases, NO agregues contexto, NO inventes. '
                'Devuelve solo el texto crudo, sin puntuacion extra ni formato. Si es silencio, ruido, '
                'respiracion o no se entiende, devuelve CADENA VACIA. Nunca adivines.'
            )
        else:
            prompt_transcripcion = (
                'Transcribe el siguiente audio en espanol (Peru). CONTEXTO: es un sistema bancario y de '
                'ventas; vocabulario frecuente: soles, monto, comision, operacion, venta, producto, stock, '
                'sucursal, registrar, eliminar, Yape, Plin, BCP, Interbank, BBVA, efectivo, tarjeta, '
                'transferencia. Los montos se dicen como numero + "soles" (ej: "cien soles", "600 soles"). '
                'Devuelve EXCLUSIVAMENTE el texto transcrito tal cual se dijo, sin comentarios, sin notas, '
                'sin formato Markdown, sin marcas de tiempo, sin formato de subtitulos (nada de "00:00" ni '
                '"-->"), sin prefijos como "Transcripcion:". Si no hay voz audible o solo hay ruido, '
                'devuelve cadena vacia.'
                ' CRITICO: si el audio es silencio, ruido, respiracion o dudoso, devuelve SIEMPRE '
                'cadena vacia. NUNCA adivines ni inventes palabras.'
            )
            # Speech adaptation: vocabulario de la BD SOLO para comandos (no para el wake-word).
            vocab_terminos = _vocabulario_dinamico(current_user)
            if vocab_terminos:
                prompt_transcripcion += (
                    ' NOMBRES REALES del sistema (sucursales, productos y medios de pago); si oyes algo '
                    'parecido a uno de estos, escribelo EXACTAMENTE asi: ' + ', '.join(vocab_terminos) + '.'
                )

        payload = {
            'contents': [{
                'parts': [
                    {'text': prompt_transcripcion},
                    {
                        'inline_data': {
                            'mime_type': 'audio/wav',
                            'data': base64.b64encode(wav_bytes).decode('ascii'),
                        }
                    },
                ]
            }],
            'generationConfig': {
                'temperature': 0.0,
                'maxOutputTokens': 2048,
            },
        }

        # Llamada a Gemini con reintentos automáticos para errores transitorios de red
        # (getaddrinfo failed, connection reset, etc.) — comunes en Windows con httpx.
        resp = None
        last_error = None
        for intento in range(3):
            try:
                resp = _httpx_local.post(url, json=payload, headers=gemini_headers, timeout=90.0)
                break  # éxito
            except _httpx_local.TimeoutException as e:
                return jsonify({
                    'success': False,
                    'message': 'Gemini tardo demasiado en responder. Intenta con un audio mas corto.',
                }), 504
            except (_httpx_local.ConnectError, _httpx_local.ReadError,
                    _httpx_local.RemoteProtocolError, OSError) as e:
                # Errores transitorios — reintentar con backoff (0.5s, 1.5s)
                last_error = e
                if intento < 2:
                    time.sleep(0.5 + intento)
                    continue
            except _httpx_local.HTTPError as e:
                return jsonify({
                    'success': False,
                    'message': f'No se pudo conectar con Gemini: {str(e)}',
                }), 502

        if resp is None:
            return jsonify({
                'success': False,
                'message': (
                    f'No se pudo conectar con Gemini despues de 3 intentos: {str(last_error)}. '
                    'Verifica tu conexion a internet y vuelve a intentar.'
                ),
            }), 502

        if resp.status_code == 400:
            try:
                err = resp.json().get('error', {}).get('message', '')[:300]
            except Exception:
                err = resp.text[:200]
            if 'API key' in err or 'API_KEY' in err:
                return jsonify({
                    'success': False,
                    'message': f'GEMINI_API_KEY invalida. Genera una nueva en https://aistudio.google.com/app/apikey',
                }), 502
            return jsonify({'success': False, 'message': f'Gemini rechazo la peticion: {err}'}), 400
        if resp.status_code == 429:
            return jsonify({
                'success': False,
                'message': 'Excediste la cuota gratuita de Gemini (1500/dia). Espera unos minutos.',
            }), 429
        if resp.status_code != 200:
            return jsonify({
                'success': False,
                'message': f'Gemini error HTTP {resp.status_code}: {resp.text[:200]}',
            }), 502

        try:
            data = resp.json()
            if not data.get('candidates'):
                # Posible bloqueo de safety filters
                prompt_feedback = data.get('promptFeedback', {})
                block_reason = prompt_feedback.get('blockReason')
                if block_reason:
                    return jsonify({
                        'success': False,
                        'message': f'Gemini bloqueo el contenido: {block_reason}',
                    }), 400
                return jsonify({'success': False, 'message': 'Gemini no devolvio resultado.'}), 502

            candidate = data['candidates'][0]
            texto_partes = []
            for part in candidate.get('content', {}).get('parts', []):
                if 'text' in part:
                    texto_partes.append(part['text'])
            texto = ''.join(texto_partes).strip()
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Respuesta inesperada de Gemini: {str(e)}',
            }), 502

        return jsonify({'success': True, 'texto': texto})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error procesando el audio: {str(e)}',
        }), 500


# ---------- System prompt y herramientas para Gemini Live (voz-a-voz) ----------

SYSTEM_PROMPT_GEMINI_LIVE = """Eres el asistente vocal de SISAGENT, un sistema bancario y de ventas en Perú. Respondes hablando.

Modelo de datos:
- Operaciones bancarias: monto (S/), comision (S/), medio (EFECTIVO, YAPE, PLIN, TARJETA, BCP, IBK, BBVA, etc. — segun lo habilitado por sucursal). NO existe campo "tipo de operacion".
- Comision AUTOMATICA: se calcula sola (S/1 por cada S/100 de monto, redondeado hacia arriba). Al registrar una operacion NUNCA preguntes ni menciones la comision, salvo que el usuario pida explicitamente un descuento o una comision distinta a la automatica (ej: "es casero, cobrale solo 1 sol", "hazle descuento"). En ese caso pasa `comision` y `motivo_descuento` a `registrar_operacion`.
- Productos: nombre, descripcion, precio, stock, foto, sucursal asignada.
- Ventas: producto x cantidad, descuenta stock, suma a caja diaria.
- Sucursales: nombre, direccion, medios de pago habilitados. IMPORTANTE: Si el usuario menciona una sucursal que no reconoces de la lista que te pasé, NO la rechaces. Asume que existe. El usuario sabe su propia sucursal mejor que tú — no discutas por el nombre.
- Usuarios: rol (admin global, admin de sucursal, usuario regular).

Tu rol:
1. Consultar informacion (productos, operaciones, ventas, usuarios, sucursales, medios de pago).
2. Ejecutar acciones (registrar/eliminar venta u operacion, crear/editar/eliminar producto, crear usuario, crear sucursal, etc.) — PERO con confirmacion verbal del usuario.

Reglas para CUALQUIER accion que MUTE datos (registrar/eliminar/crear/editar venta, operacion, producto, usuario, sucursal, etc. via `proponer_*`):
- FLUJO CRÍTICO: (1) PROPÓN (llamando proponer_*) (2) ESPERA confirmación del usuario (3) EJECUTA (el servidor hace el registro) (4) RECIÉN ENTONCES responde "Listo, registré..."
- NUNCA digas "Listo, registré" ANTES de que el usuario confirme. NUNCA confundas "propuesta" con "ejecución".
- Cuando usuario diga "S/ 100 en BCP": llama `registrar_operacion(...)` → muestra preview → aguarda "sí/confirma/dale" → servidor ejecuta → responde "Listo, registré"
- NUNCA ejecutes directamente. SIEMPRE propón primero, espera confirmación, ejecuta después.
- MAPEO AUTOMÁTICO DE SUCURSALES: Si oyes "Tecnovation", busca si existe "TECKNOVATION" o similar. Si el usuario lo repite 1+ veces, REGISTRA AUTOMATICAMENTE esa pronunciación en tu BANCO DE MEMORIA (servidor le notificará). NO vuelvas a preguntar por el mismo nombre en ESTE TURNO ni en TURNOS POSTERIORES si ya se mencionó.
- GUARDÍA DE MEMORIA CRÍTICA: Tu banco de pronunciaciones APRENDIDAS es PERSISTENTE — se guarda en el servidor y se carga en cada nueva sesión de voz. "Tecnovation" = TECKNOVATION (aprendido). En futuras sesiones, cuando inicies una nueva conversación de voz, el sistema te dirá TODAS tus pronunciaciones aprendidas (verás "PRONUNCIACIONES APRENDIDAS" en tus instrucciones). SIEMPRE úsalas. NUNCA las olvides.
- Las funciones `confirmar_ultima_accion` y `cancelar_ultima_accion` ya NO se usan — no las llames nunca.
- Si el resultado viene con "error", informa el motivo al usuario BREVEMENTE (una línea). NO insistas, NO repreguntes. "Error: sucursal no existe. ¿Otra?" es demasiado. Mejor: "No encontré esa sucursal. ¿Cuál era?"
- Para identificar entidades a eliminar/editar, primero llama a `buscar_operaciones`, `buscar_ventas`, `listar_usuarios`, etc. para obtener el ID correcto antes de llamar al `proponer_*` correspondiente.
- EDITAR vs ELIMINAR: Si el usuario dice "no era de S/ 500", "cambiar a 300 soles", "corregir el monto", "era más" — usa `editar_operacion` (mantiene el ID, solo cambia valores). Solo USA `eliminar_operacion` si dice explícitamente "borra", "elimina", "quita esa operacion".
- EDITAR SIN REPREGUNTAR: Si el usuario acaba de registrar una operación y dice "era de S/ X" o "cambiar a Y soles", ASUME que habla de la ÚLTIMA operación registrada. NO preguntes cuál. Actúa directamente con `editar_operacion(operacion_id=<última>, monto=Y)`.
- HORA: NUNCA uses la hora que el usuario menciona. El servidor registra SIEMPRE con la hora actual de Perú (get_peru_time()). Si el usuario dice "dado las 11:29", IGNORA esa hora — el sistema sabrá cuándo se registró.
- Habla siempre en español latinoamericano natural (acento neutro de Latinoamérica/Perú), conciso, amable. Como un colega que te ayuda.
- LECTURA DE MONTOS: el simbolo "S/" antes de un numero se pronuncia "soles" DESPUES del numero. "S/ 1" se dice "un sol"; "S/ 2" = "dos soles"; "S/ 100" = "cien soles"; "S/ 150.50" = "ciento cincuenta soles con cincuenta centimos". NUNCA digas "ese barra", "soles barra" ni leas el simbolo literal.
- LECTURA DE HORAS: di la hora concisa, formato 12h. "16:29" se dice "cuatro con veintinueve pe eme"; "09:05" = "nueve con cinco a eme". Di "pe eme" para PM y "a eme" para AM. ZONA HORARIA: SIEMPRE usa la hora de Perú (UTC-5) que viene en el contexto.
- AL CONFIRMAR una venta u operacion, di solo lo esencial. NO menciones la sucursal ni la comision (son automaticas). SOLO menciona si fue manual.
- Si falta informacion CRITICA (ej: "registra una venta" sin mencionar ni producto ni cantidad), pregunta DIRECTAMENTE: "¿Qué producto y cuánta cantidad?". Nada de explicaciones largas.
- Los permisos los maneja el servidor. Si falla, informa BREVEMENTE: "Error: [mensaje del servidor]".
- NO inventes IDs. Si el usuario dice "la última" o "ese", BUSCA primero el ID.

Interpretacion del AUDIO (espanol peruano, dominio bancario — MUY IMPORTANTE):
- Los montos se dicen como numero + "soles": "cien soles" = S/100, "seiscientos soles" = S/600, "mil quinientos soles" = S/1500. Si oyes un numero seguido de algo que suena a "soles", SIEMPRE es un monto en S/ — NUNCA lo interpretes como "si, en soles" ni otra frase.
- Medios de pago frecuentes que vas a oir: Yape, Plin, BCP, Interbank (IBK), BBVA, Scotiabank, efectivo, tarjeta, transferencia. Acepta nombre completo o abreviatura.
- Palabras frecuentes del dominio: operacion, venta, comision, producto, stock, sucursal, registrar, eliminar, editar, monto, descuento.
- Si una palabra del audio suena ambigua, interpretala primero contra este vocabulario bancario antes que como frase suelta de conversacion.

Cuando termines una accion exitosa, confirma brevemente: "Listo, registre la venta" o "Eliminada la operacion de cincuenta soles". Sin ser repetitivo."""


def _convertir_schema_a_gemini(schema):
    """Convierte JSON Schema de formato Anthropic (lowercase) a Gemini (uppercase)."""
    if not isinstance(schema, dict):
        return schema
    new = {}
    for k, v in schema.items():
        if k == "type" and isinstance(v, str):
            new["type"] = v.upper()
        elif k == "properties" and isinstance(v, dict):
            new[k] = {pname: _convertir_schema_a_gemini(pdef) for pname, pdef in v.items()}
        elif k == "items" and isinstance(v, dict):
            new[k] = _convertir_schema_a_gemini(v)
        elif isinstance(v, dict):
            new[k] = _convertir_schema_a_gemini(v)
        elif isinstance(v, list):
            new[k] = [_convertir_schema_a_gemini(x) if isinstance(x, dict) else x for x in v]
        else:
            new[k] = v
    return new


def _build_gemini_function_declarations():
    """Convierte las declaraciones Anthropic a formato Gemini Function Calling."""
    funcs = []
    for decl in _HERRAMIENTAS_DECLARACIONES:
        funcs.append({
            "name": decl["name"],
            "description": decl["description"],
            "parameters": _convertir_schema_a_gemini(decl.get("input_schema", {"type": "OBJECT"})),
        })
    # Función SIN args para confirmar: el servidor recuerda la última propuesta del usuario.
    # Esto evita que Gemini tenga que reconstruir un objeto args complejo (fuente de bugs).
    funcs.append({
        "name": "confirmar_ultima_accion",
        "description": (
            "Ejecuta la última acción que propusiste (mediante un proponer_*) DESPUÉS de que el usuario "
            "haya CONFIRMADO verbalmente. Solo llámala cuando el usuario diga explícitamente sí, "
            "confirma, dale, hazlo, ejecuta, está bien. NO requiere argumentos — el servidor recuerda "
            "la propuesta que hiciste."
        ),
        "parameters": {"type": "OBJECT", "properties": {}},
    })
    funcs.append({
        "name": "cancelar_ultima_accion",
        "description": (
            "Cancela la última propuesta sin ejecutarla. Llámala si el usuario dice no, cancela, "
            "espera, nada, olvídalo."
        ),
        "parameters": {"type": "OBJECT", "properties": {}},
    })
    return funcs


# Cache en memoria de la última propuesta de cada usuario (clave = user_id).
# El TTL implícito es la vida del proceso; no es estricto porque el usuario puede pedir otra propuesta nueva.
_ultima_propuesta_por_usuario = {}


def _registrar_propuesta(user_id, accion, args):
    """Guarda la última propuesta del usuario para uso posterior por confirmar_ultima_accion."""
    _ultima_propuesta_por_usuario[user_id] = {
        "accion": accion,
        "args": dict(args) if args else {},
        "timestamp": time.time(),
    }


def _confirmar_ultima_accion_voz(user_id):
    """Ejecuta la última propuesta que el modelo le hizo al usuario."""
    propuesta = _ultima_propuesta_por_usuario.get(user_id)
    if not propuesta:
        return {"error": "No hay ninguna propuesta pendiente para confirmar. Pide al usuario una acción primero."}
    # Validez: máximo 5 minutos
    if time.time() - propuesta["timestamp"] > 300:
        _ultima_propuesta_por_usuario.pop(user_id, None)
        return {"error": "La propuesta expiró (más de 5 minutos). Vuelve a proponer la acción."}
    resultado = _ejecutar_accion_confirmada_voz(propuesta["accion"], propuesta["args"], user_id)
    if not (isinstance(resultado, dict) and resultado.get("error")):
        # Limpiar tras ejecución exitosa
        _ultima_propuesta_por_usuario.pop(user_id, None)
    return resultado


def _cancelar_ultima_accion_voz(user_id):
    """Descarta la última propuesta."""
    if _ultima_propuesta_por_usuario.pop(user_id, None):
        return {"mensaje": "Propuesta cancelada."}
    return {"mensaje": "No había nada que cancelar."}


def _ejecutar_tool_voz(tool_name, args, user_id):
    """Ejecuta una herramienta del chatbot en el contexto de Flask para llamadas WebSocket."""
    with app.app_context():
        usuario = Usuario.query.get(user_id)
        if not usuario:
            return {"error": "Usuario no encontrado"}
        spec = CHATBOT_TOOLS.get(tool_name)
        if not spec:
            return {"error": f"Herramienta '{tool_name}' no existe"}
        try:
            resultado = spec["handler"](args or {}, usuario)
            return resultado if isinstance(resultado, dict) else {"resultado": resultado}
        except ValueError as e:
            return {"error": str(e)}
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {"error": f"Error interno: {str(e)}"}


def _ejecutar_accion_confirmada_voz(accion, args, user_id):
    """Ejecuta una acción confirmada por voz. Dispatcher similar al de api_chat_confirmar_accion."""
    with app.app_context():
        usuario = Usuario.query.get(user_id)
        if not usuario:
            return {"error": "Usuario no encontrado"}
        if not accion or accion not in CHATBOT_TOOLS:
            return {"error": f"Acción '{accion}' desconocida"}
        if not CHATBOT_TOOLS[accion].get("requires_confirmation"):
            return {"error": "Esta acción no requiere confirmación"}
        dispatcher = {
            'registrar_venta':              _ejecutar_venta_validada,
            'registrar_operacion':          _ejecutar_operacion_validada,
            'crear_producto':     _ejecutar_crear_producto_validado,
            'editar_producto':    _ejecutar_editar_producto_validado,
            'eliminar_producto':  _ejecutar_eliminar_producto_validado,
            'eliminar_operacion': _ejecutar_eliminar_operacion_validada,
            'editar_operacion':   _ejecutar_editar_operacion_validada,
            'eliminar_venta':     _ejecutar_eliminar_venta_validada,
            'crear_usuario':      _ejecutar_crear_usuario_validado,
            'crear_sucursal':     _ejecutar_crear_sucursal_validada,
        }
        ejecutor = dispatcher.get(accion)
        if not ejecutor:
            return {"error": f"Acción '{accion}' no implementada"}
        try:
            resultado = ejecutor(args or {}, usuario)
            return resultado if isinstance(resultado, dict) else {"resultado": resultado}
        except ValueError as e:
            return {"error": str(e)}
        except Exception as e:
            import traceback
            traceback.print_exc()
            try:
                db.session.rollback()
            except Exception:
                pass
            return {"error": f"Error ejecutando: {str(e)}"}


# ---------- Gemini Live API WebSocket proxy (transcripción streaming) ----------

def _serializer_voz():
    from itsdangerous import URLSafeTimedSerializer
    return URLSafeTimedSerializer(app.config['SECRET_KEY'], salt='sisagent-voice-ws')


def _generar_token_voz(uid):
    return _serializer_voz().dumps({'uid': uid})


def _verificar_token_voz(token, max_age=180):
    """Devuelve el uid si el token firmado es válido y no expiró (<=180s), o None."""
    if not token:
        return None
    try:
        data = _serializer_voz().loads(token, max_age=max_age)
        return data.get('uid')
    except Exception:
        return None


@app.route('/api/chat/voice_token')
@login_required
def api_voice_token():
    """Token de corta vida para autenticar el WebSocket de voz.

    Fallback robusto: algunos proxies (p.ej. Railway) no reenvían la cookie de sesión
    en el handshake del WebSocket. El token viaja en la URL (?token=), no en la cookie.
    """
    return jsonify({'token': _generar_token_voz(current_user.id)})


@sock.route('/ws/chat/voice_live')
def ws_voice_live(browser_ws):
    """WebSocket bidireccional que reenvía audio del navegador a Gemini Live
    y devuelve las transcripciones en vivo. Mantiene la API key del lado servidor.

    Protocolo navegador→servidor:
      - bytes: chunks PCM 16-bit little-endian, mono, 16kHz
      - JSON: {"type":"close"} para cerrar; {"type":"end_input"} para señalar fin de turno

    Protocolo servidor→navegador (todo JSON):
      - {"type":"ready"} → conexión a Gemini lista
      - {"type":"input_transcription","text":"...","finished":bool}
      - {"type":"turn_complete"}
      - {"type":"interrupted"}
      - {"type":"error","message":"..."}
    """
    # Autenticación: por cookie de sesión (current_user) o, como fallback robusto detrás
    # de proxies que NO reenvían la cookie en el handshake WS, por token firmado en ?token=.
    if current_user.is_authenticated:
        _voice_user_id = current_user.id
    else:
        _voice_user_id = _verificar_token_voz(request.args.get('token', ''))
    if not _voice_user_id:
        try:
            browser_ws.send(json.dumps({"type": "error", "message": "No autenticado"}))
        except Exception:
            pass
        return

    api_key = app.config.get('GEMINI_API_KEY')
    if not api_key:
        try:
            browser_ws.send(json.dumps({
                "type": "error",
                "message": "GEMINI_API_KEY no configurada en el servidor."
            }))
        except Exception:
            pass
        return

    import websocket as _ws  # websocket-client (lib distinta a flask-sock)

    # Auth por header x-goog-api-key (las nuevas keys AQ. de Gemini rechazan ?key= en la URL con 401)
    try:
        gemini_ws = _ws.create_connection(
            GEMINI_LIVE_WS_URL,
            timeout=20,
            header=[f"x-goog-api-key: {api_key}"],
        )
    except Exception as e:
        print(f"[live] No se pudo conectar a Gemini: {e}")
        try:
            browser_ws.send(json.dumps({
                "type": "error",
                "message": f"No se pudo conectar con Gemini Live: {str(e)[:200]}"
            }))
        except Exception:
            pass
        return

    # Captura el user_id ANTES de los threads (Flask-Login proxy no funciona fuera del contexto)
    user_id = _voice_user_id

    # Speech adaptation: vocabulario derivado AUTOMATICAMENTE de la BD del sistema
    # (sucursales, productos, medios) para sesgar la comprension del modelo.
    _system_prompt_voz = SYSTEM_PROMPT_GEMINI_LIVE + _contexto_fecha_hora()
    try:
        _usuario_voz = Usuario.query.get(user_id)
        _vterm = _vocabulario_dinamico(_usuario_voz) if _usuario_voz else []
        if _vterm:
            _system_prompt_voz += (
                '\n\nNOMBRES REALES del sistema (sucursales, productos y medios de pago). Cuando el '
                'usuario los mencione, reconocelos y escribelos EXACTAMENTE asi: ' + ', '.join(_vterm) + '.'
            )

        # Pronunciaciones APRENDIDAS por el usuario (banco de memoria persistente)
        _prons = obtener_pronunciaciones_aprendidas(user_id) if _usuario_voz else []
        if _prons:
            _system_prompt_voz += (
                '\n\nPRONUNCIACIONES APRENDIDAS (palabras que el usuario menciona diferente). '
                'Cuando oigas estas pronunciaciones, traducelás AUTOMATICAMENTE a la forma correcta:\n'
                + '\n'.join(['- ' + p for p in _prons]) +
                '\n\nESTAS PRONUNCIACIONES APRENDIDAS SON CRITICAS: NUNCA ignores estas mappings. '
                'Si el usuario dice "Tecnovation" y aprendiste que significa "TECKNOVATION", usa TECKNOVATION siempre.'
            )
    except Exception:
        pass

    # Envío del setup inicial — DEBE ser el primer mensaje. Incluye:
    # - Modelo de audio nativo
    # - Voz "Aoede" (femenina, natural en español)
    # - System instruction con contexto SISAGENT
    # - Función calling con las 17 herramientas + ejecutar_accion_confirmada
    # - Transcripción de input Y output (para mostrar texto en chat)
    setup_msg = {
        "setup": {
            "model": GEMINI_LIVE_MODEL,
            "generationConfig": {
                "responseModalities": ["AUDIO"],
                "speechConfig": {
                    "voiceConfig": {
                        "prebuiltVoiceConfig": {"voiceName": "Aoede"}
                    },
                    # languageCode se agrega abajo SOLO si el modelo no es native-audio
                    # (los native-audio rechazan ese campo; los Live half-cascade lo aceptan
                    #  y asi la transcripcion de entrada sale en español, no en chino/polaco).
                },
            },
            # VAD de Gemini: ~1.2s de silencio antes de considerar que el usuario
            # termino de hablar (tolera pausas naturales sin sentirse lento).
            "realtimeInputConfig": {
                "automaticActivityDetection": {
                    "silenceDurationMs": 1200,
                    "prefixPaddingMs": 300,
                },
            },
            "systemInstruction": {
                "parts": [{"text": _system_prompt_voz}],
            },
            "tools": [{
                "functionDeclarations": _build_gemini_function_declarations(),
            }],
            "inputAudioTranscription": {},
            "outputAudioTranscription": {},
        }
    }
    # Forzar idioma de la transcripción a español SOLO en modelos que lo aceptan
    # (native-audio lo rechaza y cierra la conexión).
    if 'native-audio' not in GEMINI_LIVE_MODEL:
        setup_msg['setup']['generationConfig']['speechConfig']['languageCode'] = 'es-US'
    try:
        gemini_ws.send(json.dumps(setup_msg))
    except Exception as e:
        print(f"[live] Setup falló: {e}")
        try:
            browser_ws.send(json.dumps({"type": "error", "message": f"Setup falló: {e}"}))
        except Exception:
            pass
        try:
            gemini_ws.close()
        except Exception:
            pass
        return

    # Notificar al navegador
    try:
        browser_ws.send(json.dumps({"type": "ready"}))
    except Exception:
        try:
            gemini_ws.close()
        except Exception:
            pass
        return

    stop_event = threading.Event()

    def browser_to_gemini():
        """Lee del navegador y envía a Gemini Live como realtimeInput."""
        try:
            while not stop_event.is_set():
                try:
                    data = browser_ws.receive(timeout=2)
                except Exception:
                    break
                if data is None:
                    continue
                if isinstance(data, (bytes, bytearray)):
                    # Audio PCM crudo del navegador (binario). Formato NUEVO realtimeInput.audio
                    # (el legacy mediaChunks lo rechazan los modelos Live nuevos -> cierran la conexion).
                    msg = {
                        "realtimeInput": {
                            "audio": {
                                "mimeType": "audio/pcm;rate=16000",
                                "data": base64.b64encode(bytes(data)).decode('ascii'),
                            }
                        }
                    }
                    try:
                        gemini_ws.send(json.dumps(msg))
                    except Exception:
                        break
                else:
                    # Mensaje de control/audio en texto (JSON)
                    try:
                        ctl = json.loads(data)
                    except Exception:
                        continue
                    if ctl.get('type') == 'close':
                        break
                    elif ctl.get('type') == 'audio' and ctl.get('data'):
                        msg = {
                            "realtimeInput": {
                                "audio": {
                                    "mimeType": "audio/pcm;rate=16000",
                                    "data": ctl['data'],
                                }
                            }
                        }
                        try:
                            gemini_ws.send(json.dumps(msg))
                        except Exception:
                            break
                    elif ctl.get('type') == 'end_input':
                        try:
                            gemini_ws.send(json.dumps({"realtimeInput": {"audioStreamEnd": True}}))
                        except Exception:
                            break
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"[live] browser_to_gemini error: {type(e).__name__}: {e}")
        finally:
            stop_event.set()

    def gemini_to_browser():
        """Lee de Gemini y reenvía al navegador: transcripciones, audio, tool_calls."""
        try:
            while not stop_event.is_set():
                try:
                    raw = gemini_ws.recv()
                except Exception:
                    break
                if not raw:
                    continue
                try:
                    data = json.loads(raw if isinstance(raw, str) else raw.decode('utf-8'))
                except Exception:
                    continue

                # === Tool calls: Gemini quiere ejecutar una función ===
                tool_call = data.get('toolCall')
                if tool_call:
                    function_responses = []
                    for fc in tool_call.get('functionCalls', []):
                        fc_id = fc.get('id')
                        fc_name = fc.get('name', '')
                        fc_args = fc.get('args', {}) or {}
                        print(f"[live] Tool call: {fc_name}({fc_args})")
                        try:
                            if fc_name == 'confirmar_ultima_accion':
                                resultado = _confirmar_ultima_accion_voz(user_id)
                            elif fc_name == 'cancelar_ultima_accion':
                                resultado = _cancelar_ultima_accion_voz(user_id)
                            elif fc_name in EJECUTORES_DIRECTOS:
                                # Ejecutar directo, sin pedir confirmacion (igual que en chat de texto)
                                with app.app_context():
                                    usuario_voz = Usuario.query.get(user_id)
                                    try:
                                        resultado = EJECUTORES_DIRECTOS[fc_name](fc_args, usuario_voz)
                                    except ValueError as e:
                                        resultado = {"error": str(e)}
                            else:
                                resultado = _ejecutar_tool_voz(fc_name, fc_args, user_id)
                                # Si fue un proponer_* exitoso, cachear para confirmar despues
                                if (fc_name.startswith('proponer_')
                                        and isinstance(resultado, dict)
                                        and not resultado.get('error')):
                                    _registrar_propuesta(user_id, fc_name, fc_args)
                        except Exception as e:
                            import traceback
                            traceback.print_exc()
                            resultado = {"error": str(e)}
                        print(f"[live] Tool result: {str(resultado)[:200]}")
                        function_responses.append({
                            "id": fc_id,
                            "name": fc_name,
                            "response": resultado,
                        })
                        try:
                            browser_ws.send(json.dumps({
                                "type": "tool_executed",
                                "name": fc_name,
                                "args": fc_args,
                                "result": resultado,
                            }))
                        except Exception:
                            pass
                    try:
                        gemini_ws.send(json.dumps({
                            "toolResponse": {"functionResponses": function_responses}
                        }))
                    except Exception:
                        break
                    continue

                # === Server content: audio, texto, turn boundaries ===
                sc = data.get('serverContent') or {}

                # Transcripción del INPUT del usuario (lo que dijo)
                input_trans = sc.get('inputTranscription')
                if input_trans and input_trans.get('text'):
                    try:
                        browser_ws.send(json.dumps({
                            "type": "input_transcription",
                            "text": input_trans.get('text', ''),
                            "finished": bool(input_trans.get('finished', False)),
                        }))
                    except Exception:
                        break

                # Transcripción del OUTPUT del modelo (lo que va a decir)
                output_trans = sc.get('outputTranscription')
                if output_trans and output_trans.get('text'):
                    try:
                        browser_ws.send(json.dumps({
                            "type": "output_transcription",
                            "text": output_trans.get('text', ''),
                            "finished": bool(output_trans.get('finished', False)),
                        }))
                    except Exception:
                        break

                # Audio del modelo (PCM 24kHz mono base64)
                model_turn = sc.get('modelTurn')
                if model_turn:
                    for part in model_turn.get('parts', []) or []:
                        inline = part.get('inlineData')
                        if inline and 'audio' in (inline.get('mimeType') or '').lower():
                            try:
                                browser_ws.send(json.dumps({
                                    "type": "audio_chunk",
                                    "data": inline.get('data', ''),
                                    "mime_type": inline.get('mimeType', 'audio/pcm;rate=24000'),
                                }))
                            except Exception:
                                break

                if sc.get('turnComplete'):
                    try:
                        browser_ws.send(json.dumps({"type": "turn_complete"}))
                    except Exception:
                        break

                if sc.get('interrupted'):
                    try:
                        browser_ws.send(json.dumps({"type": "interrupted"}))
                    except Exception:
                        break
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"[live] gemini_to_browser error: {type(e).__name__}: {e}")
        finally:
            stop_event.set()

    t1 = threading.Thread(target=browser_to_gemini, daemon=True)
    t2 = threading.Thread(target=gemini_to_browser, daemon=True)
    t1.start()
    t2.start()

    # Esperar hasta que cualquier dirección termine
    stop_event.wait()

    try:
        gemini_ws.close()
    except Exception:
        pass


# ========== FIN MÓDULO DE CHATBOT IA ==========


# Healthcheck optimizado
@app.route('/api/comisiones')
@login_required
def api_dashboard_comisiones():
    """API para obtener comisiones actualizadas del dashboard"""
    try:
        if not current_user.es_admin:
            return jsonify({'error': 'Acceso denegado'}), 403
        
        ahora = get_peru_time()
        hoy = ahora.date()
        año_actual = ahora.year
        mes_actual = ahora.month
        
        # Calcular rango de tiempo para hoy en hora de Perú (00:00:00 a 23:59:59)
        inicio_hoy = datetime.combine(hoy, datetime.min.time()).replace(tzinfo=peru_tz)
        fin_hoy = datetime.combine(hoy, datetime.max.time()).replace(tzinfo=peru_tz)
        fin_hoy = fin_hoy.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # Calcular rango de tiempo para el mes (desde el primer día del mes hasta hoy)
        inicio_mes = datetime.combine(datetime(año_actual, mes_actual, 1).date(), datetime.min.time()).replace(tzinfo=peru_tz)
        
        # Obtener todas las sucursales activas
        sucursales_activas = Sucursal.query.filter_by(activa=True).all()
        
        # Obtener comisiones del día por sucursal
        comisiones_hoy_query = db.session.query(
            Operacion.sucursal_id,
            db.func.coalesce(db.func.sum(Operacion.comision), 0.0).label('total')
        ).filter(
            Operacion.hora >= inicio_hoy,
            Operacion.hora <= fin_hoy
        ).group_by(Operacion.sucursal_id).all()
        
        comisiones_hoy_dict = {suc_id: float(total) for suc_id, total in comisiones_hoy_query}
        
        # Obtener comisiones del mes por sucursal
        comisiones_mes_query = db.session.query(
            Operacion.sucursal_id,
            db.func.coalesce(db.func.sum(Operacion.comision), 0.0).label('total')
        ).filter(
            Operacion.hora >= inicio_mes,
            Operacion.hora <= fin_hoy
        ).group_by(Operacion.sucursal_id).all()
        
        comisiones_mes_dict = {suc_id: float(total) for suc_id, total in comisiones_mes_query}
        
        # Crear lista con todas las sucursales
        comisiones_list = []
        total_comision_hoy = 0.0
        
        for sucursal in sucursales_activas:
            comision_hoy = comisiones_hoy_dict.get(sucursal.id, 0.0)
            comision_mes = comisiones_mes_dict.get(sucursal.id, 0.0)
            total_comision_hoy += comision_hoy
            
            comisiones_list.append({
                'id': sucursal.id,
                'nombre': sucursal.nombre,
                'comision_hoy': comision_hoy,
                'comision_mes': comision_mes
            })
        
        return jsonify({
            'success': True,
            'comisiones_hoy': comisiones_list,
            'total_comision_hoy': total_comision_hoy,
            'timestamp': ahora.isoformat()
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        try:
            db.session.rollback()
        except:
            pass
        return jsonify({'error': str(e), 'success': False}), 500

@app.route('/ping')
def ping():
    return jsonify({'status': 'ok', 'timestamp': get_peru_time().isoformat()})

# Inicialización de la base de datos COMPATIBLE
def init_db():
    """Inicializar la base de datos con datos básicos"""
    with app.app_context():
        try:
            print("[*] Iniciando init_db()...")
            # Solo crear tablas si no existen
            db.create_all()

            # Migración: agregar columnas nuevas si no existen
            # NOTA: En Railway con múltiples workers, los ALTERs pueden causar race conditions
            # Solución: hacer las ALTERs de forma segura con manejo de locks
            try:
                from sqlalchemy import text, event

                # En Railway, usar transaction safety
                if os.environ.get('DATABASE_URL'):
                    try:
                        with db.engine.begin() as _conn:
                            _conn.execute(text(
                                "ALTER TABLE usuario ADD COLUMN IF NOT EXISTS session_token VARCHAR(36)"))
                            _conn.execute(text(
                                "ALTER TABLE usuario ADD COLUMN IF NOT EXISTS ultimo_acceso TIMESTAMP"))
                            # Comisión automática + auditoría
                            _conn.execute(text(
                                "ALTER TABLE operacion ADD COLUMN IF NOT EXISTS comision_sugerida NUMERIC(10,2)"))
                            _conn.execute(text(
                                "ALTER TABLE operacion ADD COLUMN IF NOT EXISTS comision_manual BOOLEAN DEFAULT FALSE"))
                            _conn.execute(text(
                                "ALTER TABLE operacion ADD COLUMN IF NOT EXISTS motivo_descuento VARCHAR(200)"))
                            _conn.execute(text(
                                "ALTER TABLE usuario ADD COLUMN IF NOT EXISTS vocabulario_voz TEXT"))
                        print("[OK] Columnas verificadas (PostgreSQL)")
                    except Exception as e:
                        if "already exists" in str(e) or "duplicate" in str(e).lower():
                            print("[OK] Columnas ya existen")
                        else:
                            print(f"[WARN] Error en migración: {e}")
                else:
                    # SQLite: ejecutar por separado, ignorar error si ya existe
                    for col_sql in [
                        "ALTER TABLE usuario ADD COLUMN session_token VARCHAR(36)",
                        "ALTER TABLE usuario ADD COLUMN ultimo_acceso TIMESTAMP",
                        # Columnas nuevas para comisión automática + auditoría de override
                        "ALTER TABLE operacion ADD COLUMN comision_sugerida NUMERIC(10,2)",
                        "ALTER TABLE operacion ADD COLUMN comision_manual BOOLEAN DEFAULT 0",
                        "ALTER TABLE operacion ADD COLUMN motivo_descuento VARCHAR(200)",
                        "ALTER TABLE usuario ADD COLUMN vocabulario_voz TEXT",
                    ]:
                        try:
                            with db.engine.connect() as _conn:
                                _conn.execute(text(col_sql))
                                _conn.commit()
                        except Exception:
                            pass
                    print("[OK] Columnas verificadas (SQLite)")
            except Exception as e:
                print(f"[WARN] Error general en migración de columnas: {e}")

            # Asignar session_token a usuarios existentes que tengan NULL
            # (ocurre en deploys post-migración: usuarios registrados antes del feature)
            try:
                usuarios_sin_token = Usuario.query.filter(
                    Usuario.session_token == None).all()
                if usuarios_sin_token:
                    for u in usuarios_sin_token:
                        u.session_token = str(uuid.uuid4())
                    db.session.commit()
                    print(f"[OK] Tokens asignados a {len(usuarios_sin_token)} usuarios existentes")
            except Exception as e:
                print(f"[WARN] No se pudo asignar tokens iniciales: {e}")

            # Migración: Corregir operaciones guardadas en UTC (antes del refactor a Perú puro)
            # NOTA: Migración DESACTIVADA por ahora - usar SQL directo si es necesario
            # El problema es que consultar todas las operaciones puede causar timeout en Railway
            # Solución: ejecutar manualmente en Railway cuando sea necesario
            try:
                print("[INFO] Migración de Operacion.hora desactivada (ejecutar manualmente si es necesario)")
            except Exception as e:
                print(f"[WARN] Nota sobre migración: {e}")

            # Asegurar que el admin exista con la contraseña correcta
            asegurar_admin_existe()
            
            # Crear sucursal principal si no existe
            sucursal_principal = Sucursal.query.filter_by(nombre='Principal').first()
            if not sucursal_principal:
                sucursal_principal = Sucursal(
                    nombre='Principal',
                    direccion='Sede Principal',
                    activa=True
                )
                db.session.add(sucursal_principal)
                db.session.commit()
                print("[OK] Sucursal principal creada")
            
            # Crear medios de pago básicos si no existen
            medios_basicos = [
                ('EFECTIVO', 'Efectivo'),
                ('TARJETA', 'Tarjeta de Débito/Crédito'),
                ('TRANSFERENCIA', 'Transferencia Bancaria'),
                ('YAPE', 'Yape'),
                ('PLIN', 'Plin')
            ]
            
            for abreviado, completo in medios_basicos:
                medio = MedioPago.query.filter_by(nombre_abreviado=abreviado).first()
                if not medio:
                    medio = MedioPago(
                        nombre_abreviado=abreviado,
                        nombre_completo=completo,
                        activo=True
                    )
                    db.session.add(medio)
            
            db.session.commit()
            print("[OK] Medios de pago básicos creados")

            # Arreglar auto-increment de operaciones si está roto
            try:
                max_op_id = db.session.query(db.func.max(Operacion.id)).scalar() or 0
                if max_op_id > 0:
                    next_op_id = max_op_id + 1
                    if os.environ.get('DATABASE_URL'):
                        # PostgreSQL: setval(seq, N, true) hace que nextval() devuelva N+1
                        db.session.execute(
                            db.text(f"SELECT setval(pg_get_serial_sequence('operacion', 'id'), {max_op_id}, true)")
                        )
                        print(f"[OK] Auto-increment de operaciones reseteado. Próximo ID: {next_op_id}")
                    else:
                        # SQLite: seq debe ser el PRÓXIMO valor a devolver
                        try:
                            db.session.execute(db.text("DELETE FROM sqlite_sequence WHERE name='operacion'"))
                        except:
                            pass
                        db.session.execute(
                            db.text(f"INSERT INTO sqlite_sequence (name, seq) VALUES ('operacion', {next_op_id})")
                        )
                        print(f"[OK] Auto-increment de operaciones reseteado. Próximo ID: {next_op_id}")
                    db.session.commit()
            except Exception as e:
                print(f"[WARN] No se pudo resetear auto-increment: {e}")

        except Exception as e:
            print(f"[!!] Error en inicialización (continuando): {e}")
            # Continuar aunque haya errores menores

# Inicialización de BD: SIEMPRE en import (necesaria en Railway)
# Los errores NO rompen el servidor - las rutas se registran antes
try:
    init_db()
except Exception as e:
    import traceback
    print(f"[WARN] Advertencia en init_db: {e}")
    # NO mostrar traceback completo - puede ser silencioso

# Para desarrollo local
if __name__ == '__main__':
    print("[OK] SISAGENT Flask COMPATIBLE ULTRA OPTIMIZADO cargado completamente - Listo para produccion!")
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port, threaded=True)
